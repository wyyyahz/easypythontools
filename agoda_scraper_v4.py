#!/usr/bin/env python3
"""
Agoda 武汉酒店数据抓取 v4 - 首页搜索 + JS提取
"""
import json, os, time, sqlite3, re
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

BASE = os.path.dirname(os.path.abspath(__file__))
CITY = "武汉"
CHECKIN = (datetime.now() + timedelta(days=35)).strftime("%Y-%m-%d")
CHECKOUT = (datetime.now() + timedelta(days=36)).strftime("%Y-%m-%d")
DB_PATH = os.path.join(BASE, "hotels.db")
EXCEL_PATH = os.path.join(BASE, f"hotels_{CITY}.xlsx")

EXTRACT_JS = """
return (() => {
    const cards = document.querySelectorAll('[class*=PropertyCard]');
    const results = [];
    cards.forEach(card => {
        try {
            const h3 = card.querySelector('h3');
            const name = h3 ? h3.textContent.trim() : '';
            if (!name) return;

            let rating = null;
            card.querySelectorAll('[class*=review] [class*=score], [class*=Score]').forEach(el => {
                const m = el.textContent.trim().match(/(\d+\\.?\\d*)/);
                if (m) { const v = parseFloat(m[1]); if (v > 0 && v <= 10) rating = v; }
            });

            let reviewCount = 0;
            card.querySelectorAll('span').forEach(el => {
                const m = el.textContent.trim().match(/(\\d+)\\s*[条则篇]/);
                if (m) reviewCount = parseInt(m[1]);
            });

            let price = null;
            card.querySelectorAll('*').forEach(el => {
                const m = el.textContent.trim().match(/[¥￥]\\s*([\\d,]+\\.?\\d*)/);
                if (m) {
                    const v = parseFloat(m[1].replace(/,/g, ''));
                    if (v >= 10 && v <= 100000) price = v;
                }
            });

            let stars = 0;
            const starEls = card.querySelectorAll('[class*=star]:not([class*=outline]):not([class*=empty])');
            if (starEls.length > 0) stars = starEls.length;

            let url = '';
            const link = card.querySelector('a[href*=\"/hotel\"], a[href*=\"/hotels\"]');
            if (link) url = link.href;

            results.push({ name, rating, reviewCount, price, stars, url });
        } catch(e) {}
    });
    return JSON.stringify(results);
})();
"""

def setup_driver():
    opts = Options()
    opts.binary_location = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--lang=zh-CN")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    driver = webdriver.Chrome(
        service=Service(os.path.join(BASE, "chromedriver.exe")),
        options=opts)
    return driver

def search_wuhan(driver):
    """Search for Wuhan hotels via homepage search."""
    print(">> 打开 Agoda 首页...")
    driver.get("https://www.agoda.cn/")
    time.sleep(4)

    # Try to dismiss any popups
    for script in [
        "document.querySelectorAll('[class*=close],[class*=dismiss]').forEach(e=>e.click())",
        "document.querySelectorAll('[aria-label*=Close]').forEach(e=>e.click())",
    ]:
        try: driver.execute_script(script)
        except: pass
    time.sleep(1)

    # Find search input and type city
    print(">> 搜索城市: 武汉")
    search_input = None
    for xp in [
        "//input[@placeholder]",
        "//input[@aria-label]",
        "//input[@data-element-name='search-box']",
        "//div[contains(@class,'SearchBox')]//input",
        "//input[contains(@class,'search')]",
        "//input",
    ]:
        try:
            els = driver.find_elements(By.XPATH, xp)
            for el in els:
                if el.is_displayed() and el.tag_name == "input":
                    placeholder = el.get_attribute("placeholder") or ""
                    if "目的" in placeholder or "城市" in placeholder or "酒店" in placeholder or "搜索" in placeholder:
                        search_input = el
                        print(f"  找到搜索框: {placeholder}")
                        break
            if search_input: break
        except: pass

    if not search_input:
        # Fallback: find any visible input
        for el in driver.find_elements(By.TAG_NAME, "input"):
            try:
                if el.is_displayed():
                    search_input = el
                    print("  使用第一个可见输入框")
                    break
            except: pass

    if not search_input:
        print("  [错误] 未找到搜索框，打开直接搜索 URL")
        driver.get(f"https://www.agoda.cn/search?city=16958&checkIn={CHECKIN}&checkOut={CHECKOUT}&rooms=1&adults=2&currency=CNY")
        time.sleep(10)
        return

    search_input.click()
    time.sleep(0.5)
    search_input.clear()
    time.sleep(0.3)
    search_input.send_keys(CITY)
    time.sleep(3)

    # Select city suggestion
    print(">> 选择城市建议...")
    clicked = False
    for xp in [
        "//*[contains(text(),'武汉')]",
        "//*[@role='option' and contains(.,'武汉')]",
        "//li[contains(.,'武汉')]",
    ]:
        try:
            els = driver.find_elements(By.XPATH, xp)
            for el in els:
                try:
                    if el.is_displayed():
                        driver.execute_script("arguments[0].click();", el)
                        clicked = True
                        print(f"  已选择城市建议")
                        time.sleep(2)
                        break
                except: pass
            if clicked: break
        except: pass

    if not clicked:
        print("  [警告] 未找到城市建议，按 Enter")
        search_input.send_keys(Keys.RETURN)
        time.sleep(3)

    # Set dates via URL navigation (more reliable than date picker)
    print(">> 设置日期...")
    try:
        current_url = driver.current_url
        # Check if URL already has search params
        if "search" in current_url and "city" in current_url:
            # Modify the URL
            import re
            new_url = re.sub(r'checkIn=[^&]*', f'checkIn={CHECKIN}', current_url)
            new_url = re.sub(r'checkOut=[^&]*', f'checkOut={CHECKOUT}', new_url)
            if 'checkIn=' not in new_url:
                sep = '&' if '?' in new_url else '?'
                new_url += f"{sep}checkIn={CHECKIN}&checkOut={CHECKOUT}&rooms=1&adults=2&currency=CNY"
            driver.get(new_url)
            time.sleep(5)
        else:
            raise Exception("Not on search page")
    except:
        print("  >> 直接打开搜索 URL...")
        driver.get(f"https://www.agoda.cn/search?city=16958&checkIn={CHECKIN}&checkOut={CHECKOUT}&rooms=1&adults=2&currency=CNY")
        time.sleep(5)

def scroll_and_extract(driver):
    """Scroll page and extract hotel data."""
    # Wait for PropertyCards
    print(">> 等待搜索加载...")
    try:
        WebDriverWait(driver, 25).until(
            EC.presence_of_all_elements_located((By.XPATH, "//*[contains(@class,'PropertyCard')]"))
        )
        print(">> 酒店卡片已加载")
    except:
        print(">> 等待超时")

    # Additional wait for real search results to replace defaults
    time.sleep(5)

    # Scroll to load more
    print(">> 滚动加载...")
    for i in range(8):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        cards = driver.find_elements(By.XPATH, "//*[contains(@class,'PropertyCard')]")
        print(f"   滚动 {i+1}: {len(cards)} 个卡片")
        # Check if count is stable (no more loading)
        if i >= 3:
            # Try clicking "show more"
            try:
                btns = driver.find_elements(By.XPATH, "//*[contains(text(),'显示更多') or contains(text(),'Show more')]")
                for btn in btns:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        time.sleep(2)
            except: pass

    # Extract via JS
    print(">> JS提取数据...")
    raw = driver.execute_script(EXTRACT_JS)
    items = json.loads(raw) if isinstance(raw, str) else (raw or [])

    # Check names for debugging
    names = [it.get("name","") for it in items[:10]]
    has_chinese = any(any('一' <= c <= '鿿' for c in n) for n in names)
    print(f"  前10个酒店: {names}")
    print(f"  包含中文: {has_chinese}")

    # Filter for Chinese-named hotels
    hotels = []
    for it in items:
        name = it.get("name", "")
        has_chinese = any('一' <= c <= '鿿' for c in name)
        if name and (has_chinese or len(name) > 3):
            hotels.append(it)

    if not hotels and items:
        print(">> 无中文名酒店，使用全部结果")
        hotels = [it for it in items if it.get('name')]

    return hotels

def save_results(hotels):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    print(f"\n>> 保存 {len(hotels)} 家酒店...")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT, user_rating REAL, review_count INTEGER,
        min_price REAL, stars INTEGER, url TEXT, scrape_date TEXT, city TEXT
    )""")
    c.execute("DELETE FROM hotels WHERE city=?", (CITY,))
    for h in hotels:
        c.execute("""INSERT INTO hotels
            (hotel_name, user_rating, review_count, min_price, stars, url, scrape_date, city)
            VALUES (?,?,?,?,?,?,?,?)""",
            (h.get("name",""), h.get("rating"), h.get("reviewCount",0),
             h.get("price"), h.get("stars",0), h.get("url",""),
             datetime.now().strftime("%Y-%m-%d"), CITY))
    conn.commit()
    conn.close()
    print(f"  DB: {DB_PATH} ({len(hotels)}条)")

    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ["序号", "酒店名称", "用户评分", "评价数", "最低价(CNY)", "星级", "详情链接", "抓取日期"]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font, cell.fill, cell.alignment = hfont, hfill, halign

    sorted_h = sorted(hotels, key=lambda x: (x.get("price") or 99999))
    seen = set()
    uniq = []
    for h in sorted_h:
        if h["name"] not in seen:
            seen.add(h["name"])
            uniq.append(h)

    for idx, h in enumerate(uniq, 1):
        r = idx + 1
        ws.cell(row=r, column=1, value=idx).border = border
        ws.cell(row=r, column=2, value=h["name"]).border = border
        sc = ws.cell(row=r, column=3, value=h.get("rating"))
        sc.border = border
        if h.get("rating") and h["rating"] >= 9: sc.font = Font(color="FF0000", bold=True)
        ws.cell(row=r, column=4, value=h.get("reviewCount", 0)).border = border
        ws.cell(row=r, column=5, value=h.get("price")).border = border
        ws.cell(row=r, column=6, value=h.get("stars", 0)).border = border
        ws.cell(row=r, column=7, value=h.get("url", "")).border = border
        ws.cell(row=r, column=8, value=datetime.now().strftime("%Y-%m-%d")).border = border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 12
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    print(f"  Excel: {EXCEL_PATH}")

    prices = [h["price"] for h in hotels if h.get("price")]
    ratings = [h["rating"] for h in hotels if h.get("rating")]
    if prices: print(f"  价格: ¥{min(prices):.0f} ~ ¥{max(prices):.0f}, 均价 ¥{sum(prices)/len(prices):.0f}")
    if ratings: print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f}")

def main():
    print("=" * 55)
    print(f"  Agoda {CITY} 酒店抓取 v4 (首页搜索)")
    print(f"  {CHECKIN} ~ {CHECKOUT}")
    print("=" * 55)

    driver = None
    try:
        driver = setup_driver()
        search_wuhan(driver)

        # Wait for results to fully load
        print(">> 等待搜索完成...")
        time.sleep(8)

        hotels = scroll_and_extract(driver)

        if not hotels:
            print("[错误] 未找到任何酒店")
            with open("page_debug2.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(">> 页面已保存到 page_debug2.html")
            return

        print(f"\n>> 共 {len(hotels)} 家酒店")
        save_results(hotels)

    except Exception as e:
        print(f"[错误] {e}")
        import traceback; traceback.print_exc()
    finally:
        if driver:
            driver.quit()
        print(">> 完成!")

if __name__ == "__main__":
    main()
