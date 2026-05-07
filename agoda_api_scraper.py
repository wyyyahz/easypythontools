#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店数据抓取 - Selenium 获取 Cookie + GraphQL API 直连
"""
import json, os, time, sqlite3, sys, re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

BASE = os.path.dirname(os.path.abspath(__file__))
CITY = "武汉"
CHECKIN_DAYS = 35
CHECKIN = (datetime.now() + timedelta(days=CHECKIN_DAYS)).strftime("%Y-%m-%d")
CHECKOUT = (datetime.now() + timedelta(days=CHECKIN_DAYS + 1)).strftime("%Y-%m-%d")

DB_PATH = os.path.join(BASE, "hotels.db")
EXCEL_PATH = os.path.join(BASE, f"hotels_{CITY}.xlsx")

GRAPHQL_URL = "https://www.agoda.cn/graphql/search"

def get_fresh_cookies():
    """Open Agoda in Selenium to get fresh cookies, then close."""
    print(">> 启动浏览器获取 Cookie...")
    opts = Options()
    opts.binary_location = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1280,800")
    opts.add_argument("--lang=zh-CN")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/131.0.0.0 Safari/537.36")

    chromedriver = os.path.join(BASE, "chromedriver.exe")
    service = Service(chromedriver)
    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    try:
        driver.get("https://www.agoda.cn/")
        time.sleep(5)

        # Dismiss popups
        for script in [
            "document.querySelectorAll('[class*=close],[class*=dismiss]').forEach(e=>e.click())",
            "document.querySelectorAll('button:contains(关闭),button:contains(不)').forEach(e=>e.click())"
        ]:
            try: driver.execute_script(script)
            except: pass
        time.sleep(2)

        # Navigate to search page to get city-specific cookies
        search_url = (f"https://www.agoda.cn/search?city=16958"
                      f"&checkIn={CHECKIN}&checkOut={CHECKOUT}"
                      f"&rooms=1&adults=2&currency=CNY")
        driver.get(search_url)
        time.sleep(5)

        cookies = {c['name']: c['value'] for c in driver.get_cookies()}
        print(f">> 获取到 {len(cookies)} 个 Cookie")
        return cookies
    finally:
        driver.quit()

def make_headers():
    return {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        'Content-Type': 'application/json',
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Origin': 'https://www.agoda.cn',
        'Referer': f'https://www.agoda.cn/search?city=16958&checkIn={CHECKIN}&checkOut={CHECKOUT}&rooms=1&adults=2&currency=CNY',
        'AG-LANGUAGE-LOCALE': 'zh-cn',
        'AG-CID': '-1',
    }

def build_body(page_size=50, page_number=1, page_token=None, price_from=None, price_to=None):
    body = {
        "operationName": "citySearch",
        "variables": {
            "CitySearchRequest": {
                "cityId": 16958,
                "searchRequest": {
                    "searchCriteria": {
                        "bookingDate": datetime.utcnow().strftime("%Y-%m-%dT05:00:00.000Z"),
                        "checkInDate": f"{CHECKIN}T00:00:00.000Z",
                        "los": 1,
                        "rooms": 1,
                        "adults": 2,
                        "children": 0,
                        "childAges": [],
                        "ratePlans": [1],
                        "featureFlagRequest": {
                            "fiveStarDealOfTheDay": True,
                            "isAllowBookOnRequest": False,
                            "showUnAvailable": True,
                            "showRemainingProperties": True,
                            "isMultiHotelSearch": False,
                            "enableAgencySupplyForPackages": True,
                            "flags": [
                                {"feature": "FamilyChildFriendlyPopularFilter", "enable": True},
                                {"feature": "FamilyChildFriendlyPropertyTypeFilter", "enable": True},
                                {"feature": "FamilyMode", "enable": False}
                            ],
                            "isFlexibleMultiRoomSearch": False,
                            "enablePageToken": True
                        },
                        "isUserLoggedIn": False,
                        "currency": "CNY",
                        "travellerType": "Couple",
                        "isAPSPeek": False,
                        "enableOpaqueChannel": False,
                        "sorting": {"sortField": "Ranking", "sortOrder": "Desc"},
                        "requiredBasis": "PRPN",
                        "requiredPrice": "AllInclusive",
                        "suggestionLimit": 0,
                        "synchronous": False,
                        "isRoomSuggestionRequested": False,
                        "isAPORequest": False,
                        "hasAPOFilter": False,
                        "isAllowBookOnRequest": True,
                        "localCheckInDate": CHECKIN
                    },
                    "searchContext": {
                        "locale": "zh-cn",
                        "cid": -1,
                        "origin": "CN",
                        "platform": 4,
                        "deviceTypeId": 1,
                        "experiments": {"forceByExperiment": []},
                        "isRetry": False,
                        "showCMS": False,
                        "storeFrontId": 3,
                        "pageTypeId": 103,
                        "endpointSearchType": "CitySearch"
                    },
                    "filterRequest": {
                        "idsFilters": [],
                        "rangeFilters": [],
                        "textFilters": []
                    },
                    "matrixGroup": [
                        {"matrixGroup": "StarRating", "size": 100},
                        {"matrixGroup": "AccommodationType", "size": 100},
                        {"matrixGroup": "HotelAreaId", "size": 100},
                    ],
                    "page": {"pageSize": page_size, "pageNumber": page_number},
                    "searchHistory": [],
                    "isTrimmedResponseRequested": False,
                    "extraHotels": {"extraHotelIds": [], "enableFiltersForExtraHotels": False},
                    "highlyRatedAgodaHomesRequest": {
                        "numberOfAgodaHomes": 30, "minimumReviewScore": 7.5, "minimumReviewCount": 3,
                        "accommodationTypes": [28, 29, 30, 102, 103, 106, 107, 108, 109, 110, 114, 115, 120, 131],
                        "sortVersion": 0
                    },
                    "featuredPulsePropertiesRequest": {
                        "numberOfPulseProperties": 15
                    },
                    "rankingRequest": {
                        "isNhaKeywordSearch": False,
                        "isPulseRankingBoost": False
                    },
                    "searchDetailRequest": {
                        "priceHistogramBins": 30
                    }
                }
            }
        }
    }
    if price_from is not None or price_to is not None:
        rf = {"type": "Price", "currencyCode": "CNY"}
        if price_from is not None: rf["from"] = price_from
        if price_to is not None: rf["to"] = price_to
        body["variables"]["CitySearchRequest"]["searchRequest"]["filterRequest"]["rangeFilters"] = [rf]
    if page_token:
        body["variables"]["CitySearchRequest"]["searchRequest"]["page"]["pageToken"] = page_token
    return body

def extract_hotel(prop):
    if not prop or not prop.get("content"):
        return None
    info = prop["content"].get("informationSummary") or {}
    reviews = prop["content"].get("reviews") or {}
    pricing = prop.get("pricing") or {}

    name = info.get("localeName") or info.get("defaultName") or ""
    if not name:
        return None

    # Price from cheapest room
    price = None
    try:
        for offer in (pricing.get("offers") or []):
            for ro in (offer.get("roomOffers") or []):
                for pr in (ro.get("room", {}).get("pricing") or []):
                    d = pr.get("price", {}).get("perBook", {}).get("inclusive", {}).get("display")
                    if d:
                        price = round(d)
                        break
                if price: break
            if price: break
    except: pass

    # Rating
    rating = None
    try: rating = reviews.get("cumulative", {}).get("score")
    except: pass

    # Star rating
    stars = ""
    try:
        acc_type = info.get("accommodationType")
        # Map known accommodation types
        star_map = {1: "1★", 2: "2★", 3: "3★", 4: "4★", 5: "5★"}
        stars = star_map.get(acc_type, "")
    except: pass

    # Location
    location = ""
    try:
        addr = info.get("address") or {}
        if addr.get("area"): location = addr["area"].get("name", "")
        if not location and addr.get("city"): location = addr["city"].get("name", "")
        if not location and addr.get("full"): location = addr["full"]
    except: pass

    return {"name": name, "stars": stars, "rating": rating, "price": price, "location": location}

def scrape_page(session, cookies, headers, **kwargs):
    """Scrape a single page from the API."""
    body = build_body(**kwargs)
    try:
        resp = session.post(GRAPHQL_URL, json=body, headers=headers, cookies=cookies, timeout=30)
        if resp.status_code != 200:
            return [], None
        data = resp.json()
        props = data.get("data", {}).get("citySearch", {}).get("properties") or []
        token = data.get("data", {}).get("citySearch", {}).get("searchEnrichment", {}).get("pageToken")
        return props, token
    except Exception as e:
        print(f"    API请求失败: {e}")
        return [], None

def scrape_bracket(session, cookies, headers, from_price=None, to_price=None):
    """Scrape all hotels in a price bracket using pagination."""
    all_h = []
    seen = set()
    token = None
    page = 1
    MAX_P = 100
    label = f"RMB {from_price or 0}-{to_price or 'max'}"
    print(f"  [{label}] ", end="", flush=True)
    while page <= MAX_P:
        props, token = scrape_page(session, cookies, headers,
            page_size=50, page_number=page, page_token=token,
            price_from=from_price, price_to=to_price)
        if not props:
            if page > 1: break
            page += 1
            continue
        for p in props:
            h = extract_hotel(p)
            if h and h["name"] not in seen:
                seen.add(h["name"])
                all_h.append(h)
        if not token:
            break
        page += 1
        time.sleep(0.3)
    print(f"{len(all_h)}家")
    return all_h

def save_results(hotels):
    """Save to SQLite and Excel."""
    print(f"\n>> 保存 {len(hotels)} 家酒店数据...")

    # SQLite
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT, star_rating TEXT, user_rating REAL,
        price_cny INTEGER, location TEXT, scrape_date TEXT, city TEXT
    )""")
    c.execute("DELETE FROM hotels WHERE city=?", (CITY,))
    for h in hotels:
        c.execute("INSERT INTO hotels (hotel_name, star_rating, user_rating, price_cny, location, scrape_date, city) VALUES (?,?,?,?,?,?,?)",
                  (h["name"], h["stars"], h["rating"], h["price"], h["location"],
                   datetime.now().strftime("%Y-%m-%d"), CITY))
    conn.commit()
    conn.close()
    print(f"  DB: {DB_PATH} ({len(hotels)}条)")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ["序号", "酒店名称", "用户评分", "最低价(CNY)", "星级", "区域位置", "抓取日期"]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign

    sorted_h = sorted(hotels, key=lambda x: (x['price'] or 99999))
    for idx, h in enumerate(sorted_h, 1):
        r = idx + 1
        ws.cell(row=r, column=1, value=idx).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        score_cell = ws.cell(row=r, column=3, value=h['rating'])
        score_cell.border = border
        if h['rating'] and h['rating'] >= 9: score_cell.font = Font(color="FF0000", bold=True)
        ws.cell(row=r, column=4, value=h['price']).border = border
        ws.cell(row=r, column=5, value=h['stars']).border = border
        ws.cell(row=r, column=6, value=h['location']).border = border
        ws.cell(row=r, column=7, value=datetime.now().strftime("%Y-%m-%d")).border = border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 12
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    print(f"  Excel: {EXCEL_PATH}")

    # Stats
    prices = [h['price'] for h in hotels if h['price'] is not None]
    ratings = [h['rating'] for h in hotels if h['rating'] is not None]
    if prices: print(f"  价格区间: RMB {min(prices)} ~ RMB {max(prices)}")
    if ratings: print(f"  评分区间: {min(ratings):.1f} ~ {max(ratings):.1f}, >=9分: {sum(1 for r in ratings if r >= 9)}家")

def main():
    print("=" * 55)
    print(f"  Agoda {CITY}酒店数据抓取")
    print(f"  入住: {CHECKIN}  退房: {CHECKOUT}")
    print(f"  1间, 2成人, 1晚, CNY")
    print("=" * 55)

    # Step 1: Get fresh cookies via Selenium
    cookies = get_fresh_cookies()
    if not cookies:
        print("[错误] 无法获取 Cookie")
        sys.exit(1)

    headers = make_headers()
    session = requests.Session()

    # Step 2: Scrape with price brackets to get all hotels
    brackets = [
        (0, 30), (30, 50), (50, 80), (80, 120), (120, 180), (180, 260),
        (260, 380), (380, 550), (550, 800), (800, 1200), (1200, 1800),
        (1800, 3000), (3000, 100000)
    ]

    ALL = []
    SEEN = set()

    for idx, (frm, to) in enumerate(brackets, 1):
        print(f"\n[{idx}/{len(brackets)}]", flush=True)
        hotels = scrape_bracket(session, cookies, headers, frm, to)
        new = 0
        for h in hotels:
            if h['name'] not in SEEN:
                SEEN.add(h['name'])
                ALL.append(h)
                new += 1
        print(f"  -> 新增{new}家 (总计{len(ALL)}家)")
        time.sleep(0.5)

    print(f"\n{'=' * 55}")
    print(f"  总计 {len(ALL)} 家酒店")
    print(f"{'=' * 55}")

    if ALL:
        save_results(ALL)
    else:
        print("[错误] 未抓到任何酒店数据")

    print("\n>> 完成!")

if __name__ == "__main__":
    import requests
    main()