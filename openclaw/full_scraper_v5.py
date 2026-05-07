#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Full Agoda scraper v5 - 多排序方式全覆盖策略
目标: 抓取 Agoda 武汉酒店全部数据（追上页面显示的 7195 条）
策略: 用多种排序方式获取不同子集 + 健壮翻页 + 去重合并
因为 GraphQL price filter 不生效，改用多种排序突破翻页上限：
  - Ranking (默认推荐排序)
  - Price 升序
  - Price 降序
  - UserRating 降序
"""
import json, time, urllib.request, sys, os, sqlite3
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE, 'api_body.json'), 'r') as f:
    TEMPLATE = json.load(f)

HEADERS = {
    'Content-Type': 'application/json',
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE': 'zh-cn',
    'AG-CID': '-1',
    'AG-PAGE-TYPE-ID': '103',
    'AG-REQUEST-ATTEMPT': '1',
}

# 多种排序方式，每种获取不同的 hotel 子集
SORT_ORDERS = [
    {'sortField': 'Ranking', 'sortOrder': 'Desc'},
    {'sortField': 'Price', 'sortOrder': 'Asc'},
    {'sortField': 'Price', 'sortOrder': 'Desc'},
    {'sortField': 'UserRating', 'sortOrder': 'Desc'},
]

print("=" * 60)
print("  Agoda 武汉酒店 - v5 多排序全覆盖")
print(f"  排序方式: {len(SORT_ORDERS)} 种")
print("=" * 60)


def call_api(sort_field, sort_order, page=1, token=None):
    """调用 Agoda GraphQL API（无价格筛选）"""
    req = json.loads(json.dumps(TEMPLATE))
    sr = req['variables']['CitySearchRequest']['searchRequest']
    # 重置为无价格筛选
    sr['filterRequest']['rangeFilters'] = []
    # 设置排序
    sr['searchCriteria']['sorting'] = {'sortField': sort_field, 'sortOrder': sort_order}
    # 设置翻页
    sr['page'] = {'pageSize': 50, 'pageNumber': page}
    if token:
        sr['page']['pageToken'] = token
    data = json.dumps(req).encode('utf-8')
    r = urllib.request.Request('https://www.agoda.cn/graphql/search', data=data, headers=HEADERS)
    resp = urllib.request.urlopen(r, timeout=60)
    return json.loads(resp.read().decode())


def extract(prop):
    """从 property 中提取酒店信息"""
    if not prop or not prop.get('content'):
        return None
    info = prop['content'].get('informationSummary') or {}
    reviews = prop['content'].get('reviews', {})
    pricing = prop.get('pricing') or {}
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name:
        return None
    price = None
    try:
        for o in (pricing.get('offers') or []):
            for r in (o.get('roomOffers') or []):
                for p in (r.get('room', {}).get('pricing') or []):
                    d = p.get('price', {}).get('perBook', {}).get('inclusive', {}).get('display')
                    if d:
                        price = round(d)
                        break
                if price:
                    break
            if price:
                break
    except Exception:
        pass
    rating = None
    try:
        rating = reviews.get('cumulative', {}).get('score')
    except Exception:
        pass
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'):
        location = addr['area'].get('name', '')
    if not location and addr.get('city'):
        location = addr['city'].get('name', '')
    return {'name': name, 'rating': rating, 'price': price, 'location': location}


# === 主循环 ===
ALL = []
SEEN = set()
TOTAL_API = 0

for si, s in enumerate(SORT_ORDERS, 1):
    sf, so = s['sortField'], s['sortOrder']
    print(f"\n--- [{si}/{len(SORT_ORDERS)}] 排序: {sf} {so} ---")
    sys.stdout.flush()

    new_count = 0
    pages = 0
    page = 1
    token = None
    empty_pages = 0
    retries = 0
    max_retries = 3

    while pages < 500:  # 安全上限
        try:
            data = call_api(sf, so, page, token)
            props = data.get('data', {}).get('citySearch', {}).get('properties') or []
            search_info = data.get('data', {}).get('citySearch', {}).get('searchResult', {}).get('searchInfo', {})
            token = data.get('data', {}).get('citySearch', {}).get('searchEnrichment', {}).get('pageToken')

            # 记录 totalActiveHotels（第一次请求时）
            if pages == 0:
                api_total = search_info.get('totalActiveHotels', 0)
                if api_total > TOTAL_API:
                    TOTAL_API = api_total
                print(f"  totalActiveHotels: {api_total}")

            if not props:
                empty_pages += 1
                if empty_pages >= 3:  # 连续3页空结果退出
                    break
                if not token:
                    break
                page += 1
                time.sleep(0.1)
                continue

            empty_pages = 0
            retries = 0

            for p in props:
                h = extract(p)
                if h and h['name'] not in SEEN:
                    SEEN.add(h['name'])
                    ALL.append(h)
                    new_count += 1

            pages += 1
            if pages % 20 == 0:
                sys.stdout.write(f"   第{pages}页, 本排序+{new_count}新, 累计:{len(ALL)}\n")
                sys.stdout.flush()

            if not token:
                print(f"  翻页结束（无token），共{pages}页")
                break
            page += 1
            time.sleep(0.08)

        except urllib.error.HTTPError as e:
            retries += 1
            if retries > max_retries:
                print(f"  HTTP {e.code} 重试{max_retries}次放弃，共{pages}页")
                break
            sys.stdout.write(f"  HTTP {e.code} 重试{retries}/{max_retries}...\n")
            sys.stdout.flush()
            time.sleep(3)
        except Exception as e:
            retries += 1
            if retries > max_retries:
                print(f"  错误({e}) 重试{max_retries}次放弃，共{pages}页")
                break
            sys.stdout.write(f"  错误({type(e).__name__}) 重试{retries}/{max_retries}...\n")
            sys.stdout.flush()
            if 'timeout' in str(e).lower():
                time.sleep(5)
            else:
                page += 1
                time.sleep(1)

    print(f"  -> 排序 {sf} {so}: +{new_count}新酒店, 翻{pages}页")
    sys.stdout.flush()
    time.sleep(0.3)

# === 结果 ===
print(f"\n\n{'=' * 60}")
print(f"  API 报告总酒店数: {TOTAL_API}")
print(f"  实际抓取去重后:   {len(ALL)}")
print(f"  覆盖率:           {len(ALL)/TOTAL_API*100 if TOTAL_API else 0:.1f}%")
print(f"{'=' * 60}")

if not ALL:
    print("未抓取到数据！")
    sys.exit(1)

# === 保存 JSON ===
out_path = os.path.join(BASE, 'hotels_all.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)
print(f"JSON: {out_path}")

# === 保存 Excel ===
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    xlsx_path = os.path.join(BASE, f'agoda_wuhan_hotels_{len(ALL)}条.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)', '区域位置']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
    for i, h in enumerate(sorted_h, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h['rating']).border = border
        ws.cell(row=r, column=4, value=h['price']).border = border
        ws.cell(row=r, column=5, value=h.get('location', '')).border = border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"Excel: {xlsx_path}")
except Exception as e:
    print(f"Excel保存失败: {e}")

# === 保存 SQLite ===
try:
    db_path = os.path.join(BASE, 'agoda_wuhan.db')
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""CREATE TABLE hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT, user_rating REAL, price_cny INTEGER, location TEXT
    )""")
    for h in ALL:
        c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location) VALUES (?,?,?,?)",
                  (h['name'], h['rating'], h['price'], h['location']))
    conn.commit()
    conn.close()
    print(f"DB: {db_path}")
except Exception as e:
    print(f"DB保存失败: {e}")

# === 统计 ===
prices = [h['price'] for h in ALL if h['price'] is not None]
ratings = [h['rating'] for h in ALL if h['rating'] is not None]
print(f"\n{'=' * 60}")
print(f"  完成！{len(ALL)} 家酒店")
print(f"  API总数: {TOTAL_API}")
if prices:
    print(f"  价格区间: RMB {min(prices):,} ~ RMB {max(prices):,}")
    print(f"  平均价格: RMB {sum(prices)/len(prices):.0f}")
if ratings:
    print(f"  评分区间: {min(ratings):.1f} ~ {max(ratings):.1f}")
    print(f"  评分>=9.0: {sum(1 for r in ratings if r >= 9)}/{len(ratings)}")
print(f"{'=' * 60}")
