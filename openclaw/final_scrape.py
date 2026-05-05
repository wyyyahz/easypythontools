#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Final scrape - NO price filter, just full API pagination"""
import json, time, urllib.request, sys, os, sqlite3
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE, 'api_body.json'), 'r') as f:
    TEMPLATE = json.load(f)

HEADERS = {
    'Content-Type':'application/json','Accept':'*/*',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE':'zh-cn','AG-CID':'-1','AG-PAGE-TYPE-ID':'103','AG-REQUEST-ATTEMPT':'1',
}

def call_api(page=1, token=None):
    req = json.loads(json.dumps(TEMPLATE))
    # Wide price filter to get all hotels
    req['variables']['CitySearchRequest']['searchRequest']['filterRequest']['rangeFilters'][0]['ranges'] = [{'from': 0.0, 'to': 999999.0}]
    req['variables']['CitySearchRequest']['searchRequest']['page'] = {'pageSize': 50, 'pageNumber': page}
    if token: req['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = token
    data = json.dumps(req).encode('utf-8')
    r = urllib.request.Request('https://www.agoda.cn/graphql/search', data=data, headers=HEADERS)
    resp = urllib.request.urlopen(r, timeout=30)
    return json.loads(resp.read().decode())

def extract(prop):
    if not prop or not prop.get('content'): return None
    info = prop['content'].get('informationSummary') or {}
    reviews = prop['content'].get('reviews', {})
    pricing = prop.get('pricing') or {}
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name: return None
    price = None
    try:
        for o in (pricing.get('offers') or []):
            for r in (o.get('roomOffers') or []):
                for p in (r.get('room', {}).get('pricing') or []):
                    d = p.get('price',{}).get('perBook',{}).get('inclusive',{}).get('display')
                    if d: price = round(d); break
                if price: break
            if price: break
    except: pass
    rating = None
    try: rating = reviews.get('cumulative',{}).get('score')
    except: pass
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'): location = addr['area'].get('name', '')
    if not location and addr.get('city'): location = addr['city'].get('name', '')
    return {'name': name, 'rating': rating, 'price': price, 'location': location}

print("=" * 55)
print("  Agoda Wuhan - FINAL FULL SCRAPE (no filter)")
print("=" * 55)

ALL, SEEN, page, token, pages = [], set(), 1, None, 0

# First, try refreshing the captured body by triggering a new CDP capture
# to get a fresh session token
print("\nStarting pagination...")

while page <= 200:
    try:
        data = call_api(page, token)
        props = data.get('data',{}).get('citySearch',{}).get('properties') or []
        token = data.get('data',{}).get('citySearch',{}).get('searchEnrichment',{}).get('pageToken')
        total = data.get('data',{}).get('citySearch',{}).get('searchResult',{}).get('searchInfo',{}).get('totalCount',0)
        
        if not props:
            if page > 1: break
            page += 1; continue
        
        new_count = 0
        for p in props:
            h = extract(p)
            if h and h['name'] not in SEEN:
                SEEN.add(h['name']); ALL.append(h); new_count += 1
        
        pages += 1
        if page % 10 == 0:
            print(f"  Page {page}: {new_count} new (total: {len(ALL)})")
        
        if not token: 
            print(f"  No more token at page {page}")
            break
        
        page += 1
        time.sleep(0.15)
        
    except Exception as e:
        print(f"  Page {page} error: {type(e).__name__}: {e}")
        if page > 2: break
        page += 1; time.sleep(2)

print(f"\n{'='*55}")
print(f"  TOTAL: {len(ALL)} hotels ({pages} pages)")
print(f"{'='*55}")

if not ALL:
    print("No data!")
    sys.exit(1)

# Save
out_path = os.path.join(BASE, 'hotels_all.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font, c.fill, c.alignment = hfont, hfill, halign
    sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
    for i, h in enumerate(sorted_h, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h['rating']).border = border
        ws.cell(row=r, column=4, value=h['price']).border = border
        ws.cell(row=r, column=5, value=h.get('location','')).border = border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"Excel: {xlsx_path}")
except Exception as e:
    print(f"Excel error: {e}")

try:
    db_path = os.path.join(BASE, 'agoda_wuhan.db')
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""CREATE TABLE hotels (id INTEGER PRIMARY KEY AUTOINCREMENT, hotel_name TEXT, user_rating REAL, price_cny INTEGER, location TEXT)""")
    for h in ALL:
        c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location) VALUES (?,?,?,?)",
                  (h['name'], h['rating'], h['price'], h['location']))
    conn.commit()
    conn.close()
    print(f"DB: {db_path}")
except Exception as e:
    print(f"DB error: {e}")

prices = [h['price'] for h in ALL if h['price'] is not None]
ratings = [h['rating'] for h in ALL if h['rating'] is not None]
if prices: print(f"Price: RMB {min(prices):,} ~ RMB {max(prices):,}, Avg: RMB {sum(prices)/len(prices):.0f}")
if ratings: print(f"Rating: {min(ratings):.1f} ~ {max(ratings):.1f}, >=9: {sum(1 for r in ratings if r >= 9)}/{len(ratings)}")
