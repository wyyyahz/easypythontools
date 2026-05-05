#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Full Agoda scraper using the EXACT API body captured from the browser.
"""
import json, os, time, sys, sqlite3
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

# Load the captured API body template
with open(os.path.join(BASE, 'api_body.json'), 'r', encoding='utf-8') as f:
    TEMPLATE = json.load(f)

API_URL = 'https://www.agoda.cn/graphql/search'

# Price brackets from the page
BRACKETS = [
    (0, 20), (20, 30), (30, 40), (40, 50), (50, 60), (60, 70), (70, 80),
    (80, 100), (100, 120), (120, 150), (150, 180), (180, 210), (210, 260),
    (260, 310), (310, 380), (380, 460), (460, 550), (550, 670), (670, 810),
    (810, 980), (980, 1180), (1180, 1430), (1430, 1730), (1730, 2100),
    (2100, 100000)
]

def make_request(from_price, to_price, page_num=1, page_token=None, page_size=50):
    """Create API request with price filter, using the captured template."""
    req = json.loads(json.dumps(TEMPLATE))  # Deep copy
    
    # Update price range filter
    filters = req['variables']['CitySearchRequest']['searchRequest']['filterRequest']['rangeFilters']
    filters[0]['ranges'] = [{'from': float(from_price), 'to': float(to_price) + 0.49}]
    
    # Update page info
    req['variables']['CitySearchRequest']['searchRequest']['page'] = {
        'pageSize': page_size,
        'pageNumber': page_num
    }
    if page_token:
        req['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = page_token
    
    # Update timestamp
    req['variables']['CitySearchRequest']['searchRequest']['searchCriteria']['bookingDate'] = time.strftime('%Y-%m-%dT%H:%M:%S.000Z')
    
    return req

def extract_hotel(prop):
    """Extract hotel info from API property."""
    if not prop or not prop.get('content'):
        return None
    info = prop['content'].get('informationSummary') or {}
    reviews = prop['content'].get('reviews', {})
    pricing = prop.get('pricing') or {}
    
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name:
        return None
    
    # Price
    price = None
    try:
        for offer in (pricing.get('offers') or []):
            for ro in (offer.get('roomOffers') or []):
                for pr in (ro.get('room', {}).get('pricing') or []):
                    d = pr.get('price', {}).get('perBook', {}).get('inclusive', {}).get('display')
                    if d:
                        price = round(d)
                        break
                if price:
                    break
            if price:
                break
    except:
        pass
    
    # Rating
    rating = None
    try:
        rating = reviews.get('cumulative', {}).get('score')
    except:
        pass
    
    # Location
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'):
        location = addr['area'].get('name', '')
    if not location and addr.get('city'):
        location = addr['city'].get('name', '')
    
    # Star rating from accommodationType
    acc_type = info.get('accommodationType', '')
    stars_map = {1: '酒店', 2: '度假村', 3: '旅馆', 34: '酒店', 28: '公寓', 29: '别墅', 30: '民宿'}
    acc_name = stars_map.get(acc_type, '')
    
    return {
        'name': name,
        'stars': acc_name,
        'rating': rating,
        'price': price,
        'location': location,
    }

def main():
    print("=" * 55)
    print("  Agoda 武汉全量抓取 (CDP捕获body)")
    print("=" * 55)
    
    ALL_HOTELS = []
    SEEN = set()
    total_pages = 0
    
    for idx, (frm, to) in enumerate(BRACKETS, 1):
        print(f"\n[{idx}/{len(BRACKETS)}] Bracket RMB {frm}-{to}")
        
        page = 1
        token = None
        bracket_new = 0
        bracket_pages = 0
        
        while page <= 100:
            body = make_request(frm, to, page, token, 50)
            
            try:
                # Use urllib instead of requests to avoid dependency issues
                import urllib.request
                data = json.dumps(body).encode('utf-8')
                req = urllib.request.Request(API_URL, data=data, 
                    headers={'Content-Type': 'application/json', 'Accept': '*/*',
                             'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                             'AG-LANGUAGE-LOCALE': 'zh-cn', 'AG-CID': '-1',
                             'AG-PAGE-TYPE-ID': '103', 'AG-REQUEST-ATTEMPT': '1'})
                
                resp = urllib.request.urlopen(req, timeout=20)
                result = json.loads(resp.read().decode())
                
                props = result.get('data', {}).get('citySearch', {}).get('properties') or []
                token = result.get('data', {}).get('citySearch', {}).get('searchEnrichment', {}).get('pageToken')
                
                if not props:
                    if page > 1:
                        break
                    page += 1
                    continue
                
                for prop in props:
                    h = extract_hotel(prop)
                    if h and h['name'] not in SEEN:
                        SEEN.add(h['name'])
                        ALL_HOTELS.append(h)
                        bracket_new += 1
                
                bracket_pages += 1
                
                if not token:
                    break
                
                page += 1
                time.sleep(0.3)
                
            except Exception as e:
                print(f"  Error on page {page}: {e}")
                if page > 2:
                    break
                page += 1
                time.sleep(1)
        
        total_pages += bracket_pages
        print(f"  -> {bracket_new} new in {bracket_pages} pages (total: {len(ALL_HOTELS)})")
        time.sleep(0.5)
    
    print(f"\n{'=' * 55}")
    print(f"  总计: {len(ALL_HOTELS)} 家酒店 ({total_pages} 页)")
    print(f"{'=' * 55}")
    
    if not ALL_HOTELS:
        print("No data collected!")
        return
    
    # Save to SQLite
    db_path = os.path.join(BASE, 'agoda_wuhan.db')
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""CREATE TABLE hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT, star_rating TEXT, user_rating REAL,
        price_cny INTEGER, location TEXT,
        scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    for h in ALL_HOTELS:
        c.execute("INSERT INTO hotels (hotel_name, star_rating, user_rating, price_cny, location) VALUES (?,?,?,?,?)",
                  (h['name'], h['stars'], h['rating'], h['price'], h['location']))
    conn.commit()
    conn.close()
    
    # Save to Excel
    xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ['序号', '酒店名称', '类型', '用户评分', '最低价(CNY)', '区域位置']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    
    sorted_h = sorted(ALL_HOTELS, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
    for i, h in enumerate(sorted_h, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h['stars']).border = border
        ws.cell(row=r, column=4, value=h['rating']).border = border
        ws.cell(row=r, column=5, value=h['price']).border = border
        ws.cell(row=r, column=6, value=h.get('location', '')).border = border
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 24
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    
    print(f"\n  DB: {db_path}")
    print(f"  Excel: {xlsx_path}")
    
    prices = [h['price'] for h in ALL_HOTELS if h['price'] is not None]
    ratings = [h['rating'] for h in ALL_HOTELS if h['rating'] is not None]
    if prices:
        print(f"  价格: RMB {min(prices):,} ~ RMB {max(prices):,}, 均价 RMB {sum(prices)/len(prices):.0f}")
    if ratings:
        print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f}")

if __name__ == '__main__':
    main()
