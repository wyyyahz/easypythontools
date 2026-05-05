#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Multi-session Agoda scraper:
1. Open new tab, capture fresh API body
2. Paginate to get ~300 hotels
3. Close tab, repeat with new session
4. Merge all results
"""
import json, time, urllib.request, sys, os, sqlite3, asyncio, websockets
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))

HEADERS = {
    'Content-Type':'application/json','Accept':'*/*',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE':'zh-cn','AG-CID':'-1','AG-PAGE-TYPE-ID':'103','AG-REQUEST-ATTEMPT':'1',
}

ALL = []
SEEN = set()
TOTAL_SESSIONS = 5  # Try 5 fresh sessions

def extract(prop):
    if not prop or not prop.get('content'): return None
    info = prop['content'].get('informationSummary') or {}
    reviews = prop['content'].get('reviews', {})
    pricing = prop.get('pricing') or {}
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name: return None
    price = None
    try:
        for o in (pricing.get('offers') or []):
            for r in (o.get('roomOffers') or []):
                for p in (r.get('room', {}).get('pricing') or []):
                    d = p.get('price',{}).get('perBook',{}).get('inclusive',{}).get('display')
                    if d: price = round(d); break
                if price: break
            if price: break
    except: pass
    rating = None
    try: rating = reviews.get('cumulative',{}).get('score')
    except: pass
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'): location = addr['area'].get('name', '')
    if not location and addr.get('city'): location = addr['city'].get('name', '')
    return {'name': name, 'rating': rating, 'price': price, 'location': location}

def fetch_page(body_template, page=1, token=None):
    req = json.loads(json.dumps(body_template))
    req['variables']['CitySearchRequest']['searchRequest']['page'] = {'pageSize': 50, 'pageNumber': page}
    if token: req['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = token
    data = json.dumps(req).encode('utf-8')
    r = urllib.request.Request('https://www.agoda.cn/graphql/search', data=data, headers=HEADERS)
    resp = urllib.request.urlopen(r, timeout=30)
    return json.loads(resp.read().decode())

async def capture_body_via_cdp():
    """Use CDP to capture a fresh API body from a newly opened tab."""
    import urllib.request as ur
    resp = ur.urlopen('http://127.0.0.1:18800/json', timeout=5)
    tabs = json.loads(resp.read())
    
    captured = {'body': None}
    
    # Open a new tab with a fresh URL
    import urllib.parse
    nav_url = ('https://www.agoda.cn/search?city=5818&checkin=2026-06-05'
               '&checkout=2026-06-06&los=1&rooms=1&adults=2&children=0'
               '&currency=CNY&t=' + str(int(time.time() * 1000)))
    
    async with websockets.connect(tabs[0]['webSocketDebuggerUrl'], max_size=20_000_000) as ws:
        await ws.send(json.dumps({'id':1,'method':'Network.enable'}))
        await ws.recv()
        
        await ws.send(json.dumps({'id':2,'method':'Target.createTarget', 
            'params': {'url': 'about:blank', 'newWindow': False}}))
        resp = json.loads(await ws.recv())
        target_id = resp.get('result', {}).get('targetId', '')
        
        if not target_id:
            print("  Failed to create target")
            return None
        
        # Connect to the new tab
        target_ws_url = f"ws://127.0.0.1:18800/devtools/page/{target_id}"
        
    # Now connect to the new tab directly
    async with websockets.connect(target_ws_url, max_size=20_000_000) as ws:
        await ws.send(json.dumps({'id':1,'method':'Network.enable'}))
        await ws.recv()
        
        # Set up listener for graphql request
        async def listen():
            while True:
                try:
                    raw = await asyncio.wait_for(ws.recv(), timeout=15)
                    msg = json.loads(raw)
                    if msg.get('method') == 'Network.requestWillBeSent':
                        req = msg['params'].get('request', {})
                        url = req.get('url', '')
                        if 'graphql/search' in url:
                            post_data = req.get('postData', '')
                            if post_data and len(post_data) > 500:
                                captured['body'] = post_data
                                print(f"  >> Captured! {len(post_data)} bytes")
                                return
                except asyncio.TimeoutError:
                    return
                except:
                    return
        
        # Navigate to the Agoda search page
        await ws.send(json.dumps({'id':2,'method':'Page.navigate', 'params': {'url': nav_url}}))
        
        await asyncio.wait_for(listen(), timeout=20)
        
        # Close the target tab
        try:
            await ws.send(json.dumps({'id':3,'method':'Target.closeTarget', 'params': {'targetId': target_id}}))
        except:
            pass
    
    return captured['body']

async def run_session(session_num):
    """Single session: capture body + paginate."""
    print(f"\n{'='*40}")
    print(f"  Session {session_num}/{TOTAL_SESSIONS}")
    print(f"{'='*40}")
    
    print("  Capturing API body via CDP...")
    body = await capture_body_via_cdp()
    
    if not body:
        print("  Failed to capture body")
        return 0
    
    body_template = json.loads(body)
    
    print("  Paginating...")
    page, token, new_count, pages = 1, None, 0, 0
    
    while page <= 150:
        try:
            data = fetch_page(body_template, page, token)
            props = data.get('data',{}).get('citySearch',{}).get('properties') or []
            token = data.get('data',{}).get('citySearch',{}).get('searchEnrichment',{}).get('pageToken')
            
            if not props:
                if page > 1: break
                page += 1; continue
            
            for p in props:
                h = extract(p)
                if h and h['name'] not in SEEN:
                    SEEN.add(h['name']); ALL.append(h); new_count += 1
            
            pages += 1
            if not token: break
            page += 1
            time.sleep(0.15)
            
        except Exception as e:
            if page > 2: break
            page += 1; time.sleep(2)
    
    print(f"  {new_count} new in {pages} pages (total: {len(ALL)})")
    return new_count

async def main():
    print("=" * 55)
    print("  Multi-Session Agoda Scraper")
    print("=" * 55)
    
    for s in range(1, TOTAL_SESSIONS + 1):
        await run_session(s)
        time.sleep(2)
    
    print(f"\n{'='*55}")
    print(f"  FINAL TOTAL: {len(ALL)} hotels")
    print(f"{'='*55}")
    
    if not ALL:
        print("No data!")
        return
    
    # Save
    out_path = os.path.join(BASE, 'hotels_all.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "武汉酒店列表"
        headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
        hfont = Font(bold=True, color='FFFFFF', size=12)
        hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        halign = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h); c.font, c.fill, c.alignment = hfont, hfill, halign
        sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
        for i, h in enumerate(sorted_h, 1):
            r = i + 1
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=h['name']).border = border
            ws.cell(row=r, column=3, value=h['rating']).border = border
            ws.cell(row=r, column=4, value=h['price']).border = border
            ws.cell(row=r, column=5, value=h.get('location','')).border = border
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 24
        ws.freeze_panes = 'A2'
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
    except Exception as e:
        print(f"Excel error: {e}")
    
    try:
        db_path = os.path.join(BASE, 'agoda_wuhan.db')
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("DROP TABLE IF EXISTS hotels")
        c.execute("""CREATE TABLE hotels (id INTEGER PRIMARY KEY AUTOINCREMENT, hotel_name TEXT, user_rating REAL, price_cny INTEGER, location TEXT)""")
        for h in ALL:
            c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location) VALUES (?,?,?,?)",
                      (h['name'], h['rating'], h['price'], h['location']))
        conn.commit()
        conn.close()
        print(f"DB: {db_path}")
    except Exception as e:
        print(f"DB error: {e}")
    
    prices = [h['price'] for h in ALL if h['price'] is not None]
    ratings = [h['rating'] for h in ALL if h['rating'] is not None]
    if prices: print(f"Price: RMB {min(prices):,} ~ RMB {max(prices):,}, Avg: RMB {sum(prices)/len(prices):.0f}")
    if ratings: print(f"Rating: {min(ratings):.1f} ~ {max(ratings):.1f}")

if __name__ == '__main__':
    asyncio.run(main())
