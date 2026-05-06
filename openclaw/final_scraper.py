#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店最终全量抓取
策略：
  1. Selenium 获取 cookies
  2. API 价格分段精细抓取（带评分/价格/位置详细数据）
  3. Selenium 多排序滚动抓取（补充额外酒店名称）
  4. 合并去重 → 输出最终 Excel
"""
import json, os, sys, time, re
sys.stdout.reconfigure(encoding='utf-8')
BASE = os.path.dirname(os.path.abspath(__file__))

URL = ("https://www.agoda.cn/search?city=5818&checkin=2026-06-05"
       "&checkout=2026-06-06&los=1&rooms=1&adults=2&children=0&currency=CNY")
CHROMEDRIVER = os.path.join(BASE, '..', 'chromedriver.exe')

# =====================================================
# Phase 1: Selenium 获取 cookies + 页面信息
# =====================================================
print("=" * 60)
print("Phase 1: Selenium 获取 cookies")
print("=" * 60)

from selenium import webdriver
from selenium.webdriver.chrome.service import Service

service = Service(CHROMEDRIVER)
options = webdriver.ChromeOptions()
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_argument('--window-size=1920,1080')
prefs = {"profile.managed_default_content_settings.images": 2}
options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(service=service, options=options)

all_cookies = {}
shown_total = 0
try:
    driver.get(URL)
    print("等待页面加载...")
    time.sleep(10)

    for c in driver.get_cookies():
        all_cookies[c['name']] = c['value']
    print(f"获取 cookies: {len(all_cookies)} 个")

    body = driver.find_element('tag name', 'body').text
    m = re.search(r'(\d[\d,]*)\s*个住宿', body)
    if m:
        shown_total = int(m.group(1).replace(',', ''))
    print(f"页面显示住宿: {shown_total} 个")
finally:
    driver.quit()

if not all_cookies:
    print("获取 cookies 失败！")
    sys.exit(1)

# =====================================================
# Phase 2: API 价格分段爬取
# =====================================================
print(f"\n{'=' * 60}")
print("Phase 2: API 价格分段抓取")
print(f"{'=' * 60}")

with open(os.path.join(BASE, 'api_body.json'), 'r', encoding='utf-8') as f:
    TEMPLATE = json.load(f)

HEADERS = {
    'Content-Type': 'application/json',
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE': 'zh-cn',
    'AG-CID': '-1',
    'AG-PAGE-TYPE-ID': '103',
    'AG-REQUEST-ATTEMPT': '1',
    'Origin': 'https://www.agoda.cn',
    'Referer': 'https://www.agoda.cn/',
}

from datetime import datetime
bt = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000') + 'Z'

# 价格分段（与 v7 相同，覆盖全部价格区间）
BRACKETS = []
for i in range(0, 500, 10):        # 0-500: 每10元
    BRACKETS.append((i, i + 10 + 0.49))
for i in range(500, 1000, 25):     # 500-1000: 每25元
    BRACKETS.append((i, i + 25 + 0.49))
for i in range(1000, 2000, 50):    # 1000-2000: 每50元
    BRACKETS.append((i, i + 50 + 0.49))
for i in range(2000, 5000, 100):   # 2000-5000: 每100元
    BRACKETS.append((i, i + 100 + 0.49))
BRACKETS.append((5000, 99999))     # 兜底

import requests
session = requests.Session()
session.cookies.update(all_cookies)

api_hotels = []
api_seen = set()

for idx, (frm, to) in enumerate(BRACKETS, 1):
    sys.stdout.write(f"\r  [{idx}/{len(BRACKETS)}] RMB {frm:.0f}-{to:.0f}: ")
    sys.stdout.flush()

    new_count = 0
    pages = 0
    page = 1
    token = None
    retries = 0
    bracket_total = 0

    while pages < 200:
        try:
            req = json.loads(json.dumps(TEMPLATE))
            sr = req['variables']['CitySearchRequest']['searchRequest']
            sr['searchCriteria']['bookingDate'] = bt
            sr['filterRequest']['rangeFilters'] = [{
                'filterKey': 'Price',
                'ranges': [{'from': float(frm), 'to': float(to)}]
            }]
            sr['filterRequest']['idsFilters'] = []
            sr['page'] = {'pageSize': 50, 'pageNumber': page}
            if token:
                sr['page']['pageToken'] = token

            resp = session.post('https://www.agoda.cn/graphql/search',
                               json=req, headers=HEADERS, timeout=30)

            if resp.status_code != 200:
                retries += 1
                if retries > 3:
                    break
                time.sleep(2)
                continue

            retries = 0
            j = resp.json()
            if 'errors' in j:
                break

            props = j.get('data', {}).get('citySearch', {}).get('properties') or []
            si = j.get('data', {}).get('citySearch', {}).get('searchResult', {}).get('searchInfo', {})
            token = j.get('data', {}).get('citySearch', {}).get('searchEnrichment', {}).get('pageToken')

            if pages == 0:
                bracket_total = si.get('totalFilteredHotels', 0)

            if not props:
                break

            for p in props:
                if not p or not p.get('content'):
                    continue
                info = p['content'].get('informationSummary') or {}
                reviews = p['content'].get('reviews', {})
                pricing = p.get('pricing') or {}
                name = info.get('localeName') or info.get('defaultName') or ''
                if not name:
                    continue

                price = None
                try:
                    for o in (pricing.get('offers') or []):
                        for r in (o.get('roomOffers') or []):
                            for rp in (r.get('room', {}).get('pricing') or []):
                                d = rp.get('price', {}).get('perBook', {}).get('inclusive', {}).get('display')
                                if d:
                                    price = round(d)
                                    break
                            if price:
                                break
                        if price:
                            break
                except:
                    pass

                rating = None
                try:
                    rating = reviews.get('cumulative', {}).get('score')
                except:
                    pass

                location = ''
                addr = info.get('address') or {}
                if addr.get('area'):
                    location = addr['area'].get('name', '')
                if not location and addr.get('city'):
                    location = addr['city'].get('name', '')

                if name not in api_seen:
                    api_seen.add(name)
                    api_hotels.append({'name': name, 'rating': rating, 'price': price, 'location': location})
                    new_count += 1

            pages += 1
            if not token:
                break
            page += 1
            time.sleep(0.03)

        except Exception as e:
            retries += 1
            if retries > 3:
                break
            time.sleep(2)

    sys.stdout.write(f"+{new_count}新 ({pages}p/{bracket_total})\n")
    sys.stdout.flush()

print(f"\nAPI 完成: {len(api_hotels)} 家酒店 (已去重)")

# =====================================================
# Phase 3: Selenium 多排序滚动抓取补充
# =====================================================
print(f"\n{'=' * 60}")
print("Phase 3: Selenium 多排序滚动补充")
print(f"{'=' * 60}")

selenium_hotels = []
selenium_seen = set(api_seen)  # 继承 API 已见集合，避免重复

SORT_OPTIONS = [
    {"name": "推荐排序", "script": "arguments[0].click();",
     "find": "//*[contains(text(),'推荐') or contains(text(),'Ranking') or contains(text(),'默认')]"},
    {"name": "价格从低到高", "find": "//*[contains(text(),'价格从低到高') or contains(text(),'Price Low')]"},
    {"name": "价格从高到低", "find": "//*[contains(text(),'价格从高到低') or contains(text(),'Price High')]"},
    {"name": "评分", "find": "//*[contains(text(),'评分') and contains(text(),'高到低')]"},
]

MAX_SCROLL_ROUNDS = 300
STALL_LIMIT = 30  # 连续无变化上限

for sort_idx, sort_info in enumerate(SORT_OPTIONS):
    print(f"\n--- 排序 [{sort_idx+1}/{len(SORT_OPTIONS)}]: {sort_info['name']} ---")

    driver = webdriver.Chrome(service=service, options=options)
    try:
        driver.get(URL)
        time.sleep(8)

        # 尝试切换排序
        if sort_idx > 0:
            try:
                # 点击排序下拉
                sort_els = driver.find_elements('xpath',
                    "//*[contains(@class,'sort') or contains(@class,'Sort') or contains(@aria-label,'排序')]")
                for el in sort_els:
                    if el.is_displayed():
                        driver.execute_script("arguments[0].click();", el)
                        time.sleep(1)
                        break

                # 选择具体排序
                for sel_text in [
                    sort_info.get('find', ''),
                    f"//*[contains(text(),'{sort_info['name']}')]",
                    f"//*[text()='{sort_info['name']}']",
                ]:
                    if not sel_text:
                        continue
                    try:
                        els = driver.find_elements('xpath', sel_text)
                        for e in els:
                            if e.is_displayed():
                                driver.execute_script("arguments[0].click();", e)
                                print(f"  切换到: {sort_info['name']}")
                                time.sleep(3)
                                break
                        if any(e.is_displayed() for e in els if els):
                            break
                    except:
                        continue
            except Exception as e:
                print(f"  排序切换跳过: {e}")

        # 持续滚动
        last_count = 0
        stall = 0
        for round_num in range(MAX_SCROLL_ROUNDS):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.3)

            # 点击加载更多按钮
            try:
                for btn in driver.find_elements('xpath',
                    "//*[contains(text(),'加载') or contains(text(),'更多') or contains(text(),'显示')]"
                ):
                    try:
                        if btn.is_displayed():
                            driver.execute_script("arguments[0].click();", btn)
                            time.sleep(0.3)
                    except:
                        continue
            except:
                pass

            time.sleep(0.5)

            # 统计已加载酒店数
            try:
                current = driver.execute_script("""
                    return document.querySelectorAll('a[href*="/hotel/wuhan-cn"]').length;
                """)
            except:
                current = 0

            if current != last_count:
                if current - last_count >= 5 or last_count == 0:
                    sys.stdout.write(f"\r  {sort_info['name']}: {current} 家 (+{current-last_count})")
                    sys.stdout.flush()
                last_count = current
                stall = 0
            else:
                stall += 1
                if stall >= STALL_LIMIT:
                    print(f"\n  停止滚动: {stall}次无变化")
                    break

            if round_num % 50 == 0 and round_num > 0:
                print(f"  [{round_num}轮] 当前 {current} 家")

        print(f"\n  滚动完成: {last_count} 家链接")

        # 提取酒店数据
        hotel_json = driver.execute_script("""
            var seen = {}, hotels = [];
            var links = document.querySelectorAll('a[href*="/hotel/wuhan-cn"]');
            for (var i = 0; i < links.length; i++) {
                var text = links[i].textContent.trim();
                if (text.length < 10) continue;
                var pm = text.match(/RMB\\s*([0-9,]+)/);
                var rm = text.match(/([0-9]\\.[0-9])/);
                var name = '';
                var lines = text.split('\\n');
                for (var j = 0; j < lines.length; j++) {
                    var l = lines[j].trim();
                    if (l.length > 2 && l.match(/[\\u4e00-\\u9fff]/) && l.indexOf('RMB') !== 0) {
                        name = l.replace(/^[0-9]+\\s*/, '').trim();
                        break;
                    }
                }
                if (!name) name = (lines[0] || '').trim();
                name = name.replace(/^[0-9]+\\s*/, '').trim();
                if (seen[name] || name.length < 3) continue;
                seen[name] = true;
                hotels.push({
                    name: name,
                    rating: rm ? parseFloat(rm[1]) : null,
                    price: pm ? parseInt(pm[1].replace(/,/g,'')) : null
                });
            }
            return JSON.stringify({count: hotels.length, hotels: hotels});
        """)

        data = json.loads(hotel_json)
        print(f"  提取: {data['count']} 家")

        new_for_sort = 0
        for h in data['hotels']:
            if h['name'] not in selenium_seen:
                selenium_seen.add(h['name'])
                selenium_hotels.append(h)
                new_for_sort += 1
        print(f"  本轮新增: {new_for_sort} 家")

    finally:
        driver.quit()

# =====================================================
# Phase 4: 合并去重
# =====================================================
print(f"\n{'=' * 60}")
print("Phase 4: 合并")
print(f"{'=' * 60}")

# API 数据作为主数据（含详细评分价格位置）
all_hotels = list(api_hotels)  # 先放 API 的
api_names = {h['name'] for h in api_hotels}

# 补充 Selenium 独有的（API 没有的）
selenium_only = 0
for h in selenium_hotels:
    if h['name'] not in api_names:
        all_hotels.append(h)
        selenium_only += 1

print(f"  API: {len(api_hotels)} 家")
print(f"  Selenium补充: {selenium_only} 家")
print(f"  总计: {len(all_hotels)} 家 (页面显示: {shown_total})")

# =====================================================
# Phase 5: 保存 Excel
# =====================================================
print(f"\n{'=' * 60}")
print("Phase 5: 保存 Excel")
print(f"{'=' * 60}")

out_path = os.path.join(BASE, 'hotels_all.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'count': len(all_hotels), 'hotels': all_hotels}, f, ensure_ascii=False, indent=2)
print(f"JSON: {out_path}")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    xlsx_path = os.path.join(BASE, f'agoda_wuhan_hotels_最终_{len(all_hotels)}条.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"

    headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)', '区域位置', '数据来源']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign

    # 按价格排序
    sorted_all = sorted(all_hotels, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)

    for i, h in enumerate(sorted_all, 1):
        r = i + 1
        source = 'API' if h['name'] in api_names else '页面'
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h.get('rating')).border = border
        ws.cell(row=r, column=4, value=h.get('price')).border = border
        ws.cell(row=r, column=5, value=h.get('location', '')).border = border
        ws.cell(row=r, column=6, value=source).border = border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 10
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"Excel: {xlsx_path}")
except Exception as e:
    print(f"Excel失败: {e}")

# 保存 SQLite
try:
    import sqlite3
    db_path = os.path.join(BASE, 'agoda_wuhan_final.db')
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""CREATE TABLE hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT, hotel_name TEXT,
        user_rating REAL, price_cny INTEGER, location TEXT, source TEXT
    )""")
    for h in all_hotels:
        source = 'API' if h['name'] in api_names else '页面'
        c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location, source) VALUES (?,?,?,?,?)",
                  (h['name'], h.get('rating'), h.get('price'), h.get('location', ''), source))
    conn.commit()
    conn.close()
    print(f"DB: {db_path}")
except Exception as e:
    print(f"DB失败: {e}")

# 统计
prices = [h['price'] for h in all_hotels if h.get('price') is not None]
ratings = [h['rating'] for h in all_hotels if h.get('rating') is not None]
print(f"\n{'=' * 60}")
print(f"  最终完成！{len(all_hotels)} 家酒店 (页面: {shown_total})")
if prices:
    print(f"  价格: RMB {min(prices):,} ~ {max(prices):,} 均 RMB {sum(prices)/len(prices):.0f}")
if ratings:
    print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f} >=9分: {sum(1 for r in ratings if r>=9)}/{len(ratings)}")
print(f"{'=' * 60}")
