#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Save Agoda Wuhan hotel data to SQLite and Excel."""
import sqlite3
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "agoda_wuhan.db")
XLSX_PATH = os.path.join(BASE_DIR, "agoda_wuhan_hotels.xlsx")
JSON_PATH = os.path.join(BASE_DIR, "hotels_data.json")

def main():
    # Load data
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    hotels = data.get('hotels', [])
    print(f"处理 {len(hotels)} 家酒店数据...")
    
    # === Save to SQLite ===
    print("保存到 SQLite 数据库...")
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""
        CREATE TABLE hotels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hotel_name TEXT,
            star_rating TEXT,
            user_rating REAL,
            price_cny INTEGER,
            location TEXT,
            url TEXT,
            scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    for h in hotels:
        c.execute("""
            INSERT INTO hotels (hotel_name, star_rating, user_rating, price_cny, location, url)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            h.get('name', ''),
            h.get('stars', ''),
            h.get('rating'),
            h.get('price'),
            h.get('location', ''),
            h.get('url', '')
        ))
    conn.commit()
    conn.close()
    print(f"  数据库保存到: {DB_PATH}")
    
    # === Export to Excel ===
    print("导出到 Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = u"武汉酒店列表"
    
    headers = [u'序号', u'酒店名称', u'星级', u'用户评分', u'最低价(CNY)', u'地址/位置', u'详情链接']
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    sorted_hotels = sorted(hotels, key=lambda h: (h.get('price') or 999999) if h.get('price') is not None else 999999)
    
    for idx, h in enumerate(sorted_hotels, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=h.get('name', '')).border = thin_border
        ws.cell(row=row, column=3, value=h.get('stars', '')).border = thin_border
        ws.cell(row=row, column=4, value=h.get('rating')).border = thin_border
        ws.cell(row=row, column=5, value=h.get('price')).border = thin_border
        ws.cell(row=row, column=6, value=h.get('location', '')).border = thin_border
        url = h.get('url', '')
        if url:
            cell = ws.cell(row=row, column=7, value=url)
            cell.hyperlink = url
            cell.font = Font(color='0563C1', underline='single')
        else:
            ws.cell(row=row, column=7, value='').border = thin_border
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 50
    ws.freeze_panes = 'A2'
    
    wb.save(XLSX_PATH)
    print(f"  Excel保存到: {XLSX_PATH}")
    
    print(f"完成! 共 {len(hotels)} 家酒店")

if __name__ == "__main__":
    main()
