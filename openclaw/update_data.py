#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Save full hotel data to DB and Excel"""
import json, sqlite3, os, sys, re
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

def clean_name(raw):
    """Extract clean hotel name from raw text"""
    # Remove noise
    name = raw
    for prefix in ['Agoda 从此类住宿的订单中赚取的佣金更高。', '特别推荐', 'Domestic Deal', '人气提升', '广告']:
        name = name.replace(prefix, '')
    name = name.split('Rating')[0].split('tooltip')[0].strip()
    return name

conn = sqlite3.connect(os.path.join(BASE, 'agoda_wuhan.db'))
c = conn.cursor()
c.execute("DROP TABLE IF EXISTS hotels")
c.execute("""CREATE TABLE hotels (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    hotel_name TEXT, star_rating TEXT, user_rating REAL,
    price_cny INTEGER, location TEXT, url TEXT,
    scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)""")

with open(os.path.join(BASE, 'hotels_data.json'), 'r', encoding='utf-8') as f:
    data = json.load(f)

hotels = data['hotels']
for h in hotels:
    name = clean_name(h.get('name', ''))
    if len(name) < 3: continue
    c.execute("INSERT INTO hotels (hotel_name, star_rating, user_rating, price_cny, location, url) VALUES (?,?,?,?,?,?)",
              (name, h.get('stars',''), h.get('rating'), h.get('price'), h.get('location',''), h.get('url','')))
conn.commit()
conn.close()

# Excel
wb = Workbook()
ws = wb.active
ws.title = "武汉酒店列表"
headers = ['序号','酒店名称','星级','用户评分','最低价(CNY)','地址/位置','详情链接']
hfont = Font(bold=True, color='FFFFFF', size=12)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font, c.fill, c.alignment = hfont, hfill, halign

sorted_h = sorted(hotels, key=lambda h: (h.get('price') or 99999) if h.get('price') is not None else 99999)
idx = 0
for h in sorted_h:
    name = clean_name(h.get('name',''))
    if len(name) < 3: continue
    idx += 1
    r = idx + 1
    ws.cell(row=r, column=1, value=idx).border = border
    ws.cell(row=r, column=2, value=name).border = border
    ws.cell(row=r, column=3, value=h.get('stars','')).border = border
    ws.cell(row=r, column=4, value=h.get('rating')).border = border
    ws.cell(row=r, column=5, value=h.get('price')).border = border
    ws.cell(row=r, column=6, value=h.get('location','')).border = border
    url = h.get('url','')
    if url:
        cell = ws.cell(row=r, column=7, value=url)
        cell.hyperlink = url
        cell.font = Font(color='0563C1', underline='single')
    else:
        ws.cell(row=r, column=7, value='').border = border

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 22
ws.column_dimensions['G'].width = 50
ws.freeze_panes = 'A2'
wb.save(os.path.join(BASE, 'agoda_wuhan_hotels.xlsx'))

print(f"已保存 {idx} 家酒店到数据库和Excel")
