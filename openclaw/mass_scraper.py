#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
大规模多参数抓取：遍历日期×房型×产品类型的组合，每个开新会话抓一批
"""
import json, urllib.request, time, sys, os, asyncio, websockets, subprocess
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
CHROME = r'C:\Program Files\Google\Chrome\Application\chrome.exe'
H = {'Content-Type':'application/json','Accept':'*/*','User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36','AG-LANGUAGE-LOCALE':'zh-cn','AG-CID':'-1','AG-PAGE-TYPE-ID':'103','AG-REQUEST-ATTEMPT':'1'}

# Load existing data
with open(os.path.join(BASE, 'hotels_all.json'), 'r', encoding='utf-8') as f:
    ALL = json.load(f)['hotels']
SEEN = set(h['name'] for h in ALL)
print(f"Existing: {len(ALL)} hotels")

# Parameter combinations to try
COMBOS = []

# Different dates (each month has different availability)
dates = [
    ('2026-06-12', '2026-06-13'),  # 1 week later
    ('2026-07-05', '2026-07-06'),  # 1 month later
    ('2026-08-05', '2026-08-06'),  # 2 months later
    ('2026-09-05', '2026-09-06'),  # 3 months later
]

# Different room configs
rooms = [
    {'rooms': 1, 'adults': 1, 'children': 0},  # Single
    {'rooms': 1, 'adults': 3, 'children': 0},  # 3 adults
    {'rooms': 2, 'adults': 2, 'children': 0},  # 2 rooms
]

# Product types
prod_types = ['', '&productType=2']  # default + homes/apts

for ci, co in enumerate(dates + [('2026-06-05', '2026-06-06')]):  # dates + original
    ch_in, ch_out = co
    for r in rooms:
        for pt in prod_types:
            COMBOS.append({
                'checkin': ch_in, 'checkout': ch_out,
                'rooms': r['rooms'], 'adults': r['adults'], 'children': r['children'],
                'productType': pt,
                'label': f"{ch_in}_{r['adults']}a{r['rooms']}r{'_apt' if pt else ''}"
            })

# Also try different sort orders for the original config
sorts = ['&sort=Price', '&sort=PriceDesc', '&sort=ReviewScore']
for s in sorts:
    COMBOS.append({
        'checkin': '2026-06-05', 'checkout': '2026-06-06',
        'rooms': 1, 'adults': 2, 'children': 0,
        'productType': '', 'sort': s,
        'label': f"sort_{s.replace('&sort=','')}"
    })

print(f"Total combinations to try: {len(COMBOS)}")
print("Combinations:", [c['label'] for c in COMBOS])

async def capture_and_scrape(combo, port, profile_dir):
    """Start Chrome, capture API, paginate, return new count."""
    label = combo['label']
    
    # Build URL
    url = (f"https://www.agoda.cn/search?city=5818"
           f"&checkin={combo['checkin']}&checkout={combo['checkout']}"
           f"&los=1&rooms={combo['rooms']}&adults={combo['adults']}"
           f"&children={combo['children']}&currency=CNY{combo.get('productType','')}")
    
    # Start Chrome
    proc = subprocess.Popen([CHROME, f'--remote-debugging-port={port}',
        f'--user-data-dir={profile_dir}', '--no-first-run', '--no-default-browser-check',
        '--disable-sync', '--headless=new'],  # headless for speed
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    
    time.sleep(4)
    
    captured = None
    try:
        resp = urllib.request.urlopen(f'http://127.0.0.1:{port}/json/version', timeout=5)
        info = json.loads(resp.read())
        browser_ws = info['webSocketDebuggerUrl']
        
        async with websockets.connect(browser_ws, max_size=20_000_000) as bws:
            await bws.send(json.dumps({'id':1,'method':'Target.createTarget','params':{'url':'about:blank'}}))
            r = json.loads(await bws.recv())
            tid = r['result']['targetId']
            tws = f'ws://127.0.0.1:{port}/devtools/page/{tid}'
        
        async with websockets.connect(tws, max_size=20_000_000) as ws:
            await ws.send(json.dumps({'id':1,'method':'Network.enable'}))
            try: await asyncio.wait_for(ws.recv(), timeout=3)
            except: pass
            
            async def listen():
                nonlocal captured
                while True:
                    raw = await asyncio.wait_for(ws.recv(), timeout=25)
                    m = json.loads(raw)
                    if m.get('method') == 'Network.requestWillBeSent':
                        rq = m['params'].get('request', {})
                        if 'graphql/search' in rq.get('url',''):
                            pd = rq.get('postData','')
                            if pd and len(pd) > 1000:
                                captured = pd
                                return
            
            await ws.send(json.dumps({'id':2,'method':'Page.navigate','params':{'url': url}}))
            try: await asyncio.wait_for(listen(), timeout=30)
            except: pass
    
    except Exception as e:
        print(f"  [{label}] CDP error: {type(e).__name__}")
    
    # Close Chrome
    try: proc.kill()
    except: pass
    
    if not captured:
        return 0
    
    # Paginate
    template = json.loads(captured)
    template['variables']['CitySearchRequest']['searchRequest']['filterRequest']['rangeFilters'] = []
    
    # Apply sort from URL if present
    sort_param = combo.get('sort', '')
    if sort_param == 'Price':
        template['variables']['CitySearchRequest']['searchRequest']['searchCriteria']['sorting'] = {'sortField': 'Price', 'sortOrder': 'Asc'}
    elif sort_param == 'PriceDesc':
        template['variables']['CitySearchRequest']['searchRequest']['searchCriteria']['sorting'] = {'sortField': 'Price', 'sortOrder': 'Desc'}
    elif sort_param == 'ReviewScore':
        template['variables']['CitySearchRequest']['searchRequest']['searchCriteria']['sorting'] = {'sortField': 'ReviewScore', 'sortOrder': 'Desc'}
    
    tk, p, nc, pg = None, 1, 0, 0
    while p <= 100:
        try:
            r = json.loads(json.dumps(template))
            r['variables']['CitySearchRequest']['searchRequest']['page'] = {'pageSize': 50, 'pageNumber': p}
            if tk: r['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = tk
            
            d = json.loads(urllib.request.urlopen(urllib.request.Request(
                'https://www.agoda.cn/graphql/search', data=json.dumps(r).encode('utf-8'), headers=H), timeout=30).read().decode())
            ps = d.get('data',{}).get('citySearch',{}).get('properties') or []
            tk = d.get('data',{}).get('citySearch',{}).get('searchEnrichment',{}).get('pageToken')
            
            if not ps:
                if p > 1: break
                p += 1; continue
            
            for x in ps:
                info = x.get('content',{}).get('informationSummary') or {}
                n = info.get('localeName') or info.get('defaultName') or ''
                if not n or n in SEEN: continue
                SEEN.add(n)
                pr = None
                try:
                    for o in (x.get('pricing',{}).get('offers') or []):
                        for r in (o.get('roomOffers') or []):
                            for pp in (r.get('room',{}).get('pricing') or []):
                                d = pp.get('price',{}).get('perBook',{}).get('inclusive',{}).get('display')
                                if d: pr = round(d); break
                            if pr: break
                        if pr: break
                except: pass
                ra = None
                try: ra = x.get('content',{}).get('reviews',{}).get('cumulative',{}).get('score')
                except: pass
                loc = ''
                addr = info.get('address') or {}
                if addr.get('area'): loc = addr['area'].get('name','')
                ALL.append({'name': n, 'rating': ra, 'price': pr, 'location': loc})
                nc += 1
            pg += 1
            if not tk: break
            p += 1
            time.sleep(0.15)
        except:
            if p > 2: break
            p += 1; time.sleep(2)
    
    return nc

async def main():
    temp_dir = os.path.join(BASE, 'temp_profiles')
    os.makedirs(temp_dir, exist_ok=True)
    
    for idx, combo in enumerate(COMBOS):
        port = 19000 + idx
        profile = os.path.join(temp_dir, f'p{idx}')
        os.makedirs(profile, exist_ok=True)
        
        sys.stdout.write(f"\n[{idx+1}/{len(COMBOS)}] {combo['label']}... ")
        sys.stdout.flush()
        
        nc = await capture_and_scrape(combo, port, profile)
        
        sys.stdout.write(f"{nc} new (total: {len(ALL)})")
        sys.stdout.flush()
        
        time.sleep(0.5)
    
    # Clean up temp dirs
    import shutil
    try: shutil.rmtree(temp_dir)
    except: pass
    
    print(f"\n\n{'='*50}")
    print(f"FINAL: {len(ALL)} hotels")
    print(f"{'='*50}")
    
    # Save
    with open(os.path.join(BASE, 'hotels_all.json'), 'w', encoding='utf-8') as f:
        json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h); c.font, c.fill, c.alignment = hfont, hfill, halign
    sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
    for i, h in enumerate(sorted_h, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h['rating']).border = border
        ws.cell(row=r, column=4, value=h['price']).border = border
        ws.cell(row=r, column=5, value=h.get('location','')).border = border
    ws.column_dimensions['A'].width = 6; ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 14; ws.column_dimensions['E'].width = 24
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    
    pr = [h['price'] for h in ALL if h['price']]
    ra = [h['rating'] for h in ALL if h['rating']]
    if pr: print(f"Price: RMB {min(pr):,} ~ RMB {max(pr):,}, Avg: RMB {sum(pr)/len(pr):.0f}")
    if ra: print(f"Rating: {min(ra):.1f} ~ {max(ra):.1f}")
    print(f"Excel: {xlsx_path}")

if __name__ == '__main__':
    asyncio.run(main())
