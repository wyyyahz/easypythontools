#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Save 431 hotels to Excel"""
import json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE, 'hotels_all.json'), 'r', encoding='utf-8') as f:
    data = json.load(f)
ALL = data['hotels']

xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "武汉酒店列表"
headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
hfont = Font(bold=True, color='FFFFFF', size=12)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font, c.fill, c.alignment = hfont, hfill, halign

sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
for i, h in enumerate(sorted_h, 1):
    r = i + 1
    ws.cell(row=r, column=1, value=i).border = border
    ws.cell(row=r, column=2, value=h['name']).border = border
    ws.cell(row=r, column=3, value=h['rating']).border = border
    ws.cell(row=r, column=4, value=h['price']).border = border
    ws.cell(row=r, column=5, value=h.get('location','')).border = border

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 24
ws.freeze_panes = 'A2'
wb.save(xlsx_path)

prices = [h['price'] for h in ALL if h['price'] is not None]
ratings = [h['rating'] for h in ALL if h['rating'] is not None]
print(f"Excel: {xlsx_path}")
print(f"Hotels: {len(ALL)}")
if prices: print(f"Price: RMB {min(prices):,} ~ RMB {max(prices):,}, Avg: RMB {sum(prices)/len(prices):.0f}")
if ratings: print(f"Rating: {min(ratings):.1f} ~ {max(ratings):.1f}, >=9: {sum(1 for r in ratings if r >= 9)}/{len(ratings)}")
