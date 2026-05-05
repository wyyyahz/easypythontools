#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""遍历多个日期 - 每个日期不同的可用性，捕获更多酒店"""
import json, urllib.request, time, sys, os, asyncio, websockets, subprocess
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
CHROME = r'C:\Program Files\Google\Chrome\Application\chrome.exe'
H = {'Content-Type':'application/json','Accept':'*/*','User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36','AG-LANGUAGE-LOCALE':'zh-cn','AG-CID':'-1','AG-PAGE-TYPE-ID':'103','AG-REQUEST-ATTEMPT':'1'}

with open(os.path.join(BASE, 'hotels_all.json'), 'r', encoding='utf-8') as f:
    ALL = json.load(f)['hotels']
SEEN = set(h['name'] for h in ALL)
print(f"Starting: {len(ALL)} hotels")

# 20 different dates spread across 7 months
DATES = [
    ('2026-06-05', '2026-06-06'),  # original
    ('2026-06-12', '2026-06-13'),
    ('2026-06-19', '2026-06-20'),
    ('2026-06-26', '2026-06-27'),
    ('2026-07-03', '2026-07-04'),
    ('2026-07-10', '2026-07-11'),
    ('2026-07-17', '2026-07-18'),
    ('2026-07-24', '2026-07-25'),
    ('2026-07-31', '2026-08-01'),
    ('2026-08-07', '2026-08-08'),
    ('2026-08-14', '2026-08-15'),
    ('2026-08-21', '2026-08-22'),
    ('2026-09-04', '2026-09-05'),
    ('2026-09-18', '2026-09-19'),
    ('2026-10-02', '2026-10-03'),  # 国庆
    ('2026-10-16', '2026-10-17'),
    ('2026-11-06', '2026-11-07'),
    ('2026-11-20', '2026-11-21'),
    ('2026-12-04', '2026-12-05'),
    ('2026-12-18', '2026-12-19'),
]

# Room configurations
ROOMS = [
    (1, 1, 0),  # 1 adult
    (1, 2, 0),  # 2 adults
    (1, 3, 0),  # 3 adults
]

async def run_date(ci, co, rooms_n, adults_n, port, label):
    url = (f"https://www.agoda.cn/search?city=5818&checkin={ci}&checkout={co}"
           f"&los=1&rooms={rooms_n}&adults={adults_n}&children=0&currency=CNY")
    
    proc = subprocess.Popen([CHROME, f'--remote-debugging-port={port}',
        f'--user-data-dir={os.path.join(BASE, "temp_profiles", label)}',
        '--no-first-run','--no-default-browser-check','--disable-sync','--headless=new'],
        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    time.sleep(3)
    
    captured = None
    try:
        info = json.loads(urllib.request.urlopen(f'http://127.0.0.1:{port}/json/version', timeout=5).read())
        async with websockets.connect(info['webSocketDebuggerUrl'], max_size=20_000_000) as bws:
            await bws.send(json.dumps({'id':1,'method':'Target.createTarget','params':{'url':'about:blank'}}))
            r = json.loads(await bws.recv())
            tws = f'ws://127.0.0.1:{port}/devtools/page/{r["result"]["targetId"]}'
        async with websockets.connect(tws, max_size=20_000_000) as ws:
            await ws.send(json.dumps({'id':1,'method':'Network.enable'}))
            try: await asyncio.wait_for(ws.recv(), timeout=3)
            except: pass
            async def listen():
                nonlocal captured
                while True:
                    raw = await asyncio.wait_for(ws.recv(), timeout=25)
                    m = json.loads(raw)
                    if m.get('method') == 'Network.requestWillBeSent':
                        rq = m['params'].get('request', {})
                        if 'graphql/search' in rq.get('url',''):
                            pd = rq.get('postData','')
                            if pd and len(pd) > 1000: captured = pd; return
            await ws.send(json.dumps({'id':2,'method':'Page.navigate','params':{'url': url}}))
            try: await asyncio.wait_for(listen(), timeout=30)
            except: pass
    except: pass
    try: proc.kill()
    except: pass
    
    if not captured: return 0
    
    template = json.loads(captured)
    template['variables']['CitySearchRequest']['searchRequest']['filterRequest']['rangeFilters'] = []
    
    tk, p, nc, pg = None, 1, 0, 0
    while p <= 100:
        try:
            r = json.loads(json.dumps(template))
            r['variables']['CitySearchRequest']['searchRequest']['page'] = {'pageSize': 50, 'pageNumber': p}
            if tk: r['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = tk
            d = json.loads(urllib.request.urlopen(urllib.request.Request(
                'https://www.agoda.cn/graphql/search', data=json.dumps(r).encode('utf-8'), headers=H), timeout=30).read().decode())
            ps = d.get('data',{}).get('citySearch',{}).get('properties') or []
            tk = d.get('data',{}).get('citySearch',{}).get('searchEnrichment',{}).get('pageToken')
            if not ps:
                if p > 1: break
                p += 1; continue
            for x in ps:
                info = x.get('content',{}).get('informationSummary') or {}
                n = info.get('localeName') or info.get('defaultName') or ''
                if not n or n in SEEN: continue
                SEEN.add(n)
                pr = None
                try:
                    for o in (x.get('pricing',{}).get('offers') or []):
                        for r in (o.get('roomOffers') or []):
                            for pp in (r.get('room',{}).get('pricing') or []):
                                d = pp.get('price',{}).get('perBook',{}).get('inclusive',{}).get('display')
                                if d: pr = round(d); break
                            if pr: break
                        if pr: break
                except: pass
                ra = None
                try: ra = x.get('content',{}).get('reviews',{}).get('cumulative',{}).get('score')
                except: pass
                loc = ''
                addr = info.get('address') or {}
                if addr.get('area'): loc = addr['area'].get('name','')
                ALL.append({'name': n, 'rating': ra, 'price': pr, 'location': loc})
                nc += 1
            pg += 1
            if not tk: break
            p += 1
            time.sleep(0.15)
        except:
            if p > 2: break
            p += 1; time.sleep(2)
    return nc

async def main():
    temp_dir = os.path.join(BASE, 'temp_profiles')
    os.makedirs(temp_dir, exist_ok=True)
    
    combo_idx = 0
    for di, (ci, co) in enumerate(DATES):
        for rn, an, cn in ROOMS:
            combo_idx += 1
            total = len(DATES) * len(ROOMS)
            label = f"{ci}_{an}a"
            port = 19100 + combo_idx
            
            sys.stdout.write(f"\n[{combo_idx}/{total}] {label}... ")
            sys.stdout.flush()
            
            nc = await run_date(ci, co, rn, an, port, label)
            
            sys.stdout.write(f"{nc} new (total: {len(ALL)})")
            sys.stdout.flush()
            time.sleep(0.3)
    
    print(f"\n\n{'='*50}")
    print(f"FINAL: {len(ALL)} hotels")
    print(f"{'='*50}")
    
    with open(os.path.join(BASE, 'hotels_all.json'), 'w', encoding='utf-8') as f:
        json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "武汉酒店列表"
        headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
        hfont = Font(bold=True, color='FFFFFF', size=12)
        hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        halign = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h); c.font, c.fill, c.alignment = hfont, hfill, halign
        sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
        for i, h in enumerate(sorted_h, 1):
            r = i + 1
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=h['name']).border = border
            ws.cell(row=r, column=3, value=h['rating']).border = border
            ws.cell(row=r, column=4, value=h['price']).border = border
            ws.cell(row=r, column=5, value=h.get('location','')).border = border
        ws.column_dimensions['A'].width = 6; ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 14; ws.column_dimensions['E'].width = 24
        ws.freeze_panes = 'A2'
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
    except Exception as e:
        print(f"Excel error: {e}")

if __name__ == '__main__':
    asyncio.run(main())
