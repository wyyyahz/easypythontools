#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店全量抓取 v9 - Selenium 快速滚动
核心：持续向下滚动 + 自动点击"加载更多"
"""
import json, os, sys, time, sqlite3, re
sys.stdout.reconfigure(encoding='utf-8')
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

BASE = os.path.dirname(os.path.abspath(__file__))
service = Service(os.path.join(BASE, '..', 'chromedriver.exe'))
options = webdriver.ChromeOptions()
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_argument('--window-size=1920,1080')
driver = webdriver.Chrome(service=service, options=options)

try:
    url = ("https://www.agoda.cn/search?city=5818&checkin=2026-06-05"
           "&checkout=2026-06-06&los=1&rooms=1&adults=2&children=0&currency=CNY")
    print("正在打开页面...")
    driver.get(url)
    time.sleep(12)

    body = driver.find_element('tag name', 'body').text
    expected = 7000
    m = re.search(r'(\d[\d,]*)\s*个住宿', body)
    if m: expected = int(m.group(1).replace(',',''))
    print(f"页面显示: {expected} 个住宿")

    # JS 持续滚动 + 加载更多
    print("开始持续加载（最多 10 分钟）...")
    t0 = time.time()
    timeout = 600  # 10分钟

    last_count = 0
    stall_count = 0

    while time.time() - t0 < timeout:
        elapsed = time.time() - t0

        # 滚动到底部
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.3)

        # 查找并点击加载更多按钮
        try:
            btns = driver.execute_script("""
                var btns = document.querySelectorAll('button, div, span');
                var found = [];
                for (var i = 0; i < btns.length; i++) {
                    var t = btns[i].textContent || '';
                    if ((t.indexOf('加载') >= 0 || t.indexOf('更多') >= 0 || t.indexOf('显示') >= 0) &&
                        btns[i].offsetParent !== null) {
                        btns[i].click();
                        found.push(t.substring(0, 30));
                    }
                }
                return found;
            """)
            if btns and len(btns) > 0 and elapsed < 30:
                print(f"  点击了 {len(btns)} 个按钮: {btns[:2]}")
        except:
            pass

        # 短暂等待加载
        time.sleep(0.5)

        # 每3秒统计进度
        if int(elapsed) % 3 == 0:
            try:
                current = driver.execute_script("""
                    var s = new Set();
                    var links = document.querySelectorAll('a[href*="/hotel/wuhan-cn"]');
                    for (var i = 0; i < links.length; i++) {
                        var t = links[i].textContent.trim();
                        if (t.length > 10) {
                            var lines = t.split('\\n');
                            for (var j = 0; j < lines.length; j++) {
                                var l = lines[j].trim();
                                if (l.length > 2 && l.match(/[\\u4e00-\\u9fff]/) && l.indexOf('RMB') !== 0) {
                                    s.add(l.replace(/^[0-9]+\\s*/, ''));
                                    break;
                                }
                            }
                        }
                    }
                    return s.size;
                """)
                if current > last_count:
                    if current - last_count >= 10 or last_count == 0:
                        print(f"  已加载 {current} 家 ({elapsed:.0f}s)")
                    last_count = current
                    stall_count = 0
                else:
                    stall_count += 1
                    if stall_count >= 50:  # ~150 秒无变化
                        print(f"  停止：{stall_count}次无变化")
                        break
            except:
                pass

    elapsed = time.time() - t0
    print(f"\n加载阶段完成：{last_count} 家, {elapsed:.0f}秒")

    # 提取数据
    print("正在提取酒店数据...")
    hotel_json = driver.execute_script("""
        var seen = {}, hotels = [];
        var links = document.querySelectorAll('a[href*="/hotel/wuhan-cn"]');
        for (var i = 0; i < links.length; i++) {
            var text = links[i].textContent.trim();
            if (text.length < 10) continue;
            var pm = text.match(/RMB\\s*([0-9,]+)/);
            var rm = text.match(/([0-9]\\.[0-9])/);
            var name = '';
            var lines = text.split('\\n');
            for (var j = 0; j < lines.length; j++) {
                var l = lines[j].trim();
                if (l.length > 2 && l.match(/[\\u4e00-\\u9fff]/) && l.indexOf('RMB') !== 0) {
                    name = l;
                    break;
                }
            }
            if (!name) name = (lines[0] || '').trim();
            name = name.replace(/^[0-9]+\\s*/, '').trim();
            if (seen[name] || name.length < 3) continue;
            seen[name] = true;
            hotels.push({
                name: name,
                rating: rm ? parseFloat(rm[1]) : null,
                price: pm ? parseInt(pm[1].replace(/,/g,'')) : null
            });
        }
        return JSON.stringify({count: hotels.length, hotels: hotels});
    """)

    data = json.loads(hotel_json)
    print(f"提取完成：{data['count']} 家酒店 (页面: {expected})")

    if data['count'] < 100:
        print("酒店数量太少，打印页面内容诊断：")
        print(driver.find_element('tag name','body').text[:500])

    # 保存
    out_path = os.path.join(BASE, 'hotels_all.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON: {out_path}")

    try:
        from openpyxl import Workbook
        xlsx_path = os.path.join(BASE, f'agoda_wuhan_hotels_{data["count"]}条.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "武汉酒店列表"
        headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)']
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        for i, h in enumerate(data['hotels'], 1):
            ws.cell(row=i+1, column=1, value=i)
            ws.cell(row=i+1, column=2, value=h['name'])
            ws.cell(row=i+1, column=3, value=h.get('rating'))
            ws.cell(row=i+1, column=4, value=h.get('price'))
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['D'].width = 14
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
    except Exception as e:
        print(f"Excel失败: {e}")

    # 统计
    prices = [h.get('price') for h in data['hotels'] if h.get('price')]
    ratings = [h.get('rating') for h in data['hotels'] if h.get('rating')]
    print(f"\n{'='*50}")
    print(f"  完成！共 {data['count']} 家酒店 (页面: {expected})")
    if prices: print(f"  价格: RMB {min(prices):,} ~ {max(prices):,}")
    if ratings: print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f}")
    print(f"{'='*50}")

finally:
    driver.quit()
