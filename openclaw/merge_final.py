#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Merge all hotel data sources and save final Excel+DB"""
import json, os, sys, sqlite3
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

# Load all data sources
ALL = []
SEEN = set()

sources = ['hotels_all.json', 'hotels_data.json']
for src in sources:
    path = os.path.join(BASE, src)
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        hotels = data.get('hotels', [])
        for h in hotels:
            name = h.get('name', '')
            if not name or name in SEEN:
                continue
            SEEN.add(name)
            ALL.append({
                'name': name,
                'rating': h.get('rating'),
                'price': h.get('price'),
                'location': h.get('location', '')
            })
        print(f"Source '{src}': {len(hotels)} -> {len([h for h in hotels if h.get('name','') and h['name'] in SEEN and h not in ALL])} added")

# Dedup by name
unique = {}
for h in ALL:
    if h['name'] not in unique:
        unique[h['name']] = h
ALL = list(unique.values())

print(f"\nTotal unique hotels: {len(ALL)}")

if not ALL:
    print("No data!")
    sys.exit(1)

# Save JSON
out = {'count': len(ALL), 'hotels': ALL}
json_path = os.path.join(BASE, 'hotels_all.json')
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

# Save Excel
xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "武汉酒店列表"
headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)', '区域位置']
hfont = Font(bold=True, color='FFFFFF', size=12)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font, c.fill, c.alignment = hfont, hfill, halign

sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
for i, h in enumerate(sorted_h, 1):
    r = i + 1
    ws.cell(row=r, column=1, value=i).border = border
    ws.cell(row=r, column=2, value=h['name']).border = border
    ws.cell(row=r, column=3, value=h['rating']).border = border
    ws.cell(row=r, column=4, value=h['price']).border = border
    ws.cell(row=r, column=5, value=h.get('location', '')).border = border

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 24
ws.freeze_panes = 'A2'
wb.save(xlsx_path)

# Save DB
db_path = os.path.join(BASE, 'agoda_wuhan.db')
conn = sqlite3.connect(db_path)
c = conn.cursor()
c.execute("DROP TABLE IF EXISTS hotels")
c.execute("""CREATE TABLE hotels (id INTEGER PRIMARY KEY AUTOINCREMENT, hotel_name TEXT, user_rating REAL, price_cny INTEGER, location TEXT)""")
for h in ALL:
    c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location) VALUES (?,?,?,?)",
              (h['name'], h['rating'], h['price'], h.get('location', '')))
conn.commit()
conn.close()

# Stats
prices = [h['price'] for h in ALL if h['price'] is not None]
ratings = [h['rating'] for h in ALL if h['rating'] is not None]

print(f"\n{'='*50}")
print(f"  完成！{len(ALL)} 家酒店")
print(f"  Excel: {xlsx_path}")
print(f"  DB: {db_path}")
print(f"  JSON: {json_path}")
print()
if prices:
    print(f"  价格区间: RMB {min(prices):,} ~ RMB {max(prices):,}")
    print(f"  平均价格: RMB {sum(prices)/len(prices):.0f}")
if ratings:
    print(f"  评分区间: {min(ratings):.1f} ~ {max(ratings):.1f}")
    print(f"  评分>=9.0: {sum(1 for r in ratings if r >= 9)}/{len(ratings)}")
print(f"{'='*50}")
