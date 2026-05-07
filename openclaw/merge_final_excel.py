#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
最终合并脚本：
1. 先用 Selenium 快速滚动加载页面（2分钟），提取页面可见的酒店名称
2. 与 v7 的 API 数据（1795家）合并去重
3. 生成最终 Excel
"""
import json, os, sys, time, re
sys.stdout.reconfigure(encoding='utf-8')
BASE = os.path.dirname(os.path.abspath(__file__))

URL = ("https://www.agoda.cn/search?city=5818&checkin=2026-06-05"
       "&checkout=2026-06-06&los=1&rooms=1&adults=2&children=0&currency=CNY")
CHROMEDRIVER = os.path.join(BASE, '..', 'chromedriver.exe')

# ========== 1. 加载 v7 API 数据 ==========
print("=" * 60)
print("加载 v7 API 数据 (1795 家)")
print("=" * 60)

api_path = os.path.join(BASE, 'hotels_all.json')
with open(api_path, 'r', encoding='utf-8') as f:
    api_data = json.load(f)

api_hotels = api_data.get('hotels', [])
api_names = {h['name'] for h in api_hotels}
print(f"API 数据: {len(api_hotels)} 家酒店")

# ========== 2. Selenium 快速滚动提取 ==========
print(f"\n{'=' * 60}")
print("Selenium 快速滚动加载（2分钟）")
print(f"{'=' * 60}")

from selenium import webdriver
from selenium.webdriver.chrome.service import Service

service = Service(CHROMEDRIVER)
options = webdriver.ChromeOptions()
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_argument('--window-size=1920,1080')

driver = webdriver.Chrome(service=service, options=options)

selenium_names = set()
selenium_hotels = []
page_total = 0

try:
    driver.get(URL)
    time.sleep(8)

    # 获取页面显示总数
    body = driver.find_element('tag name', 'body').text
    m = re.search(r'(\d[\d,]*)\s*个住宿', body)
    if m:
        page_total = int(m.group(1).replace(',', ''))
    print(f"页面显示: {page_total} 个住宿")

    # 快速滚动 120 秒
    t0 = time.time()
    last_count = 0
    stall = 0
    scroll_round = 0

    while time.time() - t0 < 120:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.3)

        # 点击加载更多
        try:
            for btn in driver.find_elements('xpath',
                "//*[contains(text(),'加载') or contains(text(),'更多') or contains(text(),'显示')]"
            ):
                try:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        time.sleep(0.2)
                except:
                    continue
        except:
            pass

        time.sleep(0.5)

        try:
            current = driver.execute_script("""
                return document.querySelectorAll('a[href*="/hotel/wuhan-cn"]').length;
            """)
        except:
            current = 0

        if current != last_count:
            if current - last_count >= 5 or last_count == 0:
                elapsed = time.time() - t0
                sys.stdout.write(f"\r  {current} 家链接 (+{current-last_count}) [{elapsed:.0f}s]")
                sys.stdout.flush()
            last_count = current
            stall = 0
        else:
            stall += 1
            if stall >= 20:
                print(f"\n  停止：{stall}次无变化")
                break

        scroll_round += 1

    elapsed = time.time() - t0
    print(f"\n  滚动完成: {scroll_round}轮, {last_count} 家链接, {elapsed:.0f}秒")

    # 提取酒店名称
    print("正在提取酒店名称...")
    result_json = driver.execute_script("""
        var seen = {}, hotels = [];
        var links = document.querySelectorAll('a[href*="/hotel/wuhan-cn"]');
        for (var i = 0; i < links.length; i++) {
            var text = links[i].textContent.trim();
            if (text.length < 10) continue;
            var name = '';
            var lines = text.split('\\n');
            for (var j = 0; j < lines.length; j++) {
                var l = lines[j].trim();
                if (l.length > 2 && l.match(/[\\u4e00-\\u9fff]/) && l.indexOf('RMB') !== 0) {
                    name = l.replace(/^[0-9]+\\s*/, '').trim();
                    break;
                }
            }
            if (!name) name = (lines[0] || '').trim();
            name = name.replace(/^[0-9]+\\s*/, '').trim();
            if (seen[name] || name.length < 3) continue;
            seen[name] = true;

            var pm = text.match(/RMB\\s*([0-9,]+)/);
            var rm = text.match(/([0-9]\\.[0-9])/);

            hotels.push({
                name: name,
                rating: rm ? parseFloat(rm[1]) : null,
                price: pm ? parseInt(pm[1].replace(/,/g,'')) : null
            });
        }
        return JSON.stringify({count: hotels.length, hotels: hotels});
    """)

    sel_data = json.loads(result_json)
    selenium_names = {h['name'] for h in sel_data['hotels']}
    print(f"Selenium 提取: {sel_data['count']} 家唯一酒店")

finally:
    driver.quit()

# ========== 3. 合并 ==========
print(f"\n{'=' * 60}")
print("合并数据")
print(f"{'=' * 60}")

# 找出 Selenium 有但 API 没有的酒店
new_from_selenium = [h for h in sel_data['hotels'] if h['name'] not in api_names]
print(f"API 数据: {len(api_hotels)} 家")
print(f"Selenium 新发现: {len(new_from_selenium)} 家")

# 合并
all_hotels = list(api_hotels)
for h in new_from_selenium:
    all_hotels.append(h)

print(f"合并总计: {len(all_hotels)} 家 (页面: {page_total})")

# ========== 4. 保存 Excel ==========
print(f"\n{'=' * 60}")
print("保存最终 Excel")
print(f"{'=' * 60}")

out_path = os.path.join(BASE, 'hotels_all_final.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'count': len(all_hotels), 'hotels': all_hotels}, f, ensure_ascii=False, indent=2)
print(f"JSON: {out_path}")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    xlsx_path = os.path.join(BASE, f'agoda_wuhan_hotels_最终_{len(all_hotels)}条.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"

    headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)', '区域位置', '数据来源']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign

    # 价格排序
    sorted_all = sorted(all_hotels, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)

    for i, h in enumerate(sorted_all, 1):
        r = i + 1
        source = 'API' if h['name'] in api_names else '页面'
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h.get('rating')).border = border
        ws.cell(row=r, column=4, value=h.get('price')).border = border
        ws.cell(row=r, column=5, value=h.get('location', '')).border = border
        ws.cell(row=r, column=6, value=source).border = border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 10
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"Excel: {xlsx_path}")
except Exception as e:
    print(f"Excel失败: {e}")

# ========== 5. 统计 ==========
prices = [h['price'] for h in all_hotels if h.get('price') is not None]
ratings = [h['rating'] for h in all_hotels if h.get('rating') is not None]
print(f"\n{'=' * 60}")
print(f"  最终完成！{len(all_hotels)} 家酒店 (页面: {page_total})")
if prices:
    print(f"  价格: RMB {min(prices):,} ~ {max(prices):,} 均 RMB {sum(prices)/len(prices):.0f}")
if ratings:
    print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f} >=9分: {sum(1 for r in ratings if r>=9)}/{len(ratings)}")
print(f"{'=' * 60}")
