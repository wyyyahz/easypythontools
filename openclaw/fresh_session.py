#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fresh browser session scraper:
1. Start Chrome with new temp profile
2. Open Agoda, capture API body via CDP
3. Paginate from Python
4. Save data
5. Close Chrome
6. Repeat with another fresh profile
"""
import json, time, urllib.request, sys, os, sqlite3, asyncio, websockets, subprocess
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
ALL_HOTELS = []
SEEN = set()

API_HEADERS = {
    'Content-Type':'application/json','Accept':'*/*',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE':'zh-cn','AG-CID':'-1','AG-PAGE-TYPE-ID':'103','AG-REQUEST-ATTEMPT':'1',
}

def extract(prop):
    if not prop or not prop.get('content'): return None
    info = prop['content'].get('informationSummary') or {}
    reviews = prop['content'].get('reviews', {})
    pricing = prop.get('pricing') or {}
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name: return None
    price = None
    try:
        for o in (pricing.get('offers') or []):
            for r in (o.get('roomOffers') or []):
                for p in (r.get('room', {}).get('pricing') or []):
                    d = p.get('price',{}).get('perBook',{}).get('inclusive',{}).get('display')
                    if d: price = round(d); break
                if price: break
            if price: break
    except: pass
    rating = None
    try: rating = reviews.get('cumulative',{}).get('score')
    except: pass
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'): location = addr['area'].get('name', '')
    if not location and addr.get('city'): location = addr['city'].get('name', '')
    return {'name': name, 'rating': rating, 'price': price, 'location': location}

def paginate(body_str):
    """Paginate through all API results."""
    template = json.loads(body_str)
    token, page, new_count, pages = None, 1, 0, 0
    
    try:
        template['variables']['CitySearchRequest']['searchRequest']['filterRequest']['rangeFilters'] = []
    except:
        pass
    
    while page <= 150:
        try:
            req = json.loads(json.dumps(template))
            req['variables']['CitySearchRequest']['searchRequest']['page'] = {'pageSize': 50, 'pageNumber': page}
            if token: req['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = token
            
            data = json.dumps(req).encode('utf-8')
            r = urllib.request.Request('https://www.agoda.cn/graphql/search', data=data, headers=API_HEADERS)
            resp = urllib.request.urlopen(r, timeout=30)
            result = json.loads(resp.read().decode())
            
            props = result.get('data',{}).get('citySearch',{}).get('properties') or []
            token = result.get('data',{}).get('citySearch',{}).get('searchEnrichment',{}).get('pageToken')
            
            if not props:
                if page > 1: break
                page += 1; continue
            
            for p in props:
                h = extract(p)
                if h and h['name'] not in SEEN:
                    SEEN.add(h['name']); ALL_HOTELS.append(h); new_count += 1
            
            pages += 1
            if not token: break
            page += 1
            time.sleep(0.15)
            
        except Exception as e:
            if page > 2: break
            page += 1; time.sleep(2)
    
    return new_count, pages

async def run_session(session_num, port, profile_dir):
    """Start Chrome, capture API, paginate, close Chrome."""
    print(f"\n[{session_num}] Starting Chrome on port {port}...")
    
    # Start Chrome
    chrome_path = r'C:\Program Files\Google\Chrome\Application\chrome.exe'
    proc = subprocess.Popen([
        chrome_path,
        f'--remote-debugging-port={port}',
        f'--user-data-dir={profile_dir}',
        '--no-first-run', '--no-default-browser-check',
        '--disable-sync', '--disable-default-apps'
    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    
    # Wait for Chrome to start
    time.sleep(3)
    
    captured = None
    
    try:
        # Get the browser WS endpoint
        resp = urllib.request.urlopen(f'http://127.0.0.1:{port}/json/version', timeout=5)
        info = json.loads(resp.read())
        browser_ws = info['webSocketDebuggerUrl']
        
        # Create a new tab and navigate to Agoda
        async with websockets.connect(browser_ws, max_size=20_000_000) as bws:
            # Create a new page target
            await bws.send(json.dumps({'id':1,'method':'Target.createTarget',
                'params':{'url':'about:blank'}}))
            resp = await bws.recv()
            target_id = json.loads(resp).get('result',{}).get('targetId','')
            
            if not target_id:
                print(f"  [{session_num}] Failed to create target")
                proc.kill()
                return 0
            
            target_ws = f"ws://127.0.0.1:{port}/devtools/page/{target_id}"
        
        # Connect to the tab directly
        async with websockets.connect(target_ws, max_size=20_000_000) as ws:
            await ws.send(json.dumps({'id':1,'method':'Network.enable'}))
            try:
                await asyncio.wait_for(ws.recv(), timeout=5)
            except:
                pass
            
            # Listen for graphql request
            async def listen():
                nonlocal captured
                while True:
                    raw = await asyncio.wait_for(ws.recv(), timeout=20)
                    msg = json.loads(raw)
                    if msg.get('method') == 'Network.requestWillBeSent':
                        req = msg['params'].get('request', {})
                        if 'graphql/search' in req.get('url',''):
                            pd = req.get('postData','')
                            if pd and len(pd) > 1000:
                                captured = pd
                                return
            
            # Navigate
            nav_url = ('https://www.agoda.cn/search?city=5818&checkin=2026-06-05'
                       '&checkout=2026-06-06&los=1&rooms=1&adults=2&children=0&currency=CNY')
            
            await ws.send(json.dumps({'id':2,'method':'Page.navigate', 'params':{'url': nav_url}}))
            
            try:
                await asyncio.wait_for(listen(), timeout=20)
            except asyncio.TimeoutError:
                print(f"  [{session_num}] Timeout capturing API")
    
    except Exception as e:
        print(f"  [{session_num}] Error: {e}")
    
    # Close Chrome
    proc.kill()
    proc.wait()
    print(f"  [{session_num}] Chrome closed")
    
    if not captured:
        print(f"  [{session_num}] No API body captured")
        return 0
    
    print(f"  [{session_num}] Body: {len(captured)} bytes")
    
    # Paginate
    new_count, pages = paginate(captured)
    print(f"  [{session_num}] {new_count} new in {pages} pages (total: {len(ALL_HOTELS)})")
    
    return new_count

async def main():
    print("=" * 55)
    print("  Fresh Browser Session Scraper")
    print("=" * 55)
    
    NUM_SESSIONS = 6
    base_port = 18900
    base_dir = os.path.join(BASE, 'temp_profiles')
    os.makedirs(base_dir, exist_ok=True)
    
    for s in range(1, NUM_SESSIONS + 1):
        port = base_port + s
        profile = os.path.join(base_dir, f'profile_{s}')
        os.makedirs(profile, exist_ok=True)
        
        await run_session(s, port, profile)
        time.sleep(1)
    
    # Clean up temp profiles
    import shutil
    for s in range(1, NUM_SESSIONS + 1):
        try:
            shutil.rmtree(os.path.join(base_dir, f'profile_{s}'), ignore_errors=True)
        except:
            pass
    
    print(f"\n{'='*55}")
    print(f"  FINAL: {len(ALL_HOTELS)} unique hotels")
    print(f"{'='*55}")
    
    if not ALL_HOTELS:
        print("No data!")
        return
    
    # Save all formats
    out = {'count': len(ALL_HOTELS), 'hotels': ALL_HOTELS}
    json_path = os.path.join(BASE, 'hotels_all.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "武汉酒店列表"
        headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
        hfont = Font(bold=True, color='FFFFFF', size=12)
        hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        halign = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h); c.font, c.fill, c.alignment = hfont, hfill, halign
        sorted_h = sorted(ALL_HOTELS, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
        for i, h in enumerate(sorted_h, 1):
            r = i + 1
            ws.cell(row=r, column=1, value=i).border = border
            ws.cell(row=r, column=2, value=h['name']).border = border
            ws.cell(row=r, column=3, value=h['rating']).border = border
            ws.cell(row=r, column=4, value=h['price']).border = border
            ws.cell(row=r, column=5, value=h.get('location','')).border = border
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 24
        ws.freeze_panes = 'A2'
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
    except Exception as e:
        print(f"Excel error: {e}")
    
    try:
        db_path = os.path.join(BASE, 'agoda_wuhan.db')
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("DROP TABLE IF EXISTS hotels")
        c.execute("""CREATE TABLE hotels (id INTEGER PRIMARY KEY AUTOINCREMENT, hotel_name TEXT, user_rating REAL, price_cny INTEGER, location TEXT)""")
        for h in ALL_HOTELS:
            c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location) VALUES (?,?,?,?)",
                      (h['name'], h['rating'], h['price'], h['location']))
        conn.commit()
        conn.close()
        print(f"DB: {db_path}")
    except Exception as e:
        print(f"DB error: {e}")
    
    prices = [h['price'] for h in ALL_HOTELS if h['price'] is not None]
    ratings = [h['rating'] for h in ALL_HOTELS if h['rating'] is not None]
    if prices: print(f"Price: RMB {min(prices):,} ~ RMB {max(prices):,}, Avg: RMB {sum(prices)/len(prices):.0f}")
    if ratings: print(f"Rating: {min(ratings):.1f} ~ {max(ratings):.1f}")

if __name__ == '__main__':
    asyncio.run(main())
