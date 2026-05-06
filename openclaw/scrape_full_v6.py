#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店全量抓取 - 混合方案
1. 用 Selenium 打开页面获取有效 cookies + session
2. 用 requests 复用 cookies 批量 API 翻页
3. 合并去重
"""
import json, os, sys, time, urllib.request
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))

# ========================
# 第一步: 用 Selenium 获取有效 cookies
# ========================
print("=" * 60)
print("第一步: 使用 Selenium 获取 cookies...")
print("=" * 60)

from selenium import webdriver
from selenium.webdriver.chrome.service import Service

service = Service(os.path.join(BASE, '..', 'chromedriver.exe'))
options = webdriver.ChromeOptions()
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_argument('--window-size=1920,1080')
# 禁用图片加载加速
prefs = {"profile.managed_default_content_settings.images": 2}
options.add_experimental_option("prefs", prefs)

driver = webdriver.Chrome(service=service, options=options)

cookies = {}
try:
    url = ("https://www.agoda.cn/search?city=5818&checkin=2026-06-05"
           "&checkout=2026-06-06&los=1&rooms=1&adults=2&children=0&currency=CNY")
    driver.get(url)
    print("等待页面加载...")
    time.sleep(8)

    # 获取 cookies
    for c in driver.get_cookies():
        cookies[c['name']] = c['value']

    # 获取页面显示的住宿总数
    body_text = driver.find_element('tag name', 'body').text
    import re
    total_match = re.search(r'找到\s*([0-9,]+)\s*个住宿', body_text)
    if total_match:
        total_str = total_match.group(1).replace(',', '')
        shown_total = int(total_str)
        print(f"页面显示住宿总数: {shown_total}")
    else:
        shown_total = 0
        # Try "搜索结果已更新" pattern
        total_match = re.search(r'(\d[\d,]*)\s*个住宿', body_text)
        if total_match:
            shown_total = int(total_match.group(1).replace(',', ''))
            print(f"页面显示住宿总数: {shown_total}")

    print(f"获取 cookies: {len(cookies)} 个")

finally:
    driver.quit()

if not cookies:
    print("获取 cookies 失败！")
    sys.exit(1)

# ========================
# 第二步: 用 requests 复用 cookies 调用 API
# ========================
print(f"\n{'=' * 60}")
print("第二步: 复用 cookies 批量 API 抓取...")
print(f"{'=' * 60}")

# 加载模板
with open(os.path.join(BASE, 'api_body.json'), 'r') as f:
    TEMPLATE = json.load(f)

HEADERS = {
    'Content-Type': 'application/json',
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE': 'zh-cn',
    'AG-CID': '-1',
    'AG-PAGE-TYPE-ID': '103',
    'AG-REQUEST-ATTEMPT': '1',
    'Origin': 'https://www.agoda.cn',
    'Referer': 'https://www.agoda.cn/',
}

from datetime import datetime
bt = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000') + 'Z'

ALL = []
SEEN = set()

# 策略: 用不同排序 + 价格段（通过 URL 参数触发）
# 实际用 GraphQL API 无筛选，配合不同排序方式
SORT_ORDERS = [
    {'sortField': 'Ranking', 'sortOrder': 'Desc'},
    {'sortField': 'Price', 'sortOrder': 'Asc'},
    {'sortField': 'Price', 'sortOrder': 'Desc'},
]

# 用 requests.Session 带上 cookies
import requests
session = requests.Session()
session.cookies.update(cookies)

for si, s in enumerate(SORT_ORDERS, 1):
    sf, so = s['sortField'], s['sortOrder']
    print(f"\n--- [{si}/{len(SORT_ORDERS)}] 排序: {sf} {so} ---")

    new_count = 0
    pages = 0
    page = 1
    token = None
    retries = 0

    while pages < 500:
        try:
            req = json.loads(json.dumps(TEMPLATE))
            sr = req['variables']['CitySearchRequest']['searchRequest']
            sr['searchCriteria']['bookingDate'] = bt
            sr['searchCriteria']['sorting'] = {'sortField': sf, 'sortOrder': so}
            sr['page'] = {'pageSize': 50, 'pageNumber': page}
            if token:
                sr['page']['pageToken'] = token
            sr['filterRequest']['rangeFilters'] = []
            sr['filterRequest']['idsFilters'] = []

            resp = session.post('https://www.agoda.cn/graphql/search',
                               json=req, headers=HEADERS, timeout=30)

            if resp.status_code != 200:
                retries += 1
                if retries > 3:
                    break
                time.sleep(2)
                continue

            retries = 0
            j = resp.json()

            if 'errors' in j:
                print(f"  GraphQL error: {j['errors'][0]['message'][:100]}")
                break

            props = j.get('data', {}).get('citySearch', {}).get('properties') or []
            search_info = j.get('data', {}).get('citySearch', {}).get('searchResult', {}).get('searchInfo', {})
            token = j.get('data', {}).get('citySearch', {}).get('searchEnrichment', {}).get('pageToken')

            if pages == 0:
                api_total = search_info.get('totalActiveHotels', 0)
                print(f"  totalActiveHotels: {api_total}")

            if not props:
                if not token:
                    break
                page += 1
                time.sleep(0.1)
                continue

            for p in props:
                if not p or not p.get('content'):
                    continue
                info = p['content'].get('informationSummary') or {}
                reviews = p['content'].get('reviews', {})
                pricing = p.get('pricing') or {}
                name = info.get('localeName') or info.get('defaultName') or ''
                if not name:
                    continue

                price = None
                try:
                    for o in (pricing.get('offers') or []):
                        for r in (o.get('roomOffers') or []):
                            for rp in (r.get('room', {}).get('pricing') or []):
                                d = rp.get('price', {}).get('perBook', {}).get('inclusive', {}).get('display')
                                if d:
                                    price = round(d)
                                    break
                            if price:
                                break
                        if price:
                            break
                except:
                    pass

                rating = None
                try:
                    rating = reviews.get('cumulative', {}).get('score')
                except:
                    pass

                location = ''
                addr = info.get('address') or {}
                if addr.get('area'):
                    location = addr['area'].get('name', '')
                if not location and addr.get('city'):
                    location = addr['city'].get('name', '')

                if name not in SEEN:
                    SEEN.add(name)
                    ALL.append({'name': name, 'rating': rating, 'price': price, 'location': location})
                    new_count += 1

            pages += 1
            if pages % 20 == 0:
                print(f"  第{pages}页, +{new_count}新, 累计:{len(ALL)}")

            if not token:
                break
            page += 1
            time.sleep(0.05)

        except Exception as e:
            retries += 1
            if retries > 3:
                break
            time.sleep(2)

    print(f"  -> +{new_count}新酒店, {pages}页")
    time.sleep(0.3)

print(f"\n\n{'=' * 60}")
print(f"  API 抓取累计: {len(ALL)} 家酒店")
print(f"  (页面显示: {shown_total} 家)")
print(f"{'=' * 60}")

# ========================
# 第三步: 如果还不够，用 Selenium 滚动加载补充
# ========================
if len(ALL) < shown_total * 0.8:
    print(f"\n{'=' * 60}")
    print(f"第三步: API 不够 ({len(ALL)}/{shown_total}), 用 Selenium 滚动补充...")
    print(f"{'=' * 60}")

    service = Service(os.path.join(BASE, '..', 'chromedriver.exe'))
    options = webdriver.ChromeOptions()
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_argument('--window-size=1920,1080')

    driver = webdriver.Chrome(service=service, options=options)

    try:
        driver.get(url)
        time.sleep(8)

        last_count = 0
        no_change = 0
        max_rounds = 500

        for round_num in range(max_rounds):
            # Scroll down
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.5)

            # Try clicking load more buttons
            for selector in [
                "//*[contains(text(), '加载更多住宿')]",
                "//*[contains(text(), '加载更多')]",
                "//button[contains(text(), '加载')]",
                "//div[contains(text(), '加载更多')]",
                "//span[contains(text(), '更多')]"
            ]:
                try:
                    for el in driver.find_elements('xpath', selector):
                        try:
                            if el.is_displayed():
                                driver.execute_script("arguments[0].click();", el)
                                time.sleep(0.3)
                        except:
                            continue
                except:
                    continue

            # Count loaded hotels
            try:
                current = driver.execute_script("""
                    return document.querySelectorAll('[data-selenium^="hotel-item"]').length;
                """)
            except:
                current = 0

            if current == 0:
                # Try alternative selector
                try:
                    current = driver.execute_script("""
                        return document.querySelectorAll('a[href*="/hotel/wuhan-cn"]').length;
                    """)
                except:
                    pass

            if current != last_count:
                sys.stdout.write(f"\r  第{round_num}轮: {current} 家酒店 (+{current-last_count})")
                sys.stdout.flush()
                last_count = current
                no_change = 0
            else:
                no_change += 1

            if no_change >= 15:
                print(f"\n  停止: {no_change}次无变化, 共{last_count}家")
                break

        print(f"\nSelenium 加载完成: {last_count} 家酒店")

        # 提取数据
        print("正在从页面提取数据...")
        hotel_data = driver.execute_script("""
            const seen = new Set();
            const hotels = [];
            const items = document.querySelectorAll('[data-selenium^="hotel-item"]');
            if (items.length === 0) {
                // Fallback: try links
                document.querySelectorAll('a[href*="/hotel/wuhan-cn"]').forEach(link => {
                    const text = link.textContent.trim();
                    if (text.length < 10) return;
                    const pm = text.match(/RMB\\s*([0-9,]+)/);
                    const rm = text.match(/([0-9]\\.[0-9])/);
                    let name = '';
                    const lines = text.split('\\n').map(l=>l.trim()).filter(l=>l.length>3);
                    for (const l of lines) {
                        if (l.includes('酒店')||l.includes('客栈')||l.includes('民宿')||
                            l.includes('宾馆')||l.includes('公寓')||l.includes('旅馆')||
                            l.includes('RETURN')) { name = l; break; }
                    }
                    if (!name) name = lines[0]||'';
                    const cn = name.replace(/^[0-9]+\\s*/,'').trim();
                    if (seen.has(cn)||cn.length<3) return;
                    seen.add(cn);
                    hotels.push({
                        name: cn,
                        rating: rm ? parseFloat(rm[1]) : null,
                        price: pm ? parseInt(pm[1].replace(/,/g,'')) : null
                    });
                });
            }
            return JSON.stringify({count: hotels.length, hotels: hotels});
        """)

        selenium_data = json.loads(hotel_data)
        print(f"页面提取: {selenium_data['count']} 家酒店")

        # 合并
        for h in selenium_data['hotels']:
            if h['name'] not in SEEN:
                SEEN.add(h['name'])
                ALL.append(h)

        print(f"合并后总计: {len(ALL)} 家酒店")

    finally:
        driver.quit()

# ========================
# 保存结果
# ========================
print(f"\n{'=' * 60}")
print(f"  最终结果: {len(ALL)} 家酒店")
print(f"  页面显示: {shown_total} 家")
print(f"{'=' * 60}")

if not ALL:
    print("无数据！")
    sys.exit(1)

# 保存 JSON
out_path = os.path.join(BASE, 'hotels_all.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)
print(f"JSON: {out_path}")

# 保存 Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    xlsx_path = os.path.join(BASE, f'agoda_wuhan_hotels_{len(ALL)}条.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)', '区域位置']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
    for i, h in enumerate(sorted_h, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h['rating']).border = border
        ws.cell(row=r, column=4, value=h['price']).border = border
        ws.cell(row=r, column=5, value=h.get('location', '')).border = border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"Excel: {xlsx_path}")
except Exception as e:
    print(f"Excel失败: {e}")

# 保存 SQLite
try:
    import sqlite3
    db_path = os.path.join(BASE, 'agoda_wuhan.db')
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""CREATE TABLE hotels (id INTEGER PRIMARY KEY AUTOINCREMENT, hotel_name TEXT, user_rating REAL, price_cny INTEGER, location TEXT)""")
    for h in ALL:
        c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location) VALUES (?,?,?,?)",
                  (h['name'], h['rating'], h['price'], h.get('location', '')))
    conn.commit()
    conn.close()
    print(f"DB: {db_path}")
except Exception as e:
    print(f"DB失败: {e}")

# 统计
prices = [h['price'] for h in ALL if h['price'] is not None]
ratings = [h['rating'] for h in ALL if h['rating'] is not None]
print(f"\n{'=' * 60}")
print(f"  完成！{len(ALL)} 家酒店 (页面显示: {shown_total})")
if prices:
    print(f"  价格区间: RMB {min(prices):,} ~ RMB {max(prices):,}")
    print(f"  平均价格: RMB {sum(prices)/len(prices):.0f}")
if ratings:
    print(f"  评分区间: {min(ratings):.1f} ~ {max(ratings):.1f}")
print(f"{'=' * 60}")
