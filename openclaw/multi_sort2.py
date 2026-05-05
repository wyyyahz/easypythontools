#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Multi-sort scraper with fresh session"""
import json, urllib.request, time, sys, os
sys.stdout.reconfigure(encoding='utf-8')
BASE = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE, 'hotels_all.json'), 'r', encoding='utf-8') as f:
    ALL = json.load(f)['hotels']
SEEN = set(h['name'] for h in ALL)
print(f"Start: {len(ALL)} hotels")

H = {'Content-Type':'application/json','Accept':'*/*','User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36','AG-LANGUAGE-LOCALE':'zh-cn','AG-CID':'-1','AG-PAGE-TYPE-ID':'103','AG-REQUEST-ATTEMPT':'1'}

SORTS = [
    ('Price', 'Asc'), ('Price', 'Desc'), ('ReviewScore', 'Desc'),
    ('Distance', 'Asc'), ('StarRating', 'Desc'),
]

def run(sf, so):
    with open(os.path.join(BASE, 'api_body_fresh.json'), 'r') as f:
        t = json.loads(f.read())
    t['variables']['CitySearchRequest']['searchRequest']['filterRequest']['rangeFilters'] = []
    t['variables']['CitySearchRequest']['searchRequest']['searchCriteria']['sorting'] = {'sortField': sf, 'sortOrder': so}
    tk, p, nc, pg = None, 1, 0, 0
    while p <= 150:
        try:
            r = json.loads(json.dumps(t))
            r['variables']['CitySearchRequest']['searchRequest']['page'] = {'pageSize': 50, 'pageNumber': p}
            if tk: r['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = tk
            d = json.loads(urllib.request.urlopen(urllib.request.Request('https://www.agoda.cn/graphql/search',data=json.dumps(r).encode('utf-8'),headers=H),timeout=30).read().decode())
            ps = d.get('data',{}).get('citySearch',{}).get('properties') or []
            tk = d.get('data',{}).get('citySearch',{}).get('searchEnrichment',{}).get('pageToken')
            if not ps:
                if p > 1: break
                p += 1; continue
            for x in ps:
                info = x.get('content',{}).get('informationSummary') or {}
                n = info.get('localeName') or info.get('defaultName') or ''
                if not n or n in SEEN: continue
                SEEN.add(n)
                pr = None
                try:
                    for o in (x.get('pricing',{}).get('offers') or []):
                        for r in (o.get('roomOffers') or []):
                            for pp in (r.get('room',{}).get('pricing') or []):
                                d = pp.get('price',{}).get('perBook',{}).get('inclusive',{}).get('display')
                                if d: pr = round(d); break
                            if pr: break
                        if pr: break
                except: pass
                ra = None
                try: ra = x.get('content',{}).get('reviews',{}).get('cumulative',{}).get('score')
                except: pass
                loc = ''
                addr = info.get('address') or {}
                if addr.get('area'): loc = addr['area'].get('name','')
                if not loc and addr.get('city'): loc = addr['city'].get('name','')
                ALL.append({'name': n, 'rating': ra, 'price': pr, 'location': loc})
                nc += 1
            pg += 1
            if not tk: break
            p += 1
            time.sleep(0.15)
        except Exception as e:
            if p > 2: break
            p += 1; time.sleep(2)
    return nc, pg

for sf, so in SORTS:
    sys.stdout.write(f"\n{sf}({so})... ")
    sys.stdout.flush()
    nc, pg = run(sf, so)
    sys.stdout.write(f"{nc} new ({pg}p, total:{len(ALL)})")
    sys.stdout.flush()
    time.sleep(0.5)

print(f"\n\nFINAL: {len(ALL)} hotels")
with open(os.path.join(BASE, 'hotels_all.json'), 'w', encoding='utf-8') as f:
    json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "武汉酒店列表"
headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
hfont = Font(bold=True, color='FFFFFF', size=12)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h); c.font, c.fill, c.alignment = hfont, hfill, halign
sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
for i, h in enumerate(sorted_h, 1):
    r = i + 1
    ws.cell(row=r, column=1, value=i).border = border
    ws.cell(row=r, column=2, value=h['name']).border = border
    ws.cell(row=r, column=3, value=h['rating']).border = border
    ws.cell(row=r, column=4, value=h['price']).border = border
    ws.cell(row=r, column=5, value=h.get('location','')).border = border
ws.column_dimensions['A'].width = 6; ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 10; ws.column_dimensions['D'].width = 14; ws.column_dimensions['E'].width = 24
ws.freeze_panes = 'A2'
wb.save(xlsx_path)

pr = [h['price'] for h in ALL if h['price']]
ra = [h['rating'] for h in ALL if h['rating']]
if pr: print(f"Price: RMB {min(pr):,} ~ RMB {max(pr):,}, Avg: RMB {sum(pr)/len(pr):.0f}")
if ra: print(f"Rating: {min(ra):.1f} ~ {max(ra):.1f}")
print(f"Excel: {xlsx_path}")
