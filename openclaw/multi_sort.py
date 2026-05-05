#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Multi-sort scraper: modify sorting parameter in API body to get different hotel subsets"""
import json, urllib.request, time, sys, os
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))

# Load existing data
with open(os.path.join(BASE, 'hotels_all.json'), 'r', encoding='utf-8') as f:
    existing = json.load(f)
ALL = existing['hotels']
SEEN = set(h['name'] for h in ALL)
print(f"Starting: {len(ALL)} hotels")

# Load API body template
with open(os.path.join(BASE, 'api_body.json'), 'r') as f:
    base_template = json.loads(f.read())

HEADERS = {
    'Content-Type':'application/json','Accept':'*/*',
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE':'zh-cn','AG-CID':'-1','AG-PAGE-TYPE-ID':'103','AG-REQUEST-ATTEMPT':'1',
}

# Different sort configurations
SORTS = [
    {'sortField': 'Price', 'sortOrder': 'Asc'},    # Cheapest first
    {'sortField': 'Price', 'sortOrder': 'Desc'},   # Most expensive first
    {'sortField': 'ReviewScore', 'sortOrder': 'Desc'},  # Highest rated
    {'sortField': 'Distance', 'sortOrder': 'Asc'},      # Nearest
    {'sortField': 'StarRating', 'sortOrder': 'Desc'},   # Highest star rating
]

def extract_hotel(p):
    info = p.get('content',{}).get('informationSummary') or {}
    reviews = p.get('content',{}).get('reviews',{})
    pricing = p.get('pricing') or {}
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name: return None
    price = None
    try:
        for o in (pricing.get('offers') or []):
            for r in (o.get('roomOffers') or []):
                for pr in (r.get('room',{}).get('pricing') or []):
                    d = pr.get('price',{}).get('perBook',{}).get('inclusive',{}).get('display')
                    if d: price = round(d); break
                if price: break
            if price: break
    except: pass
    rating = None
    try: rating = reviews.get('cumulative',{}).get('score')
    except: pass
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'): location = addr['area'].get('name', '')
    if not location and addr.get('city'): location = addr['city'].get('name', '')
    return {'name': name, 'rating': rating, 'price': price, 'location': location}

def paginate_sort(sort_field, sort_order):
    """Paginate using a specific sort order."""
    template = json.loads(json.dumps(base_template))
    template['variables']['CitySearchRequest']['searchRequest']['filterRequest']['rangeFilters'] = []
    template['variables']['CitySearchRequest']['searchRequest']['searchCriteria']['sorting'] = {
        'sortField': sort_field, 'sortOrder': sort_order
    }
    
    token, page, new_count, pages = None, 1, 0, 0
    label = f"{sort_field} ({sort_order})"
    
    while page <= 150:
        try:
            req = json.loads(json.dumps(template))
            req['variables']['CitySearchRequest']['searchRequest']['page'] = {'pageSize': 50, 'pageNumber': page}
            if token: req['variables']['CitySearchRequest']['searchRequest']['page']['pageToken'] = token
            
            data = json.dumps(req).encode('utf-8')
            r = urllib.request.Request('https://www.agoda.cn/graphql/search', data=data, headers=HEADERS)
            resp = urllib.request.urlopen(r, timeout=30)
            result = json.loads(resp.read().decode())
            
            props = result.get('data',{}).get('citySearch',{}).get('properties') or []
            token = result.get('data',{}).get('citySearch',{}).get('searchEnrichment',{}).get('pageToken')
            
            if not props:
                if page > 1: break
                page += 1; continue
            
            for p in props:
                h = extract_hotel(p)
                if h and h['name'] not in SEEN:
                    SEEN.add(h['name']); ALL.append(h); new_count += 1
            
            pages += 1
            if not token: break
            page += 1
            time.sleep(0.15)
            
        except Exception as e:
            if page > 2: break
            page += 1; time.sleep(2)
    
    return new_count, pages

for sort in SORTS:
    print(f"\nSort: {sort['sortField']} ({sort['sortOrder']})...")
    new_count, pages = paginate_sort(sort['sortField'], sort['sortOrder'])
    print(f"  {new_count} new in {pages} pages (total: {len(ALL)})")
    time.sleep(0.5)

print(f"\n{'='*50}")
print(f"FINAL: {len(ALL)} hotels")
print(f"{'='*50}")

# Save
with open(os.path.join(BASE, 'hotels_all.json'), 'w', encoding='utf-8') as f:
    json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "武汉酒店列表"
headers = ['序号','酒店名称','用户评分','最低价(CNY)','区域位置']
hfont = Font(bold=True, color='FFFFFF', size=12)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h); c.font, c.fill, c.alignment = hfont, hfill, halign
sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
for i, h in enumerate(sorted_h, 1):
    r = i + 1
    ws.cell(row=r, column=1, value=i).border = border
    ws.cell(row=r, column=2, value=h['name']).border = border
    ws.cell(row=r, column=3, value=h['rating']).border = border
    ws.cell(row=r, column=4, value=h['price']).border = border
    ws.cell(row=r, column=5, value=h.get('location','')).border = border
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 24
ws.freeze_panes = 'A2'
wb.save(xlsx_path)

prices = [h['price'] for h in ALL if h['price'] is not None]
ratings = [h['rating'] for h in ALL if h['rating'] is not None]
if prices: print(f"Price: RMB {min(prices):,} ~ RMB {max(prices):,}, Avg: RMB {sum(prices)/len(prices):.0f}")
if ratings: print(f"Rating: {min(ratings):.1f} ~ {max(ratings):.1f}")
print(f"Excel: {xlsx_path}")
