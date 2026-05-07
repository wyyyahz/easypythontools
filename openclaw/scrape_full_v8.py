#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店全量抓取 v8 - Selenium 滚动加载
"""
import json, os, sys, time, sqlite3, re
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
from selenium import webdriver
from selenium.webdriver.chrome.service import Service

service = Service(os.path.join(BASE, '..', 'chromedriver.exe'))
options = webdriver.ChromeOptions()
options.add_argument('--disable-blink-features=AutomationControlled')
options.add_experimental_option('excludeSwitches', ['enable-automation'])
options.add_argument('--window-size=1920,1080')

driver = webdriver.Chrome(service=service, options=options)

try:
    url = ("https://www.agoda.cn/search?city=5818&checkin=2026-06-05"
           "&checkout=2026-06-06&los=1&rooms=1&adults=2&children=0&currency=CNY")
    print("打开 Agoda 页面...")
    driver.get(url)
    time.sleep(12)

    # 获取页面显示总数
    body_text = driver.find_element('tag name', 'body').text
    expected = 7000
    m = re.search(r'(\d[\d,]*)\s*个住宿', body_text)
    if m:
        expected = int(m.group(1).replace(',', ''))
    print(f"页面显示: {expected} 个住宿")

    # 滚动加载
    print("开始滚动加载（最长 5 分钟）...")
    t0 = time.time()
    timeout = 300

    last_count = 0
    no_change = 0
    last_report = 0

    while time.time() - t0 < timeout:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(0.5)

        # 点击加载更多
        for btn in driver.find_elements('xpath', "//*[contains(text(), '加载更多')]"):
            try:
                if btn.is_displayed():
                    driver.execute_script("arguments[0].click();", btn)
                    time.sleep(0.3)
            except:
                pass

        elapsed = time.time() - t0
        if elapsed - last_report >= 5:
            last_report = elapsed
            try:
                # 统计酒店链接数
                current = driver.execute_script(
                    'return document.querySelectorAll(\'a[href*="/hotel/wuhan-cn"]\').length;'
                )
                if current > last_count:
                    print(f"  {current} 家酒店 ({elapsed:.0f}s)")
                    last_count = current
                    no_change = 0
                else:
                    no_change += 1
                    if no_change >= 30:
                        print(f"  30次无变化，结束")
                        break
            except Exception as e:
                print(f"  统计出错: {e}")
                pass

    elapsed = time.time() - t0
    print(f"加载完成: {last_count} 家酒店, {elapsed:.0f}秒")

    # 提取数据
    print("提取酒店数据...")
    hotel_json = driver.execute_script("""
        var seen = {};
        var hotels = [];
        var links = document.querySelectorAll('a[href*="/hotel/wuhan-cn"]');
        for (var i = 0; i < links.length; i++) {
            var text = links[i].textContent.trim();
            if (text.length < 10) continue;
            var pm = text.match(/RMB\\s*([0-9,]+)/);
            var rm = text.match(/([0-9]\\.[0-9])/);
            var name = '';
            var lines = text.split('\\n');
            for (var j = 0; j < lines.length; j++) {
                var l = lines[j].trim();
                if (l.length > 2 && /[\\u4e00-\\u9fff]/.test(l) && l.indexOf('RMB') !== 0 && l.indexOf('\\u00b7') !== 0) {
                    name = l;
                    break;
                }
            }
            if (!name) name = lines[0] || '';
            name = name.replace(/^[0-9]+\\s*/, '').trim();
            if (seen[name] || name.length < 3) continue;
            seen[name] = true;
            hotels.push({
                name: name,
                rating: rm ? parseFloat(rm[1]) : null,
                price: pm ? parseInt(pm[1].replace(/,/g,'')) : null
            });
        }
        return JSON.stringify({count: hotels.length, hotels: hotels});
    """)

    data = json.loads(hotel_json)
    print(f"提取完成: {data['count']} 家酒店 (页面: {expected})")

    if data['count'] == 0:
        print("警告：没有提取到酒店数据！打印页面片段：")
        print(driver.find_element('tag name', 'body').text[:1000])

    # 保存
    out_path = os.path.join(BASE, 'hotels_all.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON: {out_path}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        xlsx_path = os.path.join(BASE, f'agoda_wuhan_hotels_{data["count"]}条.xlsx')
        wb = Workbook()
        ws = wb.active
        ws.title = "武汉酒店列表"
        headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)']
        hfont = Font(bold=True, color='FFFFFF', size=12)
        hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        halign = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h).font = hfont
            ws.cell(row=1, column=i, value=h).fill = hfill
            ws.cell(row=1, column=i, value=h).alignment = halign
        for i, h in enumerate(data['hotels'], 1):
            ws.cell(row=i+1, column=1, value=i).border = border
            ws.cell(row=i+1, column=2, value=h['name']).border = border
            ws.cell(row=i+1, column=3, value=h.get('rating')).border = border
            ws.cell(row=i+1, column=4, value=h.get('price')).border = border
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 14
        ws.freeze_panes = 'A2'
        wb.save(xlsx_path)
        print(f"Excel: {xlsx_path}")
    except Exception as e:
        print(f"Excel失败: {e}")

    try:
        db_path = os.path.join(BASE, 'agoda_wuhan.db')
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("DROP TABLE IF EXISTS hotels")
        c.execute("CREATE TABLE hotels (id INTEGER PRIMARY KEY AUTOINCREMENT, hotel_name TEXT, user_rating REAL, price_cny INTEGER)")
        for h in data['hotels']:
            c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny) VALUES (?,?,?)",
                      (h['name'], h.get('rating'), h.get('price')))
        conn.commit()
        conn.close()
        print(f"DB: {db_path}")
    except Exception as e:
        print(f"DB失败: {e}")

    prices = [h.get('price') for h in data['hotels'] if h.get('price') is not None]
    ratings = [h.get('rating') for h in data['hotels'] if h.get('rating') is not None]
    print(f"\n{'='*50}")
    print(f"  完成！{data['count']} 家酒店 (页面: {expected})")
    if prices:
        print(f"  价格: RMB {min(prices):,} ~ {max(prices):,}, 均价 {sum(prices)/len(prices):.0f}")
    print(f"{'='*50}")

finally:
    driver.quit()
