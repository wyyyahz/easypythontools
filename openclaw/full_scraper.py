#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Full Agoda Wuhan hotel scraper via direct GraphQL API calls.
Uses cookies from browser session + complete request body structure.
"""
import json
import os
import sys
import time
import sqlite3
sys.stdout.reconfigure(encoding='utf-8')

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))

# ========== COOKIES from browser session ==========
COOKIES = {
    'agoda.user.03': 'UserId=83133924-5561-46c1-979d-dd87b2787ff1',
    'agoda.prius': 'PriusID=0&PointsMaxTraffic=Agoda',
}

# ========== HEADERS ==========
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
    'Content-Type': 'application/json',
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Origin': 'https://www.agoda.cn',
    'Referer': 'https://www.agoda.cn/',
    'AG-LANGUAGE-LOCALE': 'zh-cn',
    'AG-CID': '-1',
    'AG-PAGE-TYPE-ID': '103',
    'AG-REQUEST-ATTEMPT': '1',
}

API_URL = 'https://www.agoda.cn/graphql/search'

# ========== FULL REQUEST BODY TEMPLATE ==========
def make_request_body(city_id, check_in, check_out, los, rooms, adults, children, 
                      currency='CNY', page_size=50, page_number=1, page_token=None,
                      price_from=None, price_to=None):
    """Build the complete GraphQL request body."""
    
    body = {
        "operationName": "citySearch",
        "variables": {
            "CitySearchRequest": {
                "cityId": city_id,
                "searchRequest": {
                    "searchCriteria": {
                        "bookingDate": time.strftime("2026-05-05T05:00:00.000Z"),
                        "checkInDate": f"{check_in}T00:00:00.000Z",
                        "los": los,
                        "rooms": rooms,
                        "adults": adults,
                        "children": children,
                        "childAges": [],
                        "ratePlans": [1],
                        "featureFlagRequest": {
                            "fiveStarDealOfTheDay": True,
                            "isAllowBookOnRequest": False,
                            "showUnAvailable": True,
                            "showRemainingProperties": True,
                            "isMultiHotelSearch": False,
                            "enableAgencySupplyForPackages": True,
                            "flags": [
                                {"feature": "FamilyChildFriendlyPopularFilter", "enable": True},
                                {"feature": "FamilyChildFriendlyPropertyTypeFilter", "enable": True},
                                {"feature": "FamilyMode", "enable": False}
                            ],
                            "isFlexibleMultiRoomSearch": False,
                            "enablePageToken": True
                        },
                        "isUserLoggedIn": False,
                        "currency": currency,
                        "travellerType": "Couple",
                        "isAPSPeek": False,
                        "enableOpaqueChannel": False,
                        "sorting": {"sortField": "Ranking", "sortOrder": "Desc"},
                        "requiredBasis": "PRPN",
                        "requiredPrice": "AllInclusive",
                        "suggestionLimit": 0,
                        "synchronous": False,
                        "isRoomSuggestionRequested": False,
                        "isAPORequest": False,
                        "hasAPOFilter": False,
                        "isAllowBookOnRequest": True,
                        "localCheckInDate": check_in
                    },
                    "searchContext": {
                        "locale": "zh-cn",
                        "cid": -1,
                        "origin": "CN",
                        "platform": 4,
                        "deviceTypeId": 1,
                        "experiments": {"forceByExperiment": []},
                        "isRetry": False,
                        "showCMS": False,
                        "storeFrontId": 3,
                        "pageTypeId": 103,
                        "endpointSearchType": "CitySearch"
                    },
                    "filterRequest": {
                        "idsFilters": [],
                        "rangeFilters": [],
                        "textFilters": []
                    },
                    "matrixGroup": [
                        {"matrixGroup": "StarRating", "size": 100},
                        {"matrixGroup": "AccommodationType", "size": 100},
                        {"matrixGroup": "HotelAreaId", "size": 100},
                        {"matrixGroup": "HotelFacilities", "size": 100},
                        {"matrixGroup": "ReviewScore", "size": 100},
                        {"matrixGroup": "PaymentOptions", "size": 100},
                        {"matrixGroup": "RoomBenefits", "size": 100},
                        {"matrixGroup": "CityCenterDistance", "size": 100},
                        {"matrixGroup": "RoomAmenities", "size": 100},
                        {"matrixGroup": "GroupedBedTypes", "size": 100},
                        {"matrixGroup": "NumberOfBedrooms", "size": 100},
                        {"matrixGroup": "LandmarkIds", "size": 100},
                        {"matrixGroup": "ReviewLocationScore", "size": 100},
                        {"matrixGroup": "BeachAccessTypeIds", "size": 100}
                    ],
                    "page": {
                        "pageSize": page_size,
                        "pageNumber": page_number
                    },
                    "searchHistory": [],
                    "isTrimmedResponseRequested": False,
                    "extraHotels": {
                        "extraHotelIds": [],
                        "enableFiltersForExtraHotels": False
                    },
                    "highlyRatedAgodaHomesRequest": {
                        "numberOfAgodaHomes": 30,
                        "minimumReviewScore": 7.5,
                        "minimumReviewCount": 3,
                        "accommodationTypes": [28, 29, 30, 102, 103, 106, 107, 108, 109, 110, 114, 115, 120, 131],
                        "sortVersion": 0
                    },
                    "featuredPulsePropertiesRequest": {
                        "numberOfPulseProperties": 15
                    },
                    "rankingRequest": {
                        "isNhaKeywordSearch": False,
                        "isPulseRankingBoost": False
                    },
                    "searchDetailRequest": {
                        "priceHistogramBins": 30
                    }
                }
            }
        }
    }
    
    # Add price range filter if provided
    if price_from is not None or price_to is not None:
        range_filter = {}
        if price_from is not None:
            range_filter["from"] = price_from
        if price_to is not None:
            range_filter["to"] = price_to
        range_filter["type"] = "Price"
        range_filter["currencyCode"] = "CNY"
        body["variables"]["CitySearchRequest"]["searchRequest"]["filterRequest"]["rangeFilters"] = [range_filter]
    
    # Add page token for subsequent pages
    if page_token:
        body["variables"]["CitySearchRequest"]["searchRequest"]["page"]["pageToken"] = page_token
    
    return body


def extract_hotel(property_obj):
    """Extract hotel info from API response property."""
    if not property_obj or not property_obj.get('content'):
        return None
    info = property_obj['content'].get('informationSummary') or {}
    reviews = property_obj['content'].get('reviews', {})
    pricing = property_obj.get('pricing') or {}
    
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name:
        return None
    
    # Price from cheapest room
    price = None
    try:
        for offer in (pricing.get('offers') or []):
            for ro in (offer.get('roomOffers') or []):
                for pr in (ro.get('room', {}).get('pricing') or []):
                    d = pr.get('price', {}).get('perBook', {}).get('inclusive', {}).get('display')
                    if d:
                        price = round(d)
                        break
                if price:
                    break
            if price:
                break
    except:
        pass
    
    # Rating
    rating = None
    try:
        rating = reviews.get('cumulative', {}).get('score')
    except:
        pass
    
    # Star rating - check accommodationType for hotel category
    stars = ''
    acc_type = info.get('accommodationType')
    
    # Location
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'):
        location = addr['area'].get('name', '')
    if not location and addr.get('city'):
        location = addr['city'].get('name', '')
    
    return {
        'name': name,
        'stars': stars,
        'rating': rating,
        'price': price,
        'location': location,
        'acc_type': acc_type
    }


def scrape_bracket(session, from_price=None, to_price=None):
    """Scrape all hotels in a price bracket using API pagination."""
    ALL = []
    SEEN = set()
    page = 1
    token = None
    MAX_PAGES = 200
    
    print(f"  Bracket {from_price or 0}-{to_price or 'max'}...", end=' ', flush=True)
    
    while page <= MAX_PAGES:
        body = make_request_body(
            city_id=5818,
            check_in='2026-06-05',
            check_out='2026-06-06',
            los=1, rooms=1, adults=2, children=0,
            page_size=50, page_number=page, page_token=token,
            price_from=from_price, price_to=to_price
        )
        
        try:
            resp = session.post(API_URL, json=body, headers=HEADERS, cookies=COOKIES, timeout=20)
            
            if resp.status_code != 200:
                if page > 1:
                    break
                page += 1
                continue
            
            data = resp.json()
            properties = data.get('data', {}).get('citySearch', {}).get('properties') or []
            token = data.get('data', {}).get('citySearch', {}).get('searchEnrichment', {}).get('pageToken')
            
            if not properties:
                if page > 1:
                    break
                page += 1
                continue
            
            for prop in properties:
                h = extract_hotel(prop)
                if h and h['name'] not in SEEN:
                    SEEN.add(h['name'])
                    ALL.append(h)
            
            if not token:
                break
            
            page += 1
            time.sleep(0.3)  # Be respectful
            
        except Exception as e:
            if page > 2:
                break
            page += 1
            time.sleep(2)
    
    print(f"{len(ALL)} hotels")
    return ALL


def save_results(all_hotels):
    """Save to SQLite and Excel."""
    print(f"\nSaving {len(all_hotels)} hotels...")
    
    # SQLite
    db_path = os.path.join(BASE, 'agoda_wuhan.db')
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""CREATE TABLE hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT, star_rating TEXT, user_rating REAL,
        price_cny INTEGER, location TEXT, acc_type INTEGER,
        scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    
    for h in all_hotels:
        c.execute("INSERT INTO hotels (hotel_name, star_rating, user_rating, price_cny, location, acc_type) VALUES (?,?,?,?,?,?)",
                  (h['name'], h['stars'], h['rating'], h['price'], h['location'], h.get('acc_type')))
    conn.commit()
    conn.close()
    
    # Excel
    xlsx_path = os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    
    headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)', '区域位置']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    
    sorted_h = sorted(all_hotels, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
    for idx, h in enumerate(sorted_h, 1):
        r = idx + 1
        ws.cell(row=r, column=1, value=idx).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h['rating']).border = border
        ws.cell(row=r, column=4, value=h['price']).border = border
        ws.cell(row=r, column=5, value=h.get('location', '')).border = border
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    
    print(f"  DB: {db_path}")
    print(f"  Excel: {xlsx_path}")
    
    # Stats
    prices = [h['price'] for h in all_hotels if h['price'] is not None]
    ratings = [h['rating'] for h in all_hotels if h['rating'] is not None]
    if prices:
        print(f"  价格: RMB {min(prices)} ~ RMB {max(prices)}, 均价 RMB {sum(prices)/len(prices):.0f}")
    if ratings:
        print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f}, 评分>=9: {sum(1 for r in ratings if r >= 9)}家")


def main():
    print("=" * 55)
    print("  Agoda 武汉酒店全量抓取 (Python API直连)")
    print("=" * 55)
    
    session = requests.Session()
    
    # Price brackets from Agoda page
    brackets = [
        (0, 20), (20, 30), (30, 40), (40, 50), (50, 60), (60, 70), (70, 80),
        (80, 100), (100, 120), (120, 150), (150, 180), (180, 210), (210, 260),
        (260, 310), (310, 380), (380, 460), (460, 550), (550, 670), (670, 810),
        (810, 980), (980, 1180), (1180, 1430), (1430, 1730), (1730, 2100),
        (2100, 100000)  # Everything above 2100
    ]
    
    ALL_HOTELS = []
    SEEN_NAMES = set()
    
    for idx, (frm, to) in enumerate(brackets, 1):
        print(f"\n[{idx}/{len(brackets)}] Price: RMB {frm} - {to}")
        
        hotels = scrape_bracket(session, frm, to)
        
        new_count = 0
        for h in hotels:
            if h['name'] not in SEEN_NAMES:
                SEEN_NAMES.add(h['name'])
                ALL_HOTELS.append(h)
                new_count += 1
        
        print(f"  -> {new_count} new (total: {len(ALL_HOTELS)})")
        
        # Small delay between brackets
        time.sleep(0.5)
    
    print(f"\n{'=' * 55}")
    print(f"  总计: {len(ALL_HOTELS)} 家酒店")
    print(f"{'=' * 55}")
    
    if ALL_HOTELS:
        save_results(ALL_HOTELS)


if __name__ == "__main__":
    main()
