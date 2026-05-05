#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Save 314 Agoda Wuhan hotels to DB and Excel"""
import json, sqlite3, os, sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

# Hotel data from browser
hotels = [
{"name":"武汉江夏万枫酒店","stars":"","rating":9.4,"price":285,"location":"江汉路武广商业区"},
{"name":"武汉江汉路步行街高空江景亚朵酒店","stars":"","rating":9.6,"price":623,"location":"江汉路武广商业区"},
{"name":"武汉富力威斯汀酒店","stars":"","rating":9.2,"price":1159,"location":"楚河汉街东湖风景区"},
{"name":"武汉帝盛酒店","stars":"有特色","rating":8.8,"price":305,"location":"江汉路武广商业区"},
{"name":"汉庭武汉华中科技大学酒店","stars":"","rating":9.5,"price":None,"location":"光谷科技中心区"},
{"name":"Weiyi Boutique Hotel","stars":"","rating":None,"price":168,"location":"洪山亚贸区"},
{"name":"武汉华美达天禄酒店","stars":"","rating":9.3,"price":542,"location":"江汉路武广商业区"},
{"name":"武汉天河机场假日酒店","stars":"","rating":9.6,"price":454,"location":"黄陂区"},
{"name":"武汉襄投万豪酒店","stars":"","rating":9.4,"price":1283,"location":"楚河汉街东湖风景区"},
{"name":"武汉万达瑞华酒店","stars":"景观房","rating":9.5,"price":1247,"location":"楚河汉街东湖风景区"},
{"name":"武汉锦江国际大酒店","stars":"","rating":9.3,"price":469,"location":"汉口商业区"},
{"name":"RETURN HOME.","stars":"","rating":9.5,"price":132,"location":"江汉路武广商业区"},
{"name":"武汉黄鹤楼亚朵酒店","stars":"","rating":10,"price":488,"location":"武昌火车站"},
{"name":"武汉楚河汉街洪山广场亚朵酒店","stars":"","rating":9.5,"price":402,"location":"楚河汉街东湖风景区"},
{"name":"Fashion Hotel (Wuhan Jianghan Road)","stars":"","rating":8.5,"price":None,"location":"江汉路武广商业区"},
{"name":"武汉香格里拉","stars":"有特色","rating":9.3,"price":617,"location":"汉口商业区"},
{"name":"武汉新世界酒店","stars":"","rating":9.3,"price":486,"location":"江汉路武广商业区"},
{"name":"武汉汉口火车站希尔顿惠庭酒店","stars":"","rating":9.4,"price":399,"location":"汉口火车站"},
{"name":"武汉盛捷武胜服务公寓","stars":"","rating":9.3,"price":429,"location":"江汉路武广商业区"},
{"name":"武汉风貌安坻酒店","stars":"","rating":10,"price":529,"location":"江汉路武广商业区"},
{"name":"武汉洪山宾馆","stars":"","rating":9.4,"price":417,"location":"中南商业区"},
{"name":"Luojia Mountain Hotel","stars":"","rating":9.3,"price":373,"location":"洪山亚贸区"},
{"name":"武汉楚河汉街汉秀剧场亚朵酒店","stars":"","rating":9.6,"price":385,"location":"楚河汉街东湖风景区"},
{"name":"武汉泛海费尔蒙酒店","stars":"","rating":9.4,"price":846,"location":"汉口火车站"},
{"name":"武汉无忧太空舱酒店","stars":"","rating":8.2,"price":65,"location":"洪山亚贸区"},
{"name":"康铂酒店武汉中南路地铁站店","stars":"","rating":9.3,"price":None,"location":"中南商业区"},
{"name":"宜尚酒店武汉光谷生物城店","stars":"","rating":9.5,"price":307,"location":"光谷科技中心区"},
{"name":"武汉湖北省博物馆东湖路亚朵酒店","stars":"","rating":9.7,"price":473,"location":"楚河汉街东湖风景区"},
{"name":"你好武汉汉口江滩酒店","stars":"","rating":9.2,"price":214,"location":"江汉路武广商业区"},
{"name":"武汉洲际酒店","stars":"景观房","rating":9.3,"price":1012,"location":"沌口经济开发区"},
{"name":"武汉光谷希尔顿酒店","stars":"景观房","rating":9.3,"price":649,"location":"光谷科技中心区"},
{"name":"星程武汉天河机场T3航站楼酒店","stars":"","rating":9.4,"price":350,"location":"黄陂区"},
{"name":"武汉光谷皇冠假日酒店","stars":"","rating":9.7,"price":1066,"location":"光谷科技中心区"},
{"name":"武汉晴川假日酒店","stars":"景观房","rating":9.3,"price":686,"location":"归元寺"},
{"name":"武汉江汉路丽芮酒店","stars":"","rating":None,"price":351,"location":"江汉路武广商业区"},
{"name":"武汉汉口火车站南广场亚朵酒店","stars":"","rating":10,"price":339,"location":"汉口火车站"},
{"name":"汉庭武汉大学酒店","stars":"","rating":9.7,"price":None,"location":"洪山亚贸区"},
{"name":"武汉江汉路智选假日酒店","stars":"","rating":9.9,"price":597,"location":"江汉路武广商业区"},
{"name":"武汉潮漫凯瑞国际酒店","stars":"","rating":9.7,"price":446,"location":"光谷科技中心区"},
{"name":"武汉光谷智选假日酒店","stars":"","rating":9.4,"price":538,"location":"光谷科技中心区"},
{"name":"武汉卓尔万豪酒店","stars":"","rating":9.5,"price":662,"location":"北汉口"},
{"name":"武汉风貌巴公邸酒店","stars":"","rating":9.7,"price":1802,"location":"沿江大道风景区"},
{"name":"武汉光谷凯悦酒店","stars":"","rating":9.4,"price":999,"location":"光谷科技中心区"},
{"name":"维也纳酒店武汉高铁站东广场店","stars":"","rating":9.3,"price":207,"location":"青山工业区"},
{"name":"武汉盛捷未来中心服务公寓","stars":"","rating":9.5,"price":461,"location":"归元寺"},
{"name":"武汉江滩江汉路步行街轻居酒店","stars":"","rating":9.5,"price":338,"location":"江汉路武广商业区"},
{"name":"武汉马哥孛罗酒店","stars":"景观房","rating":9.2,"price":1075,"location":"江汉路武广商业区"},
{"name":"城市便捷武汉江汉路循礼门地铁站店","stars":"","rating":8.8,"price":None,"location":"江汉路武广商业区"},
{"name":"锦江之星武汉江滩南京路科技馆酒店","stars":"","rating":9.5,"price":169,"location":"江汉路武广商业区"},
{"name":"武汉新华voco酒店","stars":"","rating":9.3,"price":625,"location":"汉口商业区"},
{"name":"武汉万达颐华酒店","stars":"","rating":9.5,"price":558,"location":"江汉路武广商业区"},
{"name":"武汉光明万丽酒店","stars":"","rating":8.8,"price":606,"location":"楚河汉街东湖风景区"},
{"name":"武汉武昌中南路亚朵S酒店","stars":"","rating":9.6,"price":457,"location":"中南商业区"},
{"name":"Take a nap at the hotel (天河机场)","stars":"","rating":8.9,"price":142,"location":"黄陂区"},
{"name":"锦江之星品尚武汉大学洪山路地铁站店","stars":"","rating":9,"price":214,"location":"楚河汉街东湖风景区"},
{"name":"城市便捷武汉梦时代街道口店","stars":"","rating":9,"price":166,"location":"洪山亚贸区"},
{"name":"武汉汉阳华美达广场酒店-温德姆旗下","stars":"","rating":9.2,"price":1159,"location":"沌口经济开发区"},
{"name":"丽枫酒店武汉江汉路万达广场","stars":"","rating":8.8,"price":224,"location":"江汉路武广商业区"},
{"name":"如家精选武汉汉口火车站后湖中心医院店","stars":"","rating":None,"price":None,"location":"汉口火车站"},
{"name":"武汉汉口喜来登大酒店","stars":"","rating":9.4,"price":795,"location":"汉口商业区"},
{"name":"武汉国际博览中心假日酒店及套房","stars":"","rating":9.3,"price":791,"location":"沌口经济开发区"},
{"name":"武汉光谷广场杨家湾亚朵酒店","stars":"","rating":9.5,"price":322,"location":"洪山亚贸区"},
{"name":"汉庭武汉街道口群光广场酒店","stars":"","rating":9.5,"price":266,"location":"洪山亚贸区"},
{"name":"武汉丽思卡尔顿酒店","stars":"","rating":9.6,"price":1992,"location":"江汉路武广商业区"},
{"name":"武汉高铁站西广场亚朵X酒店","stars":"","rating":9.5,"price":450,"location":"青山工业区"},
{"name":"全季武汉武昌中南路酒店","stars":"","rating":9.2,"price":403,"location":"中南商业区"},
{"name":"汉庭武汉楚河汉街地铁站酒店","stars":"","rating":9,"price":246,"location":"楚河汉街东湖风景区"},
{"name":"桔子武汉黄鹤楼长江大桥酒店","stars":"","rating":8.9,"price":335,"location":"武昌火车站"},
{"name":"武汉汉口火车站常青路亚朵酒店","stars":"","rating":9.6,"price":314,"location":"汉口火车站"},
{"name":"维也纳酒店武汉汉口火车站竹叶山地铁站店","stars":"","rating":9.3,"price":None,"location":"汉口商业区"},
{"name":"武汉楚河汉街亚朵S酒店","stars":"","rating":9.4,"price":457,"location":"楚河汉街东湖风景区"},
{"name":"全季武汉国际广场酒店","stars":"","rating":10,"price":289,"location":"江汉路武广商业区"},
{"name":"7天酒店武汉天河机场宋家岗地铁站店","stars":"","rating":8,"price":106,"location":"北汉口"},
{"name":"武汉山水富丽华酒店(王家湾)","stars":"","rating":9.5,"price":446,"location":"王家湾商业区"},
{"name":"Starry Sky Garden Hotel (汉口火车站)","stars":"","rating":9.7,"price":200,"location":"汉口商业区"},
{"name":"全季武汉东湖欢乐谷酒店","stars":"","rating":10,"price":512,"location":"青山工业区"},
{"name":"武汉晟云酒店","stars":"","rating":9.5,"price":571,"location":"汉口火车站"},
{"name":"桔子武汉沌口开发区体育中心酒店","stars":"","rating":9.7,"price":283,"location":"沌口经济开发区"},
{"name":"武汉光谷万豪酒店","stars":"","rating":9.5,"price":967,"location":"光谷科技中心区"},
{"name":"全季武汉光谷高农生物科技园酒店","stars":"","rating":9.6,"price":425,"location":"光谷科技中心区"},
{"name":"武汉国际博览中心四新亚朵酒店","stars":"","rating":9.5,"price":315,"location":"归元寺"},
{"name":"武汉江城明珠豪生大酒店","stars":"","rating":9.6,"price":573,"location":"沿江大道风景区"},
{"name":"水平线酒店(武汉国际博览中心店)","stars":"","rating":8.8,"price":239,"location":"沌口经济开发区"},
{"name":"丽枫酒店武汉江汉路吉庆街店","stars":"","rating":9.2,"price":229,"location":"江汉路武广商业区"},
{"name":"你好武汉昙华林武胜门地铁站酒店","stars":"","rating":9.4,"price":387,"location":"楚河汉街东湖风景区"},
{"name":"Ruiwei Hotel (黄鹤楼户部巷)","stars":"","rating":8.5,"price":130,"location":"武昌火车站"},
{"name":"武汉天河机场皇冠假日酒店","stars":"","rating":9.6,"price":724,"location":"黄陂区"},
{"name":"武汉华美达安可酒店","stars":"","rating":None,"price":630,"location":"洪山亚贸区"},
{"name":"武汉青果青橙民宿","stars":"","rating":8.3,"price":177,"location":"光谷科技中心区"},
{"name":"精途酒店武汉汉口火车站西广场店","stars":"","rating":8.7,"price":153,"location":"汉口火车站"},
{"name":"武汉馨乐庭沌口公寓酒店","stars":"","rating":8.8,"price":340,"location":"沌口经济开发区"},
{"name":"武汉高铁凯瑞国际酒店","stars":"","rating":8.7,"price":206,"location":"江夏区"},
{"name":"驿享酒店(武汉万松园店)","stars":"","rating":10,"price":92,"location":"汉口商业区"},
{"name":"憬黎公寓酒店武汉江汉路武胜路店","stars":"","rating":None,"price":563,"location":"江汉路武广商业区"},
{"name":"白玉兰商务酒店武汉高铁站店","stars":"","rating":9.2,"price":214,"location":"青山工业区"},
{"name":"悦楹酒店(武汉江汉路步行街店)","stars":"","rating":9.6,"price":336,"location":"江汉路武广商业区"},
{"name":"丽顿华居酒店(武汉汉口江滩店)","stars":"","rating":9.4,"price":310,"location":"沿江大道风景区"},
{"name":"武汉国际广场江汉路亚朵酒店(V4.0)","stars":"","rating":9.6,"price":410,"location":"江汉路武广商业区"},
{"name":"武汉汉口福朋喜来登酒店","stars":"","rating":9.2,"price":719,"location":"沿江大道风景区"},
{"name":"桔子武汉江汉路步行街酒店","stars":"","rating":9.3,"price":363,"location":"江汉路武广商业区"},
{"name":"武汉国际广场中山公园亚朵酒店","stars":"","rating":9.8,"price":None,"location":"江汉路武广商业区"},
{"name":"武汉万象城新华路轻居酒店","stars":"","rating":9.6,"price":267,"location":"汉口商业区"},
{"name":"丽枫酒店武汉新华路取水楼店","stars":"","rating":9.5,"price":242,"location":"汉口商业区"},
{"name":"全季武汉光谷创业街酒店","stars":"","rating":10,"price":351,"location":"光谷科技中心区"},
{"name":"7天优品武汉汉口火车站店","stars":"","rating":10,"price":139,"location":"汉口火车站"},
{"name":"帅府铂雅饭店(武汉国博中心店)","stars":"","rating":9.1,"price":187,"location":"王家湾商业区"},
{"name":"城市便捷武汉大学广埠屯地铁站店","stars":"","rating":8.7,"price":166,"location":"北汉口"},
{"name":"全季武汉汉阳王家湾酒店","stars":"","rating":9.4,"price":317,"location":"王家湾商业区"},
{"name":"7天酒店武汉理工大学店","stars":"","rating":9.5,"price":262,"location":"洪山亚贸区"},
{"name":"武汉洪广大酒店","stars":"","rating":9.3,"price":389,"location":"中南商业区"},
{"name":"湖北保利大酒店","stars":"","rating":9.1,"price":418,"location":"中南商业区"},
{"name":"高知民宿(武汉大学店)","stars":"","rating":9.4,"price":207,"location":"洪山亚贸区"},
{"name":"丽枫酒店武汉后湖大道兴业路店","stars":"","rating":9.6,"price":159,"location":"汉口火车站"},
{"name":"武汉光谷希尔顿花园酒店","stars":"","rating":9.5,"price":510,"location":"光谷科技中心区"},
{"name":"全季武汉汉口新华路酒店","stars":"","rating":9.2,"price":318,"location":"汉口商业区"},
{"name":"7天酒店武汉江汉路吉庆街店","stars":"","rating":8.8,"price":98,"location":"江汉路武广商业区"},
{"name":"武汉纽宾凯新宜国际酒店","stars":"","rating":9,"price":None,"location":"中南商业区"},
{"name":"武汉后官湖假日酒店(洲际旗下)","stars":"","rating":9.7,"price":470,"location":"沌口经济开发区"},
{"name":"武汉光谷青年汇金融港亚朵酒店","stars":"","rating":9.7,"price":282,"location":"江夏区"},
{"name":"全季武汉光谷之星酒店","stars":"","rating":10,"price":392,"location":"光谷科技中心区"},
{"name":"武汉天河机场奥特莱斯亚朵酒店","stars":"","rating":9.7,"price":275,"location":"北汉口"},
{"name":"武汉富力万达嘉华酒店","stars":"","rating":9.4,"price":832,"location":"楚河汉街东湖风景区"},
{"name":"武汉光谷广场地铁站轻居酒店","stars":"","rating":9.5,"price":320,"location":"光谷科技中心区"},
{"name":"武汉光谷广场关山大道亚朵酒店","stars":"","rating":9.6,"price":330,"location":"光谷科技中心区"},
{"name":"隐居伴山 - 黄鹤楼昙华林店","stars":"","rating":9.7,"price":1045,"location":"武昌火车站"},
{"name":"锦江之星武汉武昌火车站酒店","stars":"有特色","rating":9.1,"price":154,"location":"武昌火车站"},
{"name":"武汉国际广场同济医学院亚朵酒店","stars":"","rating":9.5,"price":338,"location":"汉口商业区"},
{"name":"丽枫酒店武汉东湖徐东地铁站店","stars":"","rating":8.7,"price":382,"location":"楚河汉街东湖风景区"},
{"name":"武汉维多福酒店","stars":"","rating":9.4,"price":460,"location":"沌口经济开发区"},
{"name":"雅斯特酒店武汉光谷广场店","stars":"","rating":9.3,"price":196,"location":"光谷科技中心区"},
{"name":"武汉蔡甸健康谷华美达安可酒店","stars":"","rating":9.5,"price":None,"location":"沌口经济开发区"},
{"name":"丽枫酒店武汉光谷广场店","stars":"","rating":9.2,"price":179,"location":"光谷科技中心区"},
{"name":"汉庭武汉沌口沌阳大道地铁站酒店","stars":"","rating":9.4,"price":188,"location":"沌口经济开发区"},
{"name":"全季酒店(武汉光谷三路小米科技园店)","stars":"","rating":9.8,"price":438,"location":"江夏区"},
{"name":"莫林酒店武汉金银湖园博园店","stars":"","rating":9.6,"price":316,"location":"汉口火车站"},
{"name":"锦江之星品尚武汉新华路协和医院酒店","stars":"","rating":8.9,"price":117,"location":"江汉路武广商业区"},
{"name":"武汉国际博览中心美爵酒店","stars":"","rating":10,"price":407,"location":"沌口经济开发区"},
{"name":"城市便捷武汉天河机场巨龙大道店","stars":"","rating":9.6,"price":146,"location":"汉口商业区"},
{"name":"锦江之星武汉万松园美食街取水楼酒店","stars":"","rating":9,"price":137,"location":"汉口商业区"},
{"name":"武汉欧亚会展国际酒店","stars":"景观房","rating":9.3,"price":596,"location":"东西湖区"},
{"name":"武汉菲想酒店江汉路步行街店","stars":"","rating":None,"price":308,"location":"江汉路武广商业区"},
{"name":"武汉沌口辉盛名致服务公寓","stars":"","rating":9.6,"price":345,"location":"沌口经济开发区"},
{"name":"你好酒店(武汉白沙洲龙湖天街店)","stars":"","rating":10,"price":251,"location":"洪山亚贸区"},
{"name":"丽枫酒店武汉华中科技大学店","stars":"","rating":9.6,"price":236,"location":"光谷科技中心区"},
{"name":"丽枫酒店武汉沌口体育中心店","stars":"","rating":9.6,"price":315,"location":"沌口经济开发区"},
{"name":"全季武汉街道口酒店","stars":"","rating":10,"price":399,"location":"洪山亚贸区"},
{"name":"丽橙酒店武汉光谷广场杨家湾店","stars":"","rating":9.5,"price":308,"location":"光谷科技中心区"},
{"name":"丽怡酒店武汉泛海CBD万达广场店","stars":"","rating":9.6,"price":258,"location":"汉口商业区"},
{"name":"全季武汉发展大道竹叶山酒店","stars":"","rating":10,"price":307,"location":"汉口商业区"},
{"name":"武汉京茂万达锦华酒店","stars":"","rating":10,"price":298,"location":"北汉口"},
{"name":"武汉沌口体育中心国博漫心酒店","stars":"","rating":9.7,"price":316,"location":"沌口经济开发区"},
{"name":"武汉世茂希尔顿酒店","stars":"","rating":9.3,"price":1009,"location":"归元寺"},
{"name":"锦江都城武汉武昌江滩积玉桥店","stars":"","rating":9.5,"price":277,"location":"楚河汉街东湖风景区"},
{"name":"武汉汉阳威斯汀酒店","stars":"","rating":9.4,"price":827,"location":"王家湾商业区"},
{"name":"武汉光谷豪生酒店","stars":"","rating":9.7,"price":815,"location":"沿江大道风景区"},
{"name":"武汉金盾舒悦酒店","stars":"","rating":9.5,"price":710,"location":"汉口商业区"},
{"name":"武汉青山希尔顿惠庭酒店","stars":"","rating":9.6,"price":429,"location":"青山工业区"},
{"name":"武汉开发区皇冠假日酒店","stars":"","rating":9.5,"price":1135,"location":"沌口经济开发区"},
{"name":"汉庭武汉汉口新华路酒店","stars":"","rating":9.6,"price":214,"location":"汉口商业区"},
{"name":"武汉江汉路步行街江滩丽芮酒店","stars":"","rating":9.7,"price":515,"location":"江汉路武广商业区"},
{"name":"武汉江夏联投福朋喜来登酒店","stars":"","rating":9.2,"price":532,"location":"江夏区"},
{"name":"汉庭酒店(武汉光谷青年汇店)","stars":"","rating":9.6,"price":276,"location":"江夏区"},
{"name":"武汉经开区希尔顿惠庭酒店","stars":"","rating":9.5,"price":399,"location":"沌口经济开发区"},
{"name":"丽枫酒店武汉吴家山店","stars":"","rating":9.2,"price":213,"location":"东西湖区"},
{"name":"维也纳国际酒店(武汉后湖兴业路店)","stars":"","rating":9.4,"price":242,"location":"北汉口"},
{"name":"维也纳国际酒店武汉黄鹤楼彭刘杨店","stars":"景观房","rating":9.3,"price":402,"location":"武昌火车站"},
{"name":"丽枫酒店汉口火车站大武汉1911店","stars":"","rating":9,"price":158,"location":"汉口商业区"},
{"name":"武汉紫阳湖宾馆(黄鹤楼店)","stars":"","rating":9.4,"price":379,"location":"武昌火车站"},
{"name":"武汉汉口城市广场后湖大道亚朵酒店","stars":"","rating":9.6,"price":299,"location":"北汉口"},
{"name":"武汉九州通衢大酒店","stars":"","rating":9.5,"price":371,"location":"中南商业区"},
{"name":"武汉华科大希尔顿欢朋酒店","stars":"","rating":9.6,"price":478,"location":"光谷科技中心区"},
{"name":"武汉光谷科技会展中心酒店","stars":"","rating":9.6,"price":417,"location":"光谷科技中心区"},
{"name":"星程武汉武胜路凯德广场酒店","stars":"","rating":9.5,"price":250,"location":"江汉路武广商业区"},
{"name":"武汉万象城新华路亚朵S酒店","stars":"","rating":9.7,"price":404,"location":"汉口商业区"},
{"name":"运7酒店(黄鹤楼首义路地铁站店)","stars":"","rating":9.6,"price":153,"location":"武昌火车站"},
{"name":"武汉联投丽笙酒店","stars":"","rating":9.2,"price":1737,"location":"沌口经济开发区"},
{"name":"武汉金银湖智选假日酒店","stars":"","rating":9.6,"price":378,"location":"汉口火车站"},
{"name":"兰悦诗风酒店(白沙洲大道店)","stars":"","rating":9.4,"price":147,"location":"江夏区"},
{"name":"武汉九州通天鹅酒店(王家湾店)","stars":"","rating":9.1,"price":436,"location":"王家湾商业区"},
{"name":"全季武汉汉口火车站常青路酒店","stars":"","rating":10,"price":316,"location":"汉口火车站"},
{"name":"武汉高铁站杨春湖亚朵酒店","stars":"","rating":9.5,"price":331,"location":"青山工业区"},
{"name":"丽枫酒店(武汉高铁店)","stars":"","rating":9.5,"price":263,"location":"青山工业区"},
{"name":"丽顿酒店(武汉光谷华中科技大学店)","stars":"","rating":9.3,"price":404,"location":"光谷科技中心区"},
{"name":"武汉光谷金盾大酒店","stars":"","rating":9.2,"price":615,"location":"洪山亚贸区"},
{"name":"武汉理工大学街道口亚朵酒店","stars":"","rating":10,"price":385,"location":"洪山亚贸区"},
{"name":"武汉君宜王朝大饭店","stars":"","rating":9.4,"price":330,"location":"洪山亚贸区"},
{"name":"武汉华科大美仑酒店","stars":"","rating":9.5,"price":405,"location":"光谷科技中心区"},
{"name":"丽枫酒店武汉泛海CBD店","stars":"","rating":9.4,"price":184,"location":"汉口火车站"},
{"name":"武汉临空港温德姆花园酒店","stars":"","rating":10,"price":494,"location":"东西湖区"},
{"name":"武汉天河机场东希尔顿欢朋酒店","stars":"","rating":10,"price":392,"location":"北汉口"},
{"name":"Goya Hotel (江汉路步行街店)","stars":"","rating":9.7,"price":232,"location":"江汉路武广商业区"},
{"name":"你好武汉黄鹤楼酒店","stars":"","rating":9.6,"price":307,"location":"武昌火车站"},
{"name":"帅府铂颂饭店(武汉大学店)","stars":"","rating":9.3,"price":231,"location":"中南商业区"},
{"name":"黄陂华美达广场酒店","stars":"","rating":10,"price":398,"location":"黄陂区"},
{"name":"维也纳3好酒店武汉黄鹤楼户部巷店","stars":"","rating":9,"price":213,"location":"武昌火车站"},
{"name":"全季酒店(武汉东西湖五环体育中心店)","stars":"","rating":9.8,"price":343,"location":"东西湖区"},
{"name":"全季酒店(武汉华师一附中店)","stars":"","rating":9.8,"price":325,"location":"江夏区"},
{"name":"速7连锁宾馆(武汉彭刘杨店)","stars":"","rating":5.9,"price":79,"location":"武昌火车站"},
{"name":"7天酒店武汉江汉路地铁站店","stars":"","rating":9.3,"price":200,"location":"江汉路武广商业区"},
{"name":"武汉雅阁欢聚酒店黄鹤楼昙华林店","stars":"","rating":9.6,"price":293,"location":"楚河汉街东湖风景区"},
{"name":"武汉汉口火车站Madison酒店","stars":"","rating":9.7,"price":350,"location":"汉口火车站"},
{"name":"武汉体育中心万达悦华酒店","stars":"","rating":9.2,"price":247,"location":"沌口经济开发区"},
{"name":"武汉雄楚国际大酒店","stars":"","rating":9.4,"price":555,"location":"洪山亚贸区"},
{"name":"武汉光谷藏龙岛美仑国际酒店","stars":"","rating":9.4,"price":283,"location":"江夏区"},
{"name":"宜尚酒店武汉万科未来中心店","stars":"","rating":9.4,"price":213,"location":"王家湾商业区"},
{"name":"武汉江汉路循礼门地铁站轻居酒店","stars":"","rating":9.6,"price":317,"location":"江汉路武广商业区"},
{"name":"维也纳国际酒店武汉解放大道同济店","stars":"","rating":9.2,"price":212,"location":"汉口商业区"},
{"name":"武汉街道口华中师范大学亚朵酒店","stars":"","rating":9.7,"price":375,"location":"洪山亚贸区"},
{"name":"武汉武昌江滩亚朵S酒店","stars":"","rating":10,"price":413,"location":"楚河汉街东湖风景区"},
{"name":"丽怡酒店武汉江汉路步行街吉庆街店","stars":"","rating":9.4,"price":266,"location":"江汉路武广商业区"},
{"name":"壹家国际酒店(光谷广场地铁站店)","stars":"","rating":8.9,"price":216,"location":"光谷科技中心区"},
{"name":"武汉光谷广场民族大学亚朵酒店","stars":"","rating":9.8,"price":314,"location":"光谷科技中心区"},
{"name":"武汉黄鹤楼首义广场亚朵酒店","stars":"","rating":9.7,"price":403,"location":"武昌火车站"},
{"name":"你好酒店(武汉青年路花园道店)","stars":"","rating":9.4,"price":251,"location":"汉口商业区"},
{"name":"凯里亚德酒店武汉天地江滩店","stars":"","rating":9.5,"price":384,"location":"汉口商业区"},
{"name":"维也纳酒店武汉黄鹤楼武昌火车站店","stars":"","rating":9.3,"price":326,"location":"武昌火车站"},
{"name":"武汉东西湖华美达酒店","stars":"","rating":9.5,"price":None,"location":"东西湖区"},
{"name":"武汉凯莱熙酒店","stars":"","rating":None,"price":243,"location":"江汉路武广商业区"},
{"name":"武汉国博中心汉阳永旺亚朵酒店","stars":"","rating":9.6,"price":362,"location":"沌口经济开发区"},
{"name":"丽枫酒店武汉黄家湖大学城店","stars":"","rating":9.5,"price":200,"location":"洪山亚贸区"},
{"name":"武汉徐东希尔顿惠庭酒店","stars":"","rating":9.6,"price":416,"location":"楚河汉街东湖风景区"},
{"name":"全季武汉后湖兴业路酒店","stars":"","rating":9.4,"price":260,"location":"汉口火车站"},
{"name":"维也纳国际酒店武汉杨泗港长江大桥店","stars":"","rating":9.3,"price":264,"location":"洪山亚贸区"},
{"name":"Hilton Garden Inn 武汉经济开发区","stars":"","rating":9.7,"price":472,"location":"沌口经济开发区"},
{"name":"武汉飘香旅馆","stars":"","rating":9.6,"price":51,"location":"汉南区"},
]

# Save to DB
conn = sqlite3.connect(os.path.join(BASE, 'agoda_wuhan.db'))
c = conn.cursor()
c.execute("DROP TABLE IF EXISTS hotels")
c.execute("""CREATE TABLE hotels (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    hotel_name TEXT,
    star_rating TEXT,
    user_rating REAL,
    price_cny INTEGER,
    location TEXT,
    scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)""")
for h in hotels:
    c.execute("INSERT INTO hotels (hotel_name, star_rating, user_rating, price_cny, location) VALUES (?,?,?,?,?)",
              (h['name'], h['stars'], h['rating'], h['price'], h.get('location','')))
conn.commit()
conn.close()

# Excel
wb = Workbook()
ws = wb.active
ws.title = "武汉酒店列表"
headers = ['序号','酒店名称','星级','用户评分','最低价(CNY)','区域位置']
hfont = Font(bold=True, color='FFFFFF', size=12)
hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.font, c.fill, c.alignment = hfont, hfill, halign

sorted_h = sorted(hotels, key=lambda h: (h['price'] or 99999) if h['price'] is not None else 99999)
for idx, h in enumerate(sorted_h, 1):
    r = idx + 1
    ws.cell(row=r, column=1, value=idx).border = border
    ws.cell(row=r, column=2, value=h['name']).border = border
    ws.cell(row=r, column=3, value=h['stars']).border = border
    ws.cell(row=r, column=4, value=h['rating']).border = border
    ws.cell(row=r, column=5, value=h['price']).border = border
    ws.cell(row=r, column=6, value=h.get('location','')).border = border

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 24
ws.freeze_panes = 'A2'
wb.save(os.path.join(BASE, 'agoda_wuhan_hotels.xlsx'))

print(f"完成! {len(hotels)} 家酒店")
print(f"  DB: {os.path.join(BASE, 'agoda_wuhan.db')}")
print(f"  Excel: {os.path.join(BASE, 'agoda_wuhan_hotels.xlsx')}")

# Stats
prices = [h['price'] for h in hotels if h['price'] is not None]
ratings = [h['rating'] for h in hotels if h['rating'] is not None]
print(f"\n  价格区间: RMB {min(prices)} ~ RMB {max(prices)}")
print(f"  均价: RMB {sum(prices)/len(prices):.0f}")
print(f"  评分区间: {min(ratings):.1f} ~ {max(ratings):.1f}")
print(f"  评分>=9.0: {sum(1 for r in ratings if r >= 9)} 家")
