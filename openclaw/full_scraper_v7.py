#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店全量抓取 v7 - 价格分段 + 正确 filter 格式
终于找到正确的价格筛选格式: [{"filterKey": "Price", "ranges": [{"from": X, "to": Y}]}]
策略：超细价格段（每10-50元一段）+ 完整翻页 + 去重合并
"""
import json, urllib.request, sys, os, sqlite3, time
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE, 'api_body.json'), 'r') as f:
    TEMPLATE = json.load(f)

HEADERS = {
    'Content-Type': 'application/json',
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'AG-LANGUAGE-LOCALE': 'zh-cn',
    'AG-CID': '-1',
    'AG-PAGE-TYPE-ID': '103',
    'AG-REQUEST-ATTEMPT': '1',
}

from datetime import datetime
booking_date = datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000') + 'Z'

# 超细价格分段
BRACKETS = []
# 0-500: 每10元
for i in range(0, 500, 10):
    BRACKETS.append((i, i + 10 + 0.49))  # 0.49重叠确保不漏
# 500-1000: 每25元
for i in range(500, 1000, 25):
    BRACKETS.append((i, i + 25 + 0.49))
# 1000-2000: 每50元
for i in range(1000, 2000, 50):
    BRACKETS.append((i, i + 50 + 0.49))
# 2000-5000: 每100元
for i in range(2000, 5000, 100):
    BRACKETS.append((i, i + 100 + 0.49))
# 兜底
BRACKETS.append((5000, 99999))

print("=" * 60)
print(f"  Agoda 武汉酒店全量抓取 v7")
print(f"  价格分段数: {len(BRACKETS)}")
print(f"  目标: 约 7000+ 酒店")
print("=" * 60)


def call_api(frm, to, page=1, token=None):
    req = json.loads(json.dumps(TEMPLATE))
    sr = req['variables']['CitySearchRequest']['searchRequest']
    sr['searchCriteria']['bookingDate'] = booking_date
    sr['page'] = {'pageSize': 50, 'pageNumber': page}
    if token:
        sr['page']['pageToken'] = token
    # 正确的价格筛选格式
    sr['filterRequest']['rangeFilters'] = [{
        'filterKey': 'Price',
        'ranges': [{'from': float(frm), 'to': float(to)}]
    }]
    sr['filterRequest']['idsFilters'] = []
    data = json.dumps(req).encode('utf-8')
    r = urllib.request.Request('https://www.agoda.cn/graphql/search', data=data, headers=HEADERS)
    return json.loads(urllib.request.urlopen(r, timeout=30).read().decode())


def extract(prop):
    if not prop or not prop.get('content'):
        return None
    info = prop['content'].get('informationSummary') or {}
    reviews = prop['content'].get('reviews', {})
    pricing = prop.get('pricing') or {}
    name = info.get('localeName') or info.get('defaultName') or ''
    if not name:
        return None
    price = None
    try:
        for o in (pricing.get('offers') or []):
            for r in (o.get('roomOffers') or []):
                for p in (r.get('room', {}).get('pricing') or []):
                    d = p.get('price', {}).get('perBook', {}).get('inclusive', {}).get('display')
                    if d:
                        price = round(d)
                        break
                if price:
                    break
            if price:
                break
    except Exception:
        pass
    rating = None
    try:
        rating = reviews.get('cumulative', {}).get('score')
    except Exception:
        pass
    location = ''
    addr = info.get('address') or {}
    if addr.get('area'):
        location = addr['area'].get('name', '')
    if not location and addr.get('city'):
        location = addr['city'].get('name', '')
    return {'name': name, 'rating': rating, 'price': price, 'location': location}


# === 主循环 ===
ALL = []
SEEN = set()
total_api_reported = 0

for idx, (frm, to) in enumerate(BRACKETS, 1):
    sys.stdout.write(f"[{idx}/{len(BRACKETS)}] RMB {frm:.0f}-{to:.0f}: ")
    sys.stdout.flush()

    new_count = 0
    pages = 0
    page = 1
    token = None
    retries = 0
    bracket_api_count = 0

    while pages < 200:
        try:
            j = call_api(frm, to, page, token)
            props = j.get('data', {}).get('citySearch', {}).get('properties') or []
            si = j.get('data', {}).get('citySearch', {}).get('searchResult', {}).get('searchInfo', {})
            token = j.get('data', {}).get('citySearch', {}).get('searchEnrichment', {}).get('pageToken')

            if pages == 0:
                bracket_api_count = si.get('totalFilteredHotels', 0)
                if bracket_api_count > total_api_reported:
                    total_api_reported = bracket_api_count

            if not props:
                break  # 无结果，直接退出该价格段

            retries = 0
            for p in props:
                h = extract(p)
                if h and h['name'] not in SEEN:
                    SEEN.add(h['name'])
                    ALL.append(h)
                    new_count += 1

            pages += 1

            if not token:
                break
            page += 1
            time.sleep(0.03)

        except Exception as e:
            retries += 1
            if retries > 3:
                break
            time.sleep(2)

    sys.stdout.write(f"+{new_count}新 ({pages}p)\n")
    sys.stdout.flush()

print(f"\n\n{'=' * 60}")
print(f"  API 报告总数: {total_api_reported}")
print(f"  实际抓取去重: {len(ALL)} 家酒店")
print(f"{'=' * 60}")

if not ALL:
    print("无数据！")
    sys.exit(1)

# === 保存 JSON ===
out_path = os.path.join(BASE, 'hotels_all.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump({'count': len(ALL), 'hotels': ALL}, f, ensure_ascii=False, indent=2)
print(f"JSON: {out_path}")

# === 保存 Excel ===
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    xlsx_path = os.path.join(BASE, f'agoda_wuhan_hotels_{len(ALL)}条.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ['序号', '酒店名称', '用户评分', '最低价(CNY)', '区域位置']
    hfont = Font(bold=True, color='FFFFFF', size=12)
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    halign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font, c.fill, c.alignment = hfont, hfill, halign
    sorted_h = sorted(ALL, key=lambda x: (x['price'] or 99999) if x['price'] is not None else 99999)
    for i, h in enumerate(sorted_h, 1):
        r = i + 1
        ws.cell(row=r, column=1, value=i).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border
        ws.cell(row=r, column=3, value=h['rating']).border = border
        ws.cell(row=r, column=4, value=h['price']).border = border
        ws.cell(row=r, column=5, value=h.get('location', '')).border = border
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.freeze_panes = 'A2'
    wb.save(xlsx_path)
    print(f"Excel: {xlsx_path}")
except Exception as e:
    print(f"Excel失败: {e}")

# === 保存 SQLite ===
try:
    db_path = os.path.join(BASE, 'agoda_wuhan.db')
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""CREATE TABLE hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT, user_rating REAL, price_cny INTEGER, location TEXT
    )""")
    for h in ALL:
        c.execute("INSERT INTO hotels (hotel_name, user_rating, price_cny, location) VALUES (?,?,?,?)",
                  (h['name'], h['rating'], h['price'], h.get('location', '')))
    conn.commit()
    conn.close()
    print(f"DB: {db_path}")
except Exception as e:
    print(f"DB失败: {e}")

# === 统计 ===
prices = [h['price'] for h in ALL if h['price'] is not None]
ratings = [h['rating'] for h in ALL if h['rating'] is not None]
print(f"\n{'=' * 60}")
print(f"  完成！{len(ALL)} 家酒店")
if prices:
    print(f"  价格区间: RMB {min(prices):,} ~ RMB {max(prices):,}")
    print(f"  平均价格: RMB {sum(prices)/len(prices):.0f}")
if ratings:
    print(f"  评分区间: {min(ratings):.1f} ~ {max(ratings):.1f}")
    print(f"  评分>=9.0: {sum(1 for r in ratings if r >= 9)}/{len(ratings)}")
print(f"{'=' * 60}")
