#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店抓取工具
- 入住: 2026-06-05, 1晚, 1间, 2成人
- 币种: CNY
- 保存到 SQLite + 导出 Excel
"""
import json
import os
import sys
import time
import sqlite3
from datetime import datetime, timedelta

sys.stdout.reconfigure(encoding='utf-8')

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))

# ============ 配置 ============
CITY_ID = 5818          # 武汉
CITY_NAME = "武汉"
CHECK_IN = "2026-06-05"
LOS = 1                 # 1晚
ROOMS = 1
ADULTS = 2
CHILDREN = 0
CURRENCY = "CNY"
PAGE_SIZE = 50
# 延迟控制(秒) - 抓取频率不要太快
PAGE_DELAY = 1.0
BRACKET_DELAY = 1.5

API_URL = "https://www.agoda.cn/graphql/search"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    ),
    "Content-Type": "application/json",
    "Accept": "*/*",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Origin": "https://www.agoda.cn",
    "Referer": "https://www.agoda.cn/",
    "AG-LANGUAGE-LOCALE": "zh-cn",
    "AG-CID": "-1",
    "AG-PAGE-TYPE-ID": "103",
    "AG-REQUEST-ATTEMPT": "1",
}

COOKIES = {
    "agoda.user.03": "UserId=83133924-5561-46c1-979d-dd87b2787ff1",
    "agoda.prius": "PriusID=0&PointsMaxTraffic=Agoda",
}


def make_request_body(
    city_id, check_in, los, rooms, adults, children,
    currency="CNY", page_size=50, page_number=1, page_token=None,
    price_from=None, price_to=None,
):
    """构建 GraphQL 请求体"""
    checkout = (
        datetime.strptime(check_in, "%Y-%m-%d") + timedelta(days=los)
    ).strftime("%Y-%m-%d")

    body = {
        "operationName": "citySearch",
        "variables": {
            "CitySearchRequest": {
                "cityId": city_id,
                "searchRequest": {
                    "searchCriteria": {
                        "bookingDate": datetime.now().strftime("%Y-%m-%dT05:00:00.000Z"),
                        "checkInDate": f"{check_in}T00:00:00.000Z",
                        "los": los,
                        "rooms": rooms,
                        "adults": adults,
                        "children": children,
                        "childAges": [],
                        "ratePlans": [1],
                        "featureFlagRequest": {
                            "fiveStarDealOfTheDay": True,
                            "isAllowBookOnRequest": False,
                            "showUnAvailable": True,
                            "showRemainingProperties": True,
                            "isMultiHotelSearch": False,
                            "enableAgencySupplyForPackages": True,
                            "flags": [
                                {"feature": "FamilyChildFriendlyPopularFilter", "enable": True},
                                {"feature": "FamilyChildFriendlyPropertyTypeFilter", "enable": True},
                                {"feature": "FamilyMode", "enable": False},
                            ],
                            "isFlexibleMultiRoomSearch": False,
                            "enablePageToken": True,
                        },
                        "isUserLoggedIn": False,
                        "currency": currency,
                        "travellerType": "Couple",
                        "isAPSPeek": False,
                        "enableOpaqueChannel": False,
                        "sorting": {"sortField": "Ranking", "sortOrder": "Desc"},
                        "requiredBasis": "PRPN",
                        "requiredPrice": "AllInclusive",
                        "suggestionLimit": 0,
                        "synchronous": False,
                        "isRoomSuggestionRequested": False,
                        "isAPORequest": False,
                        "hasAPOFilter": False,
                        "isAllowBookOnRequest": True,
                        "localCheckInDate": check_in,
                    },
                    "searchContext": {
                        "locale": "zh-cn",
                        "cid": -1,
                        "origin": "CN",
                        "platform": 4,
                        "deviceTypeId": 1,
                        "experiments": {"forceByExperiment": []},
                        "isRetry": False,
                        "showCMS": False,
                        "storeFrontId": 3,
                        "pageTypeId": 103,
                        "endpointSearchType": "CitySearch",
                    },
                    "filterRequest": {
                        "idsFilters": [],
                        "rangeFilters": [],
                        "textFilters": [],
                    },
                    "matrixGroup": [
                        {"matrixGroup": "StarRating", "size": 100},
                        {"matrixGroup": "AccommodationType", "size": 100},
                        {"matrixGroup": "HotelAreaId", "size": 100},
                        {"matrixGroup": "HotelFacilities", "size": 100},
                        {"matrixGroup": "ReviewScore", "size": 100},
                        {"matrixGroup": "PaymentOptions", "size": 100},
                        {"matrixGroup": "RoomBenefits", "size": 100},
                        {"matrixGroup": "CityCenterDistance", "size": 100},
                        {"matrixGroup": "RoomAmenities", "size": 100},
                        {"matrixGroup": "GroupedBedTypes", "size": 100},
                        {"matrixGroup": "NumberOfBedrooms", "size": 100},
                        {"matrixGroup": "LandmarkIds", "size": 100},
                        {"matrixGroup": "ReviewLocationScore", "size": 100},
                        {"matrixGroup": "BeachAccessTypeIds", "size": 100},
                    ],
                    "page": {
                        "pageSize": page_size,
                        "pageNumber": page_number,
                    },
                    "searchHistory": [],
                    "isTrimmedResponseRequested": False,
                    "extraHotels": {
                        "extraHotelIds": [],
                        "enableFiltersForExtraHotels": False,
                    },
                    "highlyRatedAgodaHomesRequest": {
                        "numberOfAgodaHomes": 30,
                        "minimumReviewScore": 7.5,
                        "minimumReviewCount": 3,
                        "accommodationTypes": [28, 29, 30, 102, 103, 106, 107, 108, 109, 110, 114, 115, 120, 131],
                        "sortVersion": 0,
                    },
                    "featuredPulsePropertiesRequest": {
                        "numberOfPulseProperties": 15,
                    },
                    "rankingRequest": {
                        "isNhaKeywordSearch": False,
                        "isPulseRankingBoost": False,
                    },
                    "searchDetailRequest": {
                        "priceHistogramBins": 30,
                    },
                },
            }
        },
    }

    # 价格区间筛选
    if price_from is not None or price_to is not None:
        rf = {}
        if price_from is not None:
            rf["from"] = price_from
        if price_to is not None:
            rf["to"] = price_to
        rf["type"] = "Price"
        rf["currencyCode"] = currency
        body["variables"]["CitySearchRequest"]["searchRequest"]["filterRequest"]["rangeFilters"] = [rf]

    # 分页 token
    if page_token:
        body["variables"]["CitySearchRequest"]["searchRequest"]["page"]["pageToken"] = page_token

    return body


def extract_hotel(prop):
    """从 API 返回的 property 对象中提取酒店信息"""
    if not prop or not prop.get("content"):
        return None

    info = prop["content"].get("informationSummary") or {}
    reviews = prop["content"].get("reviews", {})
    pricing = prop.get("pricing") or {}

    name = info.get("localeName") or info.get("defaultName") or ""
    if not name:
        return None

    # --- 地址 ---
    address = ""
    area_name = ""
    addr = info.get("address") or {}
    city_name = addr.get("city", {}).get("name", "")
    area_name = addr.get("area", {}).get("name", "")
    country_name = addr.get("country", {}).get("name", "")
    parts = [p for p in [area_name, city_name, country_name] if p]
    address = " ".join(parts)

    # --- 用户评分 ---
    rating = None
    try:
        rating = reviews.get("cumulative", {}).get("score")
    except Exception:
        pass

    # --- 星级(酒店星级, 非用户评分) ---
    star_rating = None
    try:
        sr = info.get("rating", 0)
        if sr and sr > 0:
            star_rating = round(sr)
    except Exception:
        pass

    # --- 最低价 ---
    price = None
    try:
        for offer in (pricing.get("offers") or []):
            for ro in (offer.get("roomOffers") or []):
                for pr in (ro.get("room", {}).get("pricing") or []):
                    d = pr.get("price", {}).get("perBook", {}).get("inclusive", {}).get("display")
                    if d:
                        price = round(d)
                        break
                if price:
                    break
            if price:
                break
    except Exception:
        pass

    # --- 评价数 ---
    review_count = 0
    try:
        review_count = reviews.get("cumulative", {}).get("reviewCount", 0)
    except Exception:
        pass

    # --- 酒店类型 ---
    acc_type = info.get("accommodationType")

    return {
        "name": name,
        "address": address,
        "area": area_name,
        "rating": rating,
        "star_rating": star_rating,
        "price": price,
        "review_count": review_count,
        "acc_type": acc_type,
    }


def scrape_bracket(session, from_price=None, to_price=None):
    """抓取指定价格区间内的所有酒店(含分页)"""
    all_hotels = []
    seen = set()
    page = 1
    token = None
    max_pages = 200

    label = f"{from_price or 0}-{to_price or 'max'}"
    print(f"  区间 ¥{label}...", end=" ", flush=True)

    while page <= max_pages:
        body = make_request_body(
            city_id=CITY_ID,
            check_in=CHECK_IN,
            los=LOS,
            rooms=ROOMS,
            adults=ADULTS,
            children=CHILDREN,
            currency=CURRENCY,
            page_size=PAGE_SIZE,
            page_number=page,
            page_token=token,
            price_from=from_price,
            price_to=to_price,
        )

        try:
            resp = session.post(API_URL, json=body, headers=HEADERS, cookies=COOKIES, timeout=30)

            if resp.status_code != 200:
                if page > 1:
                    break
                page += 1
                continue

            data = resp.json()
            city_search = data.get("data", {}).get("citySearch", {})
            properties = city_search.get("properties") or []
            token = city_search.get("searchEnrichment", {}).get("pageToken")

            if not properties:
                if page > 1:
                    break
                page += 1
                continue

            for prop in properties:
                if prop.get("propertyResultType") != "NormalProperty":
                    continue
                h = extract_hotel(prop)
                if h and h["name"] not in seen:
                    seen.add(h["name"])
                    all_hotels.append(h)

            if not token:
                break

            page += 1
            time.sleep(PAGE_DELAY)  # 抓取频率不要太快

        except Exception as e:
            if page > 2:
                break
            page += 1
            time.sleep(3)

    print(f"{len(all_hotels)} 家")
    return all_hotels


def save_to_sqlite(hotels, db_path):
    """保存到 SQLite 数据库"""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute("DROP TABLE IF EXISTS hotels")
    c.execute("""
        CREATE TABLE hotels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hotel_name TEXT NOT NULL,
            address TEXT,
            area TEXT,
            user_rating REAL,
            star_rating INTEGER,
            price_cny INTEGER,
            review_count INTEGER,
            acc_type INTEGER,
            scraped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    for h in hotels:
        c.execute(
            """
            INSERT INTO hotels (hotel_name, address, area, user_rating, star_rating, price_cny, review_count, acc_type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                h["name"],
                h["address"],
                h["area"],
                h["rating"],
                h["star_rating"],
                h["price"],
                h["review_count"],
                h.get("acc_type"),
            ),
        )
    conn.commit()
    conn.close()
    print(f"  SQLite: {db_path} ({len(hotels)} 条记录)")


def save_to_excel(hotels, xlsx_path):
    """导出到格式化的 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{CITY_NAME}酒店列表"

    headers = ["序号", "酒店名称", "地址", "区域", "用户评分", "星级", "最低价(CNY)", "评价数"]
    col_widths = [6, 42, 40, 22, 10, 8, 14, 10]

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 表头
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
        c.border = border

    # 按价格排序(无价格的排在最后)
    sorted_hotels = sorted(
        hotels,
        key=lambda x: (x["price"] or 999999) if x["price"] is not None else 999999,
    )

    for idx, h in enumerate(sorted_hotels, 1):
        r = idx + 1
        ws.cell(row=r, column=1, value=idx).border = border
        ws.cell(row=r, column=1, alignment=Alignment(horizontal="center"))
        ws.cell(row=r, column=2, value=h["name"]).border = border
        ws.cell(row=r, column=3, value=h["address"]).border = border
        ws.cell(row=r, column=4, value=h["area"]).border = border

        rating_cell = ws.cell(row=r, column=5)
        if h["rating"] is not None:
            rating_cell.value = h["rating"]
            rating_cell.number_format = "0.0"
        else:
            rating_cell.value = "暂无"
        rating_cell.border = border
        rating_cell.alignment = Alignment(horizontal="center")

        star_cell = ws.cell(row=r, column=6)
        if h["star_rating"] is not None:
            star_cell.value = f"{h['star_rating']}星"
        else:
            star_cell.value = ""
        star_cell.border = border
        star_cell.alignment = Alignment(horizontal="center")

        price_cell = ws.cell(row=r, column=7)
        if h["price"] is not None:
            price_cell.value = h["price"]
            price_cell.number_format = "¥#,##0"
        else:
            price_cell.value = ""
        price_cell.border = border
        price_cell.alignment = Alignment(horizontal="right")

        review_cell = ws.cell(row=r, column=8)
        if h["review_count"] and h["review_count"] > 0:
            review_cell.value = h["review_count"]
        else:
            review_cell.value = ""
        review_cell.border = border
        review_cell.alignment = Alignment(horizontal="center")

    # 列宽
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(hotels) + 1}"

    wb.save(xlsx_path)
    print(f"  Excel: {xlsx_path}")


def print_stats(hotels):
    """打印统计信息"""
    prices = [h["price"] for h in hotels if h["price"] is not None]
    ratings = [h["rating"] for h in hotels if h["rating"] is not None]
    stars = [h["star_rating"] for h in hotels if h["star_rating"] is not None]

    print(f"\n{'=' * 50}")
    print(f"  抓取完成! 共 {len(hotels)} 家酒店")
    print(f"{'=' * 50}")

    if prices:
        print(f"  价格区间: ¥{min(prices)} ~ ¥{max(prices)}")
        print(f"  平均价格: ¥{sum(prices) / len(prices):.0f}")
        print(f"  有价格的酒店: {len(prices)}/{len(hotels)}")

    if ratings:
        print(f"  评分区间: {min(ratings):.1f} ~ {max(ratings):.1f}")
        print(f"  平均评分: {sum(ratings) / len(ratings):.1f}")
        print(f"  评分≥9.0: {sum(1 for r in ratings if r >= 9.0)} 家")

    if stars:
        print(f"  有星级标注: {len(stars)} 家")
        star_dist = {}
        for s in stars:
            star_dist[s] = star_dist.get(s, 0) + 1
        for s in sorted(star_dist.keys(), reverse=True):
            print(f"    {s}星: {star_dist[s]} 家")


def main():
    print("=" * 55)
    print(f"  Agoda {CITY_NAME} 酒店抓取")
    print(f"  入住: {CHECK_IN}, {LOS}晚, {ROOMS}间, {ADULTS}成人")
    print(f"  币种: {CURRENCY}")
    print("=" * 55)

    session = requests.Session()

    # Agoda 价格区间(根据平台实际分段)
    brackets = [
        (0, 50),
        (50, 80),
        (80, 120),
        (120, 180),
        (180, 260),
        (260, 380),
        (380, 550),
        (550, 810),
        (810, 1180),
        (1180, 1730),
        (1730, 2600),
        (2600, 5000),
        (5000, 999999),
    ]

    all_hotels = []
    seen_names = set()

    for idx, (frm, to) in enumerate(brackets, 1):
        print(f"\n[{idx}/{len(brackets)}]")
        hotels = scrape_bracket(session, frm, to)

        new_count = 0
        for h in hotels:
            if h["name"] not in seen_names:
                seen_names.add(h["name"])
                all_hotels.append(h)
                new_count += 1

        print(f"  => 新增 {new_count} 家 (累计 {len(all_hotels)})")

        # 区间间延迟
        time.sleep(BRACKET_DELAY)

    if not all_hotels:
        print("\n未抓取到任何酒店数据。")
        return

    print_stats(all_hotels)

    # 保存到 SQLite
    db_path = os.path.join(BASE, "agoda_wuhan.db")
    save_to_sqlite(all_hotels, db_path)

    # 导出 Excel
    xlsx_path = os.path.join(BASE, f"Agoda_{CITY_NAME}酒店_{CHECK_IN}.xlsx")
    save_to_excel(all_hotels, xlsx_path)

    print(f"\n完成! 数据已保存到数据库和Excel。")


if __name__ == "__main__":
    main()
