#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Agoda 武汉酒店数据抓取 - Selenium 获取 Cookie + GraphQL API 直连
抓取全部酒店并保存到 SQLite + Excel
"""
import json, os, time, sqlite3, sys, re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
CITY = "武汉"
CHECKIN = (datetime.now() + timedelta(days=35)).strftime("%Y-%m-%d")
CHECKOUT = (datetime.now() + timedelta(days=36)).strftime("%Y-%m-%d")

DB_PATH = os.path.join(BASE, "hotels.db")
EXCEL_PATH = os.path.join(BASE, f"hotels_{CITY}.xlsx")
GRAPHQL_URL = "https://www.agoda.cn/graphql/search"

# ========== Selenium: Get Fresh Cookies ==========

def get_fresh_cookies_and_headers():
    """Open Agoda search page, get fresh cookies and a valid Referer."""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service

    print(">> 启动浏览器获取 Cookie ...")
    opts = Options()
    opts.binary_location = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1280,800")
    opts.add_argument("--lang=zh-CN")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/131.0.0.0 Safari/537.36")

    chromedriver = os.path.join(BASE, "chromedriver.exe")
    service = Service(chromedriver)
    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

    search_url = (f"https://www.agoda.cn/search?city=16958"
                  f"&checkIn={CHECKIN}&checkOut={CHECKOUT}"
                  f"&rooms=1&adults=2&currency=CNY")

    try:
        # First visit homepage
        driver.get("https://www.agoda.cn/")
        time.sleep(4)

        # Dismiss popups
        for script in [
            "document.querySelectorAll('[class*=close],[class*=dismiss]').forEach(e=>e.click())",
            "document.querySelectorAll('[aria-label*=Close]').forEach(e=>e.click())",
        ]:
            try: driver.execute_script(script)
            except: pass
        time.sleep(2)

        # Navigate to search page
        driver.get(search_url)
        time.sleep(8)

        cookies_list = driver.get_cookies()
        cookies = {c['name']: c['value'] for c in cookies_list}
        print(f">> 获取到 {len(cookies)} 个 Cookie")

        # Also get any useful data from localStorage
        try:
            ls = driver.execute_script(
                "return JSON.stringify(window.localStorage);")
            print(f">> localStorage 条目数: {len(json.loads(ls)) if ls else 0}")
        except:
            pass

        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
            'Content-Type': 'application/json',
            'Accept': '*/*',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Origin': 'https://www.agoda.cn',
            'Referer': search_url,
            'AG-LANGUAGE-LOCALE': 'zh-cn',
            'AG-CID': '-1',
            'AG-PAGE-TYPE-ID': '103',
            'AG-REQUEST-ATTEMPT': '1',
            'AG-DEBUG-OVERRIDE-ORIGIN': 'CN',
            'sec-ch-ua': '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
        }

        return cookies, headers, driver
    except Exception as e:
        driver.quit()
        raise e


# ========== GraphQL Request Builder ==========

def build_payload(city_id, checkin, checkout, los, rooms, adults, currency,
                  page_size=50, page_number=1, page_token=None,
                  price_from=None, price_to=None):
    """Build the GraphQL request body for Agoda city search."""

    checkin_str = f"{checkin}T00:00:00.000Z"
    checkout_str = f"{checkout}T00:00:00.000Z"

    body = {
        "operationName": "citySearch",
        "variables": {
            "CitySearchRequest": {
                "cityId": city_id,
                "searchRequest": {
                    "searchCriteria": {
                        "bookingDate": datetime.utcnow().strftime("%Y-%m-%dT05:00:00.000Z"),
                        "checkInDate": checkin_str,
                        "los": los,
                        "rooms": rooms,
                        "adults": adults,
                        "children": 0,
                        "childAges": [],
                        "ratePlans": [1],
                        "featureFlagRequest": {
                            "fiveStarDealOfTheDay": True,
                            "isAllowBookOnRequest": False,
                            "showUnAvailable": True,
                            "showRemainingProperties": True,
                            "isMultiHotelSearch": False,
                            "enableAgencySupplyForPackages": True,
                            "flags": [
                                {"feature": "FamilyChildFriendlyPopularFilter", "enable": True},
                                {"feature": "FamilyChildFriendlyPropertyTypeFilter", "enable": True},
                                {"feature": "FamilyMode", "enable": False}
                            ],
                            "isFlexibleMultiRoomSearch": False,
                            "enablePageToken": True
                        },
                        "isUserLoggedIn": False,
                        "currency": currency,
                        "travellerType": "Couple",
                        "isAPSPeek": False,
                        "enableOpaqueChannel": False,
                        "sorting": {"sortField": "Ranking", "sortOrder": "Desc"},
                        "requiredBasis": "PRPN",
                        "requiredPrice": "AllInclusive",
                        "suggestionLimit": 0,
                        "synchronous": False,
                        "isRoomSuggestionRequested": False,
                        "isAPORequest": False,
                        "hasAPOFilter": False,
                        "isAllowBookOnRequest": True,
                        "localCheckInDate": checkin
                    },
                    "searchContext": {
                        "locale": "zh-cn",
                        "cid": -1,
                        "origin": "CN",
                        "platform": 4,
                        "deviceTypeId": 1,
                        "experiments": {"forceByExperiment": []},
                        "isRetry": False,
                        "showCMS": False,
                        "storeFrontId": 3,
                        "pageTypeId": 103,
                        "endpointSearchType": "CitySearch"
                    },
                    "filterRequest": {
                        "idsFilters": [],
                        "rangeFilters": [],
                        "textFilters": []
                    },
                    "matrixGroup": [
                        {"matrixGroup": "StarRating", "size": 100},
                        {"matrixGroup": "AccommodationType", "size": 100},
                        {"matrixGroup": "HotelAreaId", "size": 100},
                    ],
                    "page": {"pageSize": page_size, "pageNumber": page_number},
                    "searchHistory": [],
                    "isTrimmedResponseRequested": False,
                    "extraHotels": {"extraHotelIds": [], "enableFiltersForExtraHotels": False},
                    "highlyRatedAgodaHomesRequest": {
                        "numberOfAgodaHomes": 30, "minimumReviewScore": 7.5, "minimumReviewCount": 3,
                        "accommodationTypes": [28, 29, 30, 102, 103, 106, 107, 108, 109, 110, 114, 115, 120, 131],
                        "sortVersion": 0
                    },
                    "featuredPulsePropertiesRequest": {"numberOfPulseProperties": 15},
                    "rankingRequest": {"isNhaKeywordSearch": False, "isPulseRankingBoost": False},
                    "searchDetailRequest": {"priceHistogramBins": 30}
                }
            },
            "ContentSummaryRequest": {
                "context": {
                    "rawUserId": "",
                    "memberId": 0,
                    "userOrigin": "CN",
                    "locale": "zh-cn",
                    "forceExperimentsByIdNew": [],
                    "apo": False,
                    "searchCriteria": {"cityId": city_id},
                    "platform": {"id": 4},
                    "cid": "-1",
                    "storeFrontId": 3,
                    "occupancy": {
                        "numberOfAdults": adults,
                        "numberOfChildren": 0,
                        "travelerType": 1,
                        "checkIn": checkin_str
                    },
                    "deviceTypeId": 1,
                    "whiteLabelKey": "",
                    "correlationId": ""
                },
                "summary": {"includeHotelCharacter": False},
                "rateCategories": True,
                "contentRateCategories": {"escapeRateCategories": {}},
                "reviews": {
                    "demographics": {"filter": {"defaultProviderOnly": True}},
                    "summaries": {"apo": True, "limit": 1, "travellerType": 1},
                    "cumulative": {}
                },
                "images": {
                    "page": {"pageNumber": 1, "pageSize": 3},
                    "maxWidth": 0, "maxHeight": 0,
                    "imageSizes": [
                        {"key": "normal", "size": {"width": 167, "height": 85}},
                        {"key": "retina", "size": {"width": 334, "height": 170}}
                    ]
                },
                "nonHotelAccommodation": False,
                "engagement": True,
                "highlights": {"includeCollection": False}
            },
            "PricingSummaryRequest": {
                "cheapestOnly": True,
                "context": {
                    "clientInfo": {
                        "cid": -1, "languageId": 8, "origin": "CN",
                        "platform": 4, "storefront": 3
                    },
                    "isAllowBookOnRequest": True,
                    "sessionInfo": {"isLogin": False, "memberId": 0, "sessionId": 1}
                },
                "roomSortingStrategy": None,
                "isSSR": True,
                "pricing": {
                    "bookingDate": datetime.utcnow().strftime("%Y-%m-%dT05:00:00.000Z"),
                    "checkIn": checkin_str,
                    "checkout": checkout_str,
                    "currency": "CNY",
                    "details": {"cheapestPriceOnly": False, "itemBreakdown": False, "priceBreakdown": False},
                    "featureFlag": [
                        "ClientDiscount", "VipPlatinum", "VipDiamond", "Coupon",
                        "CreditCardPromotionPeek", "EnableCashback", "DispatchGoLocalForInternational",
                        "EnableGoToTravelCampaign", "EnableCofundedCashback", "EnableCashbackMildlyAggressiveDisplay",
                        "AutoApplyPromos", "EnableAgencySupplyForPackages", "MixAndSave",
                        "ReturnHotelNotReadyIfPullNotReady", "APSPeek", "PromosCumulative",
                        "DomesticTaxReceipt", "QuantumPaymentsEnabled"
                    ],
                    "features": {
                        "crossOutRate": False, "isAPSPeek": False, "isAllOcc": False,
                        "isApsEnabled": False, "isIncludeUsdAndLocalCurrency": False,
                        "isMSE": True, "isRPM2Included": True, "maxSuggestions": 0,
                        "newRateModel": False, "overrideOccupancy": False, "priusId": 0,
                        "synchronous": False, "filterCheapestRoomEscapesPackage": False,
                        "calculateRareRoomBadge": True, "enableRichContentOffer": True,
                        "enablePushDayUseRates": True, "returnCheapestEscapesOfferOnSSR": True,
                        "enableEscapesPackage": True, "disableEscapesPackage": False,
                        "isEnableSupplierFinancialInfo": False, "isLoggingAuctionData": False,
                        "enableRatePlanCheckInCheckOut": True, "enableSuggestPriceExclusiveWithFees": True
                    },
                    "filters": {
                        "cheapestRoomFilters": [], "filterAPO": False,
                        "ratePlans": [1], "secretDealOnly": False, "suppliers": []
                    },
                    "includedPriceInfo": False,
                    "localCheckInDate": checkin,
                    "localCheckoutDate": checkout,
                    "occupancy": {
                        "adults": adults, "children": 0, "rooms": rooms,
                        "childAges": [], "childrenTypes": []
                    },
                    "supplierPullMetadata": {"requiredPrecheckAccuracyLevel": 0},
                    "mseHotelIds": [], "mseClicked": "",
                    "bookingDurationType": None,
                    "ppLandingHotelIds": [], "searchedHotelIds": [], "paymentId": -1
                },
                "suggestedPrice": "NA"
            },
            "PriceStreamMetaLabRequest": {"attributesId": [2, 3, 1, 8, 6]}
        }
    }

    if price_from is not None or price_to is not None:
        rf = {"type": "Price", "currencyCode": "CNY"}
        if price_from is not None: rf["from"] = price_from
        if price_to is not None: rf["to"] = price_to
        body["variables"]["CitySearchRequest"]["searchRequest"]["filterRequest"]["rangeFilters"] = [rf]

    if page_token:
        body["variables"]["CitySearchRequest"]["searchRequest"]["page"]["pageToken"] = page_token

    return body


# ========== Hotel Data Extraction ==========

def extract_hotel(prop):
    """Extract hotel info from a GraphQL property object."""
    if not prop or not prop.get("content"):
        return None
    content = prop["content"]
    info = content.get("informationSummary") or {}
    reviews_data = content.get("reviews") or {}
    pricing = prop.get("pricing") or {}

    # Hotel name
    name = info.get("localeName") or info.get("displayName") or info.get("defaultName") or ""
    if not name:
        return None

    # Address
    addr = info.get("address") or {}
    area_name = addr.get("area", {}).get("name", "") if addr.get("area") else ""
    city_name = addr.get("city", {}).get("name", "") if addr.get("city") else ""
    address = f"{area_name} {city_name}".strip()

    # Rating
    rating = None
    try:
        cum = reviews_data.get("cumulative")
        if cum and isinstance(cum, dict):
            rating = cum.get("score")
    except:
        pass

    # Review count
    review_count = 0
    try:
        cum = reviews_data.get("cumulative")
        if cum and isinstance(cum, dict):
            review_count = cum.get("reviewCount", 0) or 0
    except:
        pass

    # Star rating (accommodationType maps to star level)
    stars = None
    try:
        acc_type = info.get("rating") or info.get("accommodationType")
        if acc_type and acc_type > 0:
            stars = acc_type
    except:
        pass

    # Price from cheapest room
    price = None
    try:
        offers = pricing.get("offers") or []
        for offer in offers:
            room_offers = offer.get("roomOffers") or []
            for ro in room_offers:
                room = ro.get("room") or {}
                pricing_list = room.get("pricing") or []
                for pr in pricing_list:
                    per_room = pr.get("price", {}).get("perRoomPerNight", {})
                    inclusive = per_room.get("inclusive", {})
                    display = inclusive.get("display")
                    if display:
                        price = round(float(display))
                        break
                    exclusive = per_room.get("exclusive", {})
                    display = exclusive.get("display")
                    if display:
                        price = round(float(display))
                        break
                if price: break
            if price: break

        # Fallback: perBook price
        if not price:
            for offer in offers:
                room_offers = offer.get("roomOffers") or []
                for ro in room_offers:
                    room = ro.get("room") or {}
                    pricing_list = room.get("pricing") or []
                    for pr in pricing_list:
                        per_book = pr.get("price", {}).get("perBook", {})
                        inclusive = per_book.get("inclusive", {})
                        display = inclusive.get("display")
                        if display:
                            price = round(float(display))
                            break
                    if price: break
                if price: break
    except:
        pass

    # Property URL
    property_links = info.get("propertyLinks") or {}
    property_page = property_links.get("propertyPage", "")
    hotel_url = ""
    if property_page:
        hotel_url = f"https://www.agoda.cn{property_page}" if property_page.startswith("/") else property_page

    return {
        "name": name,
        "address": address,
        "rating": rating,
        "review_count": review_count,
        "stars": stars,
        "price": price,
        "url": hotel_url
    }


# ========== Scrape a Single Page ==========

def scrape_page(session, headers, cookies, city_id, payload_kwargs):
    """Scrape a single page from the GraphQL API."""
    body = build_payload(**payload_kwargs)
    try:
        resp = session.post(GRAPHQL_URL, json=body, headers=headers, cookies=cookies, timeout=60)
        if resp.status_code != 200:
            print(f"  HTTP {resp.status_code}")
            return [], None, None

        data = resp.json()
        if "errors" in data:
            print(f"  GraphQL Error: {data['errors'][0].get('message', '')[:100]}")
            return [], None, None

        city_search = data.get("data", {}).get("citySearch", {})
        properties = city_search.get("properties") or []
        page_token = city_search.get("searchEnrichment", {}).get("pageToken")
        search_info = city_search.get("searchResult", {}).get("searchInfo", {})
        total = search_info.get("totalActiveHotels", 0)

        return properties, page_token, total
    except Exception as e:
        print(f"  Exception: {e}")
        return [], None, None


# ========== Scrape a Price Bracket ==========

def scrape_bracket(session, headers, cookies, city_id,
                   from_price=None, to_price=None, page_size=50):
    """Scrape all hotels in a price bracket using pagination."""
    all_hotels = []
    seen_names = set()
    page_token = None
    page = 1
    max_pages = 50

    label = f"¥{from_price or 0}-{to_price or 'max'}"
    print(f"  [{label}] ", end="", flush=True)

    while page <= max_pages:
        props, next_token, total = scrape_page(session, headers, cookies, city_id, {
            "city_id": city_id,
            "checkin": CHECKIN,
            "checkout": CHECKOUT,
            "los": 1,
            "rooms": 1,
            "adults": 2,
            "currency": "CNY",
            "page_size": page_size,
            "page_number": page,
            "page_token": page_token,
            "price_from": from_price,
            "price_to": to_price
        })

        if not props:
            if page > 1:
                break
            page += 1
            continue

        new_count = 0
        for p in props:
            if p.get("propertyResultType") != "NormalProperty":
                continue
            h = extract_hotel(p)
            if h and h["name"] and h["name"] not in seen_names:
                seen_names.add(h["name"])
                all_hotels.append(h)
                new_count += 1

        if new_count > 0:
            print(f"+{new_count}", end=" ", flush=True)

        if not next_token:
            break

        page_token = next_token
        page += 1
        time.sleep(0.5)  # Rate limiting

    # If we got no results via pagination, wait and try simple page 1 again
    if not all_hotels:
        time.sleep(2)
        props, _, _ = scrape_page(session, headers, cookies, city_id, {
            "city_id": city_id,
            "checkin": CHECKIN,
            "checkout": CHECKOUT,
            "los": 1,
            "rooms": 1,
            "adults": 2,
            "currency": "CNY",
            "page_size": page_size,
            "page_number": 1,
            "page_token": None,
            "price_from": from_price,
            "price_to": to_price
        })
        for p in (props or []):
            if p.get("propertyResultType") != "NormalProperty":
                continue
            h = extract_hotel(p)
            if h and h["name"] and h["name"] not in seen_names:
                seen_names.add(h["name"])
                all_hotels.append(h)
                new_count += 1
        if new_count > 0:
            print(f"+{new_count}", end=" ", flush=True)

    print(f" => {len(all_hotels)}家")
    return all_hotels


# ========== Save to DB & Excel ==========

def save_results(hotels, city):
    """Save hotel data to SQLite and Excel."""
    if not hotels:
        print("[错误] 无数据可保存")
        return

    print(f"\n>> 保存 {len(hotels)} 家酒店数据 ...")

    # ---- Deduplicate ----
    seen = set()
    unique = []
    for h in hotels:
        if h["name"] not in seen:
            seen.add(h["name"])
            unique.append(h)
    hotels = unique
    print(f">> 去重后: {len(hotels)} 家")

    # ---- SQLite ----
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT,
        address TEXT,
        user_rating REAL,
        review_count INTEGER,
        stars INTEGER,
        min_price REAL,
        url TEXT,
        scrape_date TEXT,
        city TEXT
    )""")
    c.execute("DELETE FROM hotels WHERE city=?", (city,))
    for h in hotels:
        c.execute("""INSERT INTO hotels
            (hotel_name, address, user_rating, review_count, stars, min_price, url, scrape_date, city)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (h["name"], h["address"], h["rating"], h["review_count"],
             h["stars"], h["price"], h["url"],
             datetime.now().strftime("%Y-%m-%d"), city))
    conn.commit()
    conn.close()
    print(f"  DB: {DB_PATH} ({len(hotels)} 条)")

    # ---- Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = f"{city}酒店列表"
    headers = ["序号", "酒店名称", "地址", "用户评分", "评价数", "星级", "最低价(CNY)", "详情链接", "抓取日期"]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    for i, h_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h_text)
        cell.font, cell.fill, cell.alignment = hfont, hfill, halign
        cell.border = thin_border

    # Sort by price ascending
    sorted_h = sorted(hotels, key=lambda x: (x["price"] or 99999))

    for idx, h in enumerate(sorted_h, 1):
        r = idx + 1
        ws.cell(row=r, column=1, value=idx).border = thin_border
        ws.cell(row=r, column=2, value=h["name"]).border = thin_border
        ws.cell(row=r, column=3, value=h["address"]).border = thin_border

        score_cell = ws.cell(row=r, column=4, value=h["rating"])
        score_cell.border = thin_border
        if h["rating"] and h["rating"] >= 9:
            score_cell.font = Font(color="FF0000", bold=True)

        ws.cell(row=r, column=5, value=h["review_count"]).border = thin_border

        star_cell = ws.cell(row=r, column=6, value=h["stars"])
        star_cell.border = thin_border

        price_cell = ws.cell(row=r, column=7, value=h["price"])
        price_cell.border = thin_border

        ws.cell(row=r, column=8, value=h["url"]).border = thin_border
        ws.cell(row=r, column=9, value=datetime.now().strftime("%Y-%m-%d")).border = thin_border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 55
    ws.column_dimensions['I'].width = 12
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    print(f"  Excel: {EXCEL_PATH}")

    # Stats
    prices = [h["price"] for h in hotels if h["price"] is not None]
    ratings = [h["rating"] for h in hotels if h["rating"] is not None]
    if prices:
        print(f"  价格: ¥{min(prices):.0f} ~ ¥{max(prices):.0f}, 均价 ¥{sum(prices)/len(prices):.0f}")
    if ratings:
        print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f}, >=9分: {sum(1 for r in ratings if r >= 9)}家")

    return EXCEL_PATH


# ========== Main ==========

def main():
    print("=" * 55)
    print(f"  Agoda {CITY} 酒店数据抓取")
    print(f"  入住: {CHECKIN}  退房: {CHECKOUT}  (1晚)")
    print(f"  1间, 2成人, CNY")
    print("=" * 55)

    # Step 1: Get fresh cookies via Selenium
    print("\n[Step 1/4] 获取 Cookie ...")
    cookies = None
    headers = None
    driver = None

    try:
        cookies, headers, driver = get_fresh_cookies_and_headers()
    except Exception as e:
        print(f"  Cookie获取失败: {e}")
        sys.exit(1)

    if not cookies or not headers:
        print("[错误] Cookie 或 Headers 为空")
        if driver: driver.quit()
        sys.exit(1)

    # Try city IDs in order (16958 is destination id, 5818 is alternate)
    city_ids = [16958, 5818]

    import requests

    # Step 2: Test API connection
    print("\n[Step 2/4] 测试 API 连接 ...")
    session = requests.Session()
    api_ok = False
    active_city_id = None

    for cid in city_ids:
        props, token, total = scrape_page(session, headers, cookies, cid, {
            "city_id": cid,
            "checkin": CHECKIN,
            "checkout": CHECKOUT,
            "los": 1,
            "rooms": 1,
            "adults": 2,
            "currency": "CNY",
            "page_size": 10,
            "page_number": 1,
            "page_token": None,
            "price_from": None,
            "price_to": None
        })
        if props:
            api_ok = True
            active_city_id = cid
            print(f"  API连接成功! cityId={cid}, 总酒店数: {total}, 本页: {len(props)}")
            # Show a sample
            sample = extract_hotel(props[0])
            if sample:
                print(f"  样例: {sample['name']} | ¥{sample['price']} | 评分{sample['rating']}")
            break
        else:
            print(f"  cityId={cid} 未返回数据, 尝试下一个...")

    if not api_ok:
        print("[错误] API 连接失败, 尝试直接 Selenium 方式...")
        # Fallback: use Selenium to scrape directly
        if driver:
            print(">> 使用 Selenium 直接抓取页面数据...")
            fallback_selenium(driver)
        else:
            print("[错误] 无可用抓取方式")
        return

    # Close browser now that API is working
    if driver:
        print(">> 关闭浏览器...")
        driver.quit()
        driver = None

    # Step 3: Scrape all hotels using price brackets
    print(f"\n[Step 3/4] 爬取全部酒店 (cityId={active_city_id}) ...")

    # Price brackets to get around API result limits
    brackets = [
        (None, 30), (30, 50), (50, 80), (80, 120), (120, 180),
        (180, 260), (260, 380), (380, 550), (550, 800),
        (800, 1200), (1200, 1800), (1800, 3000), (3000, 5000),
        (5000, 8000), (8000, 15000), (15000, 50000)
    ]

    all_hotels = []
    global_seen = set()
    total_brackets = len(brackets)

    for idx, (frm, to) in enumerate(brackets, 1):
        print(f"\n[{idx}/{total_brackets}]", end=" ")
        hotels = scrape_bracket(session, headers, cookies, active_city_id,
                                from_price=frm, to_price=to)
        new_count = 0
        for h in hotels:
            if h["name"] not in global_seen:
                global_seen.add(h["name"])
                all_hotels.append(h)
                new_count += 1
        print(f"  => 新增{new_count}家 (累计{len(all_hotels)}家)")
        time.sleep(1.0)  # Polite delay between brackets

    # Try a broader bracket to catch outliers
    print(f"\n[{total_brackets+1}/{total_brackets}]", end=" ")
    extra = scrape_bracket(session, headers, cookies, active_city_id,
                           from_price=None, to_price=None)
    new_count = 0
    for h in extra:
        if h["name"] not in global_seen:
            global_seen.add(h["name"])
            all_hotels.append(h)
            new_count += 1
    print(f"  => 新增{new_count}家 (累计{len(all_hotels)}家)")

    print(f"\n{'=' * 55}")
    print(f"  总计 {len(all_hotels)} 家酒店")
    print(f"{'=' * 55}")

    # Step 4: Save results
    print(f"\n[Step 4/4] 保存数据 ...")
    if all_hotels:
        excel_path = save_results(all_hotels, CITY)
        print(f"\n{'=' * 55}")
        print(f"  [完成] 共抓取 {len(all_hotels)} 家酒店")
        if excel_path:
            print(f"  文件: {excel_path}")
        print(f"{'=' * 55}")
    else:
        print("[错误] 未抓取到任何酒店数据")
        # Save debug page
        if driver:
            with open("page_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(">> 页面已保存到 page_debug.html")


def fallback_selenium(driver):
    """Fallback: Use Selenium to scrape hotels directly from search page."""
    print(">> 使用 Selenium 直接提取 ...")

    extract_js = """
    return (() => {
        const cards = document.querySelectorAll('[class*=PropertyCard]');
        const results = [];
        cards.forEach(card => {
            try {
                const h3 = card.querySelector('h3');
                const name = h3 ? h3.textContent.trim() : '';
                if (!name) return;

                let rating = null;
                card.querySelectorAll('[class*=review] [class*=score], [class*=Score]').forEach(el => {
                    const m = el.textContent.trim().match(/(\\d+\\.?\\d*)/);
                    if (m) { const v = parseFloat(m[1]); if (v > 0 && v <= 10) rating = v; }
                });

                let reviewCount = 0;
                card.querySelectorAll('span').forEach(el => {
                    const m = el.textContent.trim().match(/(\\d+)\\s*[条则篇]/);
                    if (m) reviewCount = parseInt(m[1]);
                });

                let price = null;
                card.querySelectorAll('*').forEach(el => {
                    const m = el.textContent.trim().match(/[¥￥]\\s*([\\d,]+\\.?\\d*)/);
                    if (m) {
                        const v = parseFloat(m[1].replace(/,/g, ''));
                        if (v >= 10 && v <= 100000) price = v;
                    }
                });

                let stars = 0;
                const starEls = card.querySelectorAll('[class*=star]:not([class*=outline]):not([class*=empty])');
                if (starEls.length > 0) stars = starEls.length;

                let url = '';
                const link = card.querySelector('a[href*="/hotel"], a[href*="/hotels"]');
                if (link) url = link.href;

                results.push({ name, rating, reviewCount, price, stars, url });
            } catch(e) {}
        });
        return JSON.stringify(results);
    })();
    """

    try:
        time.sleep(5)
        # Scroll to load more
        for i in range(10):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            cards = driver.find_elements("xpath", "//*[contains(@class,'PropertyCard')]")
            print(f"  滚动 {i+1}: {len(cards)} 个卡片")
            try:
                btns = driver.find_elements("xpath",
                    "//*[contains(text(),'显示更多') or contains(text(),'Show more')]")
                for btn in btns:
                    if btn.is_displayed():
                        driver.execute_script("arguments[0].click();", btn)
                        time.sleep(2)
            except:
                pass

        raw = driver.execute_script(extract_js)
        items = json.loads(raw) if isinstance(raw, str) else (raw or [])

        hotels = []
        for it in items:
            name = it.get("name", "")
            has_chinese = any('一' <= c <= '鿿' for c in name)
            if name and (has_chinese or len(name) > 3):
                hotels.append(it)

        if not hotels:
            hotels = [it for it in items if it.get('name')]

        if hotels:
            save_results([{
                "name": h.get("name",""),
                "address": "",
                "rating": h.get("rating"),
                "review_count": h.get("reviewCount", 0),
                "stars": h.get("stars", 0),
                "price": h.get("price"),
                "url": h.get("url", "")
            } for h in hotels], CITY)
        else:
            print("[错误] Selenium 也未提取到酒店数据")
            with open("page_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(">> 页面已保存到 page_debug.html")

    except Exception as e:
        print(f"[错误] Selenium 提取失败: {e}")
        import traceback; traceback.print_exc()


if __name__ == "__main__":
    main()
