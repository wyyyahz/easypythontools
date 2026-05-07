#!/usr/bin/env python3
"""
Capture Agoda's actual GraphQL request/response via CDP,
then use the real API payload to scrape ALL hotels.
"""
import json, os, time, sqlite3, sys
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

BASE = os.path.dirname(os.path.abspath(__file__))
CITY = "武汉"
CHECKIN = (datetime.now() + timedelta(days=35)).strftime("%Y-%m-%d")
CHECKOUT = (datetime.now() + timedelta(days=36)).strftime("%Y-%m-%d")
DB_PATH = os.path.join(BASE, "hotels.db")
EXCEL_PATH = os.path.join(BASE, f"hotels_{CITY}.xlsx")

def start_driver():
    opts = Options()
    opts.binary_location = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--lang=zh-CN")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    opts.set_capability("goog:loggingPrefs", {"performance": "ALL"})
    chromedriver = os.path.join(BASE, "chromedriver.exe")
    driver = webdriver.Chrome(service=Service(chromedriver), options=opts)
    return driver

def capture_graphql_response(driver):
    """Wait for search page to load and capture graphql response."""
    url = (f"https://www.agoda.cn/search?city=16958"
           f"&checkIn={CHECKIN}&checkOut={CHECKOUT}"
           f"&rooms=1&adults=2&currency=CNY")
    print(f">> 打开搜索页...")
    driver.get(url)

    # Wait for search results to load
    time.sleep(8)

    # Wait for PropertyCards to appear
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//*[contains(@class,'PropertyCard')]")))
        print(">> PropertyCards 已加载")
    except:
        print(">> PropertyCards 超时")

    time.sleep(3)

    # Get performance logs from Chrome
    logs = driver.get_log("performance")
    print(f">> 性能日志条目: {len(logs)}")

    # Extract graphql response bodies from network events
    # Chrome performance log format: each entry has a message with params
    responses = {}
    for entry in logs:
        try:
            msg = json.loads(entry["message"])
            params = msg.get("message", {}).get("params", {})

            # Capture response
            if (params.get("type") == "Network" and
                "response" in params and
                "graphql" in str(params.get("response", {}).get("url", "")).lower()):
                req_id = params.get("requestId")
                resp = params["response"]
                responses[req_id] = {
                    "url": resp.get("url"),
                    "status": resp.get("status"),
                    "requestId": req_id
                }
        except:
            pass

    # Try to get response bodies via CDP
    raw_responses = {}
    for req_id, info in responses.items():
        if info.get("status") == 200:
            try:
                body = driver.execute_cdp_cmd("Network.getResponseBody",
                    {"requestId": req_id})
                raw_responses[req_id] = json.loads(body["body"])
                print(f">> 捕获到 GraphQL 响应: {len(raw_responses[req_id])} 字节")
            except Exception as e:
                print(f">> 获取响应体失败: {e}")

    return raw_responses, driver

def extract_cookies(driver):
    """Extract cookies from the browser session."""
    return {c["name"]: c["value"] for c in driver.get_cookies()}

def parse_hotel_data(responses):
    """Parse hotel data from GraphQL responses."""
    hotels = []
    for req_id, data in responses.items():
        try:
            props = (data.get("data", {})
                     .get("citySearch", {})
                     .get("properties", []))
            if props:
                print(f">> 解析到 {len(props)} 个酒店属性")
                for prop in props:
                    h = extract_hotel(prop)
                    if h:
                        hotels.append(h)
        except Exception as e:
            print(f">> 解析错误: {e}")
    return hotels

def extract_hotel(prop):
    """Extract hotel from a property object."""
    if not prop or not prop.get("content"):
        return None
    info = prop["content"].get("informationSummary") or {}
    reviews = prop["content"].get("reviews") or {}
    pricing = prop.get("pricing") or {}

    name = info.get("localeName") or info.get("defaultName") or ""
    if not name:
        return None

    # Price
    price = None
    try:
        for offer in (pricing.get("offers") or []):
            for ro in (offer.get("roomOffers") or []):
                for pr in (ro.get("room", {}).get("pricing") or []):
                    d = pr.get("price", {}).get("perBook", {}).get("inclusive", {}).get("display")
                    if d:
                        price = round(d)
                        break
                if price: break
            if price: break
    except: pass

    # Rating
    rating = None
    try: rating = reviews.get("cumulative", {}).get("score")
    except: pass

    # Star rating from accommodationType
    stars = ""
    try:
        acc_type = info.get("accommodationType")
        star_map = {1: "1★", 2: "2★", 3: "3★", 4: "4★", 5: "5★"}
        stars = star_map.get(acc_type, "")
    except: pass

    # Location
    location = ""
    try:
        addr = info.get("address") or {}
        if addr.get("area"):
            location = addr["area"].get("name", "")
        if not location and addr.get("city"):
            location = addr["city"].get("name", "")
        if not location:
            location = addr.get("full", "")
    except: pass

    # Review count
    review_count = 0
    try: review_count = reviews.get("cumulative", {}).get("totalCount", 0)
    except: pass

    return {
        "name": name, "stars": stars, "rating": rating,
        "price": price, "location": location,
        "review_count": review_count
    }

def save_results(hotels):
    """Save to SQLite and Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    print(f"\n>> 保存 {len(hotels)} 家酒店...")

    # DB
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS hotels (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hotel_name TEXT, star_rating TEXT, user_rating REAL,
        review_count INTEGER, price_cny INTEGER, location TEXT,
        scrape_date TEXT, city TEXT
    )""")
    c.execute("DELETE FROM hotels WHERE city=?", (CITY,))
    for h in hotels:
        c.execute("""INSERT INTO hotels
            (hotel_name, star_rating, user_rating, review_count, price_cny, location, scrape_date, city)
            VALUES (?,?,?,?,?,?,?,?)""",
            (h["name"], h["stars"], h["rating"], h.get("review_count", 0),
             h["price"], h["location"],
             datetime.now().strftime("%Y-%m-%d"), CITY))
    conn.commit()
    conn.close()
    print(f"  DB: {DB_PATH} ({len(hotels)}条)")

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "武汉酒店列表"
    headers = ["序号", "酒店名称", "用户评分", "评价数", "最低价(CNY)", "星级", "区域位置", "抓取日期"]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font, cell.fill, cell.alignment = hfont, hfill, halign

    sorted_h = sorted(hotels, key=lambda x: (x['price'] or 99999))
    for idx, h in enumerate(sorted_h, 1):
        r = idx + 1
        ws.cell(row=r, column=1, value=idx).border = border
        ws.cell(row=r, column=2, value=h['name']).border = border

        score_cell = ws.cell(row=r, column=3, value=h['rating'])
        score_cell.border = border
        if h['rating'] and h['rating'] >= 9:
            score_cell.font = Font(color="FF0000", bold=True)

        ws.cell(row=r, column=4, value=h.get('review_count', 0)).border = border
        ws.cell(row=r, column=5, value=h['price']).border = border
        ws.cell(row=r, column=6, value=h['stars']).border = border
        ws.cell(row=r, column=7, value=h['location']).border = border
        ws.cell(row=r, column=8, value=datetime.now().strftime("%Y-%m-%d")).border = border

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 24
    ws.column_dimensions['H'].width = 12
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    print(f"  Excel: {EXCEL_PATH}")

    # Stats
    prices = [h['price'] for h in hotels if h['price'] is not None]
    ratings = [h['rating'] for h in hotels if h['rating'] is not None]
    if prices: print(f"  价格: ¥{min(prices)} ~ ¥{max(prices)}, 均价 ¥{sum(prices)/len(prices):.0f}")
    if ratings: print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f}")

def main():
    print("=" * 55)
    print(f"  Agoda {CITY} 酒店数据抓取 (CDP捕获)")
    print(f"  {CHECKIN} ~ {CHECKOUT}, 1间2成人1晚CNY")
    print("=" * 55)

    driver = None
    try:
        driver = start_driver()

        # Capture the GraphQL response from the browser
        responses, driver = capture_graphql_response(driver)

        if not responses:
            print(">> 未捕获到 GraphQL 响应，尝试直接从页面提取...")
            # Fallback: try to extract from page
            cards = driver.find_elements(By.XPATH,
                "//*[contains(@class,'PropertyCard')]")
            print(f">> 找到 {len(cards)} 个卡片元素")

            # Try to get data from the page's __INITIAL_STATE__ or similar
            page_data = driver.execute_script("""
                // Try to find Apollo cache or React data
                try {
                    let data = {};
                    // Check for __NEXT_DATA__ or similar
                    let scripts = document.querySelectorAll('script');
                    for (let s of scripts) {
                        let text = s.textContent || '';
                        if (text.includes('citySearch') || text.includes('properties')) {
                            return text.substring(0, 5000);
                        }
                    }
                    return 'not found';
                } catch(e) { return 'error: ' + e.message; }
            """)
            print(f">> 页面数据查找: {page_data[:500] if page_data else 'none'}")
        else:
            hotels = parse_hotel_data(responses)

            if not hotels:
                print(">> GraphQL 响应无酒店数据，输出响应结构...")
                for req_id, data in list(responses.items())[:1]:
                    s = json.dumps(data, ensure_ascii=False, indent=2)
                    print(s[:3000])
                return

        # Fallback: scrape the visible page
        if not hotels:
            print(">> 开始页面提取...")
            cards = driver.find_elements(By.XPATH,
                "//*[contains(@class,'PropertyCard')]")
            print(f">> 可见卡片: {len(cards)}")
            hotels = []
            for card in cards:
                try:
                    name_el = card.find_element(By.XPATH, ".//h3")
                    name = name_el.text.strip()
                    if name and all(ord(c) < 256 for c in name):
                        continue
                    if name and len(name) > 1:
                        hotels.append({"name": name, "stars": "", "rating": None,
                            "price": None, "location": "", "review_count": 0})
                        print(f"  {name[:30]}")
                except: pass

            if not hotels:
                print(">> h3 提取失败，尝试 JS 提取...")
                html = driver.execute_script("""
                    let items = [];
                    document.querySelectorAll('[class*=PropertyCard]').forEach(c => {
                        let name = c.querySelector('h3');
                        let price = c.querySelector('[class*=Price]');
                        let rating = c.querySelector('[class*=review]');
                        items.push({
                            name: name ? name.textContent.trim() : '',
                            price: price ? price.textContent.trim() : '',
                            rating: rating ? rating.textContent.trim() : ''
                        });
                    });
                    return JSON.stringify(items);
                """)
                items = json.loads(html)
                print(f">> JS 提取到 {len(items)} 个:")
                for it in items[:5]:
                    print(f"  {json.dumps(it, ensure_ascii=False)}")
                for it in items:
                    if it.get('name') and len(it['name']) > 1:
                        hotels.append({"name": it['name'], "stars": "", "rating": None,
                            "price": None, "location": "", "review_count": 0})

        if hotels:
            # Deduplicate
            seen = set()
            unique = []
            for h in hotels:
                if h['name'] not in seen:
                    seen.add(h['name'])
                    unique.append(h)
            print(f"\n>> 去重后 {len(unique)} 家酒店")
            save_results(unique)
        else:
            print("[错误] 未找到任何酒店数据")

    except Exception as e:
        print(f"[错误] {e}")
        import traceback; traceback.print_exc()
    finally:
        if driver:
            driver.quit()

if __name__ == "__main__":
    main()
