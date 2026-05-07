#!/usr/bin/env python3
"""
Agoda 武汉酒店数据抓取 - 纯JS提取版
"""
import json, os, time, sqlite3
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

BASE = os.path.dirname(os.path.abspath(__file__))
CITY = "武汉"
CHECKIN = (datetime.now() + timedelta(days=35)).strftime("%Y-%m-%d")
CHECKOUT = (datetime.now() + timedelta(days=36)).strftime("%Y-%m-%d")
DB_PATH = os.path.join(BASE, "hotels.db")
EXCEL_PATH = os.path.join(BASE, f"hotels_{CITY}.xlsx")

EXTRACT_JS = """
return (() => {
    const cards = document.querySelectorAll('[class*=PropertyCard]');
    const results = [];

    cards.forEach((card, idx) => {
        try {
            const h3 = card.querySelector('h3');
            const name = h3 ? h3.textContent.trim() : '';
            if (!name) return;

            let rating = null;
            const scoreEls = card.querySelectorAll('[class*=review] [class*=score], [class*=Score]');
            scoreEls.forEach(el => {
                const m = el.textContent.trim().match(/(\d+\.?\d*)/);
                if (m) {
                    const v = parseFloat(m[1]);
                    if (v > 0 && v <= 10) rating = v;
                }
            });

            let reviewCount = 0;
            const allSpans = card.querySelectorAll('span');
            allSpans.forEach(el => {
                const m = el.textContent.trim().match(/(\d+)\s*[条则篇]/);
                if (m) reviewCount = parseInt(m[1]);
            });

            let price = null;
            const allEls = card.querySelectorAll('*');
            allEls.forEach(el => {
                const m = el.textContent.trim().match(/[¥￥]\s*([\d,]+\.?\d*)/);
                if (m) {
                    const v = parseFloat(m[1].replace(/,/g, ''));
                    if (v >= 10 && v <= 100000) price = v;
                }
            });

            let stars = 0;
            const starEls = card.querySelectorAll('[class*=star]:not([class*=outline]):not([class*=empty])');
            if (starEls.length > 0) stars = starEls.length;

            let url = '';
            const link = card.querySelector('a[href*="/hotel"], a[href*="/hotels"]');
            if (link) url = link.href;

            results.push({
                name: name,
                rating: rating,
                reviewCount: reviewCount,
                price: price,
                stars: stars,
                url: url
            });
        } catch(e) {}
    });

    return JSON.stringify(results);
})();
"""

SCROLL_JS = """
() => {
    const maxScrolls = 50;
    let lastCount = 0;
    let noChangeCount = 0;

    return new Promise((resolve) => {
        const check = () => {
            window.scrollTo(0, document.body.scrollHeight);

            setTimeout(() => {
                // Click any "show more" buttons
                document.querySelectorAll('[class*=load-more], button:contains(显示更多), button:contains(Show more)')
                    .forEach(btn => { try { btn.click(); } catch(e) {} });

                const cards = document.querySelectorAll('[class*=PropertyCard]');
                const count = cards.length;

                if (count > lastCount) {
                    lastCount = count;
                    noChangeCount = 0;
                } else {
                    noChangeCount++;
                }

                if (noChangeCount >= 3) {
                    resolve(lastCount);
                } else {
                    setTimeout(check, 2000);
                }
            }, 2000);
        };
        check();
    });
}
"""

def run():
    print("=" * 55)
    print(f"  Agoda {CITY} 酒店抓取 (JS提取)")
    print(f"  {CHECKIN} ~ {CHECKOUT}")
    print("=" * 55)

    opts = Options()
    opts.binary_location = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--lang=zh-CN")
    opts.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")

    driver = webdriver.Chrome(
        service=Service(os.path.join(BASE, "chromedriver.exe")),
        options=opts)

    try:
        url = (f"https://www.agoda.cn/search?city=16958"
               f"&checkIn={CHECKIN}&checkOut={CHECKOUT}"
               f"&rooms=1&adults=2&currency=CNY")
        print(f">> 打开: {url}")
        driver.get(url)

        print(">> 等待页面加载...")
        time.sleep(10)

        # Wait for PropertyCards
        try:
            WebDriverWait(driver, 20).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, "//*[contains(@class,'PropertyCard')]")))
            print(">> PropertyCards 已出现")
        except:
            print(">> 等待 PropertyCards 超时")

        time.sleep(3)

        # Scroll to load more
        print(">> 滚动加载更多...")
        try:
            final_count = driver.execute_script(SCROLL_JS + "\nreturn " + SCROLL_JS.split("() =>")[1].strip() if False else SCROLL_JS)
            # Actually let's just scroll using simple JS
            for i in range(5):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
                cards = driver.find_elements(By.XPATH, "//*[contains(@class,'PropertyCard')]")
                print(f"   滚动 {i+1}: {len(cards)} 个")
        except:
            pass

        # Extract data using JavaScript
        print(">> JS提取酒店数据...")
        raw = driver.execute_script(EXTRACT_JS)
        items = json.loads(raw) if isinstance(raw, str) else raw

        print(f">> 提取到 {len(items)} 个结果")

        # Filter out non-Chinese hotels
        hotels = []
        for it in items:
            name = it.get("name", "")
            # Skip if name seems non-Chinese (hotels in Wuhan should have Chinese names)
            has_chinese = any('一' <= c <= '鿿' for c in name)
            if name and (has_chinese or len(name) > 3):
                hotels.append(it)
                price_str = f"¥{it['price']}" if it.get('price') else "?"
                print(f"  {it['name'][:30]:30s} {price_str:>8s}  评分:{it.get('rating','?')}")

        if not hotels:
            print(">> 过滤后无结果，使用全部结果")
            hotels = [it for it in items if it.get('name')]

        if not hotels:
            print("[错误] 未找到任何酒店")
            # Save debug info
            with open("page_debug.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(">> 页面已保存到 page_debug.html")
            return

        print(f"\n>> 共 {len(hotels)} 家酒店")

        # Save
        print(">> 保存数据...")
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS hotels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hotel_name TEXT, user_rating REAL, review_count INTEGER,
            min_price REAL, stars INTEGER, url TEXT, scrape_date TEXT, city TEXT
        )""")
        c.execute("DELETE FROM hotels WHERE city=?", (CITY,))
        for h in hotels:
            c.execute("""INSERT INTO hotels
                (hotel_name, user_rating, review_count, min_price, stars, url, scrape_date, city)
                VALUES (?,?,?,?,?,?,?,?)""",
                (h.get("name",""), h.get("rating"), h.get("reviewCount",0),
                 h.get("price"), h.get("stars",0), h.get("url",""),
                 datetime.now().strftime("%Y-%m-%d"), CITY))
        conn.commit()
        conn.close()

        # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "武汉酒店列表"
        headers = ["序号", "酒店名称", "用户评分", "评价数", "最低价(CNY)", "星级", "详情链接", "抓取日期"]
        hfont = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        halign = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font, cell.fill, cell.alignment = hfont, hfill, halign

        sorted_h = sorted(hotels, key=lambda x: (x.get("price") or 99999))
        seen = set()
        uniq = []
        for h in sorted_h:
            if h["name"] not in seen:
                seen.add(h["name"])
                uniq.append(h)

        for idx, h in enumerate(uniq, 1):
            r = idx + 1
            ws.cell(row=r, column=1, value=idx).border = border
            ws.cell(row=r, column=2, value=h["name"]).border = border
            sc = ws.cell(row=r, column=3, value=h.get("rating"))
            sc.border = border
            if h.get("rating") and h["rating"] >= 9: sc.font = Font(color="FF0000", bold=True)
            ws.cell(row=r, column=4, value=h.get("reviewCount", 0)).border = border
            ws.cell(row=r, column=5, value=h.get("price")).border = border
            ws.cell(row=r, column=6, value=h.get("stars", 0)).border = border
            ws.cell(row=r, column=7, value=h.get("url", "")).border = border
            ws.cell(row=r, column=8, value=datetime.now().strftime("%Y-%m-%d")).border = border

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 50
        ws.column_dimensions['H'].width = 12
        ws.freeze_panes = "A2"
        wb.save(EXCEL_PATH)

        print(f"  DB: {DB_PATH}")
        print(f"  Excel: {EXCEL_PATH}")

        prices = [h["price"] for h in hotels if h.get("price")]
        ratings = [h["rating"] for h in hotels if h.get("rating")]
        if prices: print(f"  价格: ¥{min(prices):.0f} ~ ¥{max(prices):.0f}, 均价 ¥{sum(prices)/len(prices):.0f}")
        if ratings: print(f"  评分: {min(ratings):.1f} ~ {max(ratings):.1f}")

    except Exception as e:
        print(f"[错误] {e}")
        import traceback; traceback.print_exc()
    finally:
        driver.quit()
        print(">> 完成!")

if __name__ == "__main__":
    run()
